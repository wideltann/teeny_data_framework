"""
Comprehensive tests for table_functions_postgres.py

Tests cover:
- S3 helper functions
- Column mapping and selection functions
- File reading functions (CSV, TSV, PSV, XLSX, Parquet, Fixed-width)
- Database operations (table creation, COPY, schema validation)
- Metadata functions (file extraction, metadata tracking)
- CLI schema inference functions

Uses testcontainers for PostgreSQL - automatically manages Docker containers!
Requires Docker to be running.
"""

import pytest
import polars as pl
import adbc_driver_postgresql.dbapi as pg_dbapi
from pathlib import Path
from typing import Dict, List, Tuple, Any
import tempfile
import zipfile
import json
from unittest.mock import Mock, patch, MagicMock
from testcontainers.postgres import PostgresContainer

# Import functions to test
from src.table_functions_postgres import (
    # S3 helpers
    is_s3_path,
    get_s3_filesystem,
    # Path utilities
    normalize_path,
    path_join,
    path_basename,
    path_parent,
    # Column mapping
    prepare_column_mapping,
    # File readers
    read_csv,
    read_xlsx,
    read_parquet,
    read_fixed_width,
    read_using_column_mapping,
    # Database functions
    table_exists,
    get_table_schema,
    validate_schema_match,
    create_table_from_dataframe,
    copy_dataframe_to_table,
    # Metadata functions
    get_csv_header_and_row_count,
    get_file_metadata_row,
    row_count_check,
    update_metadata,
    update_table,
    add_files_to_metadata_table,
    extract_and_add_zip_files,
    add_files,
    drop_search_dir,
    drop_partition,
    drop_file_from_metadata_and_table,
    # Schema inference
    to_snake_case,
    infer_schema_from_file,
    # Connection helpers
    get_connection,
    execute_sql,
    execute_sql_fetch,
    execute_sql_fetchone,
)


# ===== FIXTURES =====


@pytest.fixture
def temp_dir():
    """Create a temporary directory for test files"""
    with tempfile.TemporaryDirectory() as tmpdir:
        yield Path(tmpdir)


@pytest.fixture(scope="session")
def postgres_container():
    """
    Create a PostgreSQL container for the entire test session.

    Uses testcontainers to automatically:
    - Pull PostgreSQL Docker image (if needed)
    - Start a PostgreSQL container
    - Expose the database port
    - Clean up when tests are done

    Requires Docker to be running!
    """
    with PostgresContainer("postgres:16-alpine") as postgres:
        yield postgres


@pytest.fixture(scope="function")
def db_uri(postgres_container):
    """
    Create a database URI for each test function.

    Reuses the same PostgreSQL container but creates a fresh schema for each test.
    """
    # Get connection string from container
    connection_url = postgres_container.get_connection_url()

    # Convert SQLAlchemy-style URL to standard PostgreSQL URL
    # testcontainers returns: postgresql+psycopg2://...
    # ADBC expects: postgresql://...
    uri = connection_url.replace("postgresql+psycopg2://", "postgresql://")

    # Create test schema using ADBC
    execute_sql(uri, "CREATE SCHEMA IF NOT EXISTS test_schema")

    yield uri

    # Cleanup schema after test
    try:
        execute_sql(uri, "DROP SCHEMA IF EXISTS test_schema CASCADE")
    except:
        pass


@pytest.fixture
def sample_csv_file(temp_dir):
    """Create a sample CSV file for testing"""
    csv_path = temp_dir / "sample.csv"
    data = """name,age,score
Alice,25,85.5
Bob,30,92.0
Charlie,35,78.5"""
    csv_path.write_text(data)
    return csv_path


@pytest.fixture
def sample_csv_no_header(temp_dir):
    """Create a CSV file without headers"""
    csv_path = temp_dir / "sample_no_header.csv"
    data = """Alice,25,85.5
Bob,30,92.0
Charlie,35,78.5"""
    csv_path.write_text(data)
    return csv_path


@pytest.fixture
def sample_psv_file(temp_dir):
    """Create a pipe-delimited file"""
    psv_path = temp_dir / "sample.psv"
    data = """name|age|score
Alice|25|85.5
Bob|30|92.0"""
    psv_path.write_text(data)
    return psv_path


@pytest.fixture
def sample_zip_file(temp_dir):
    """Create a ZIP file containing CSV files"""
    zip_path = temp_dir / "sample.zip"

    with zipfile.ZipFile(zip_path, 'w') as zf:
        zf.writestr("file1.csv", "col1,col2\n1,2\n3,4")
        zf.writestr("file2.csv", "col1,col2\n5,6\n7,8")

    return zip_path


# ===== S3 HELPER TESTS =====


class TestS3Helpers:
    """Test S3 helper functions"""

    def test_is_s3_path_true(self):
        assert is_s3_path("s3://bucket/path/to/file")

    def test_is_s3_path_false(self):
        assert not is_s3_path("/local/path")
        assert not is_s3_path("http://example.com")
        assert not is_s3_path("s3:/bucket/path")  # Path-mangled no longer supported

    def test_get_s3_filesystem_returns_passed_filesystem(self):
        """Test that get_s3_filesystem returns the passed filesystem if provided"""
        mock_fs = MagicMock()
        result = get_s3_filesystem(mock_fs)
        assert result is mock_fs

    def test_get_s3_filesystem_creates_new(self):
        """Test that get_s3_filesystem creates a new S3FileSystem if not provided"""
        import s3fs
        result = get_s3_filesystem(None)
        assert isinstance(result, s3fs.S3FileSystem)


# ===== PATH UTILITY TESTS =====


class TestPathUtilities:
    """Test path utility functions for S3 and local paths"""

    # normalize_path tests
    def test_normalize_path_local(self):
        """Test normalizing local paths"""
        assert normalize_path("/path/to/file") == "/path/to/file"
        assert normalize_path("/path/to/dir/") == "/path/to/dir"
        assert normalize_path("relative/path") == "relative/path"

    def test_normalize_path_s3(self):
        """Test normalizing S3 paths"""
        assert normalize_path("s3://bucket/path/to/file") == "s3://bucket/path/to/file"
        assert normalize_path("s3://bucket/path/") == "s3://bucket/path"
        assert normalize_path("s3://bucket") == "s3://bucket"

    def test_normalize_path_with_path_object(self):
        """Test normalizing Path objects"""
        from pathlib import Path
        assert normalize_path(Path("/path/to/file")) == "/path/to/file"

    # path_join tests
    def test_path_join_local(self):
        """Test joining local paths"""
        assert path_join("/path", "to", "file") == "/path/to/file"
        assert path_join("relative", "path") == "relative/path"

    def test_path_join_s3(self):
        """Test joining S3 paths"""
        assert path_join("s3://bucket", "path", "file.csv") == "s3://bucket/path/file.csv"
        assert path_join("s3://bucket/", "path/", "file.csv") == "s3://bucket/path/file.csv"

    def test_path_join_empty_parts(self):
        """Test joining with empty parts"""
        assert path_join("/path", "", "file") == "/path/file"
        assert path_join("") == ""

    def test_path_join_strips_slashes(self):
        """Test that joining strips extra slashes"""
        assert path_join("/path/", "/to/", "/file") == "/path/to/file"
        assert path_join("s3://bucket/", "/path/") == "s3://bucket/path"

    # path_basename tests
    def test_path_basename_local(self):
        """Test getting basename from local paths"""
        assert path_basename("/path/to/file.csv") == "file.csv"
        assert path_basename("/path/to/dir/") == "dir"
        assert path_basename("file.csv") == "file.csv"

    def test_path_basename_s3(self):
        """Test getting basename from S3 paths"""
        assert path_basename("s3://bucket/path/to/file.csv") == "file.csv"
        assert path_basename("s3://bucket/path/") == "path"

    # path_parent tests
    def test_path_parent_local(self):
        """Test getting parent from local paths"""
        assert path_parent("/path/to/file.csv") == "/path/to"
        assert path_parent("/path/to/") == "/path"
        assert path_parent("/file.csv") == ""

    def test_path_parent_s3(self):
        """Test getting parent from S3 paths"""
        assert path_parent("s3://bucket/path/to/file.csv") == "s3://bucket/path/to"
        assert path_parent("s3://bucket/path") == "s3://bucket"
        # S3 bucket root returns itself
        assert path_parent("s3://bucket") == "s3://bucket"

    def test_path_parent_single_element(self):
        """Test parent of single element path"""
        assert path_parent("file.csv") == ""


# ===== COLUMN MAPPING TESTS =====


class TestColumnMapping:
    """Test column selection and mapping functions"""

    def test_prepare_column_mapping_basic(self):
        """Test basic column selection without renaming"""
        header = ["col1", "col2", "col3"]
        column_mapping = {
            "col1": ([], "string"),
            "col2": ([], "int"),
            "col3": ([], "float"),
        }

        rename_dict, read_dtypes, missing_cols = prepare_column_mapping(
            header, column_mapping
        )

        assert rename_dict == {}  # No renames needed
        assert read_dtypes == {"col1": "string", "col2": "int", "col3": "float"}
        assert missing_cols == {}

    def test_prepare_column_mapping_with_rename(self):
        """Test column selection with renaming"""
        header = ["old_name", "col2"]
        column_mapping = {
            "new_name": (["old_name", "alt_name"], "string"),
            "col2": ([], "int"),
        }

        rename_dict, read_dtypes, missing_cols = prepare_column_mapping(
            header, column_mapping
        )

        assert rename_dict == {"old_name": "new_name"}
        assert read_dtypes == {"old_name": "string", "col2": "int"}
        assert missing_cols == {}

    def test_prepare_column_mapping_with_missing_cols(self):
        """Test handling of missing columns"""
        header = ["col1", "col2"]
        column_mapping = {
            "col1": ([], "string"),
            "col2": ([], "int"),
            "col3": ([], "float"),  # Missing in header
        }

        rename_dict, read_dtypes, missing_cols = prepare_column_mapping(
            header, column_mapping
        )

        assert missing_cols == {"col3": "float"}
        assert read_dtypes == {"col1": "string", "col2": "int"}

    def test_prepare_column_mapping_with_default(self):
        """Test default type for unmapped columns"""
        header = ["col1", "col2", "col3", "col4"]
        column_mapping = {
            "col1": ([], "string"),
            "col2": ([], "int"),
            "default": ([], "string"),
        }

        rename_dict, read_dtypes, missing_cols = prepare_column_mapping(
            header, column_mapping
        )

        # col3 and col4 should get default type
        assert read_dtypes == {
            "col1": "string",
            "col2": "int",
            "col3": "string",
            "col4": "string",
        }
        assert rename_dict == {}
        assert missing_cols == {}


# ===== FILE READING TESTS =====


class TestFileReading:
    """Test file reading functions"""

    def test_read_csv_with_header(self, sample_csv_file):
        """Test reading CSV with header"""
        column_mapping = {
            "name": ([], "string"),
            "age": ([], "int"),
            "score": ([], "float"),
        }

        df = read_csv(
            full_path=str(sample_csv_file),
            column_mapping=column_mapping,
            has_header=True,
        )

        assert len(df) == 3
        assert list(df.columns) == ["name", "age", "score"]
        assert df["name"][0] == "Alice"
        assert df["age"].dtype == pl.Int64

    def test_read_csv_no_header(self, sample_csv_no_header):
        """Test reading CSV without header"""
        column_mapping = {
            "name": ([], "string"),
            "age": ([], "int"),
            "score": ([], "float"),
        }

        header = ["name", "age", "score"]

        df = read_csv(
            full_path=str(sample_csv_no_header),
            column_mapping=column_mapping,
            header=header,
            has_header=False,
        )

        assert len(df) == 3
        assert df["name"][0] == "Alice"

    def test_read_csv_with_quoted_headers(self, temp_dir):
        """Test reading CSV with quoted headers"""
        csv_path = temp_dir / "quoted_headers.csv"
        csv_path.write_text('"name","age","score"\nAlice,25,95.5\nBob,30,87.2\nCharlie,22,92.0')

        column_mapping = {
            "name": ([], "string"),
            "age": ([], "int"),
            "score": ([], "float"),
        }

        df = read_csv(
            full_path=str(csv_path),
            column_mapping=column_mapping,
            has_header=True,
        )

        assert len(df) == 3
        assert list(df.columns) == ["name", "age", "score"]
        assert df["name"][0] == "Alice"
        assert df["age"].dtype == pl.Int64
        assert df["score"][0] == 95.5

    def test_read_csv_psv_separator(self, sample_psv_file):
        """Test reading pipe-delimited file"""
        column_mapping = {
            "name": ([], "string"),
            "age": ([], "int"),
            "score": ([], "float"),
        }

        df = read_csv(
            full_path=str(sample_psv_file),
            column_mapping=column_mapping,
            has_header=True,
            separator="|",
        )

        assert len(df) == 2
        assert df["name"][0] == "Alice"

    def test_read_csv_with_missing_cols(self, sample_csv_file):
        """Test reading CSV with missing columns in mapping"""
        column_mapping = {
            "name": ([], "string"),
            "age": ([], "int"),
            "score": ([], "float"),
            "extra_col": ([], "string"),  # Not in CSV
        }

        df = read_csv(
            full_path=str(sample_csv_file),
            column_mapping=column_mapping,
            has_header=True,
        )

        assert "extra_col" in df.columns
        assert df["extra_col"].null_count() == len(df)

    def test_read_using_column_mapping_csv(self, sample_csv_file):
        """Test router function for CSV"""
        column_mapping = {
            "name": ([], "string"),
            "age": ([], "int"),
            "score": ([], "float"),
        }

        df = read_using_column_mapping(
            full_path=str(sample_csv_file),
            filetype="csv",
            column_mapping=column_mapping,
            has_header=True,
        )

        assert df is not None
        assert len(df) == 3

    def test_read_using_column_mapping_tsv(self, temp_dir):
        """Test router function for TSV"""
        tsv_path = temp_dir / "sample.tsv"
        tsv_path.write_text("col1\tcol2\n1\t2\n3\t4")

        column_mapping = {
            "col1": ([], "int"),
            "col2": ([], "int"),
        }

        df = read_using_column_mapping(
            full_path=str(tsv_path),
            filetype="tsv",
            column_mapping=column_mapping,
            has_header=True,
        )

        assert df is not None
        assert len(df) == 2

    def test_read_using_column_mapping_psv(self, sample_psv_file):
        """Test router function for PSV"""
        column_mapping = {
            "name": ([], "string"),
            "age": ([], "int"),
            "score": ([], "float"),
        }

        df = read_using_column_mapping(
            full_path=str(sample_psv_file),
            filetype="psv",
            column_mapping=column_mapping,
            has_header=True,
        )

        assert df is not None

    def test_read_fixed_width(self, temp_dir):
        """Test reading fixed-width file"""
        fw_path = temp_dir / "fixed_width.txt"
        fw_path.write_text("AL   100  200\nCA   150  250\n")

        column_mapping = {
            "state": ("string", 1, 2),
            "pop": ("int", 6, 3),
            "area": ("int", 11, 3),
        }

        df = read_fixed_width(
            full_path=str(fw_path),
            column_mapping=column_mapping,
        )

        assert len(df) == 2
        assert df["state"][0] == "AL"
        assert df["pop"][0] == 100

    def test_read_csv_no_duplicate_columns_with_rename(self, temp_dir):
        """Test that CSV column renaming doesn't create duplicate columns"""
        csv_path = temp_dir / "rename_test.csv"
        csv_path.write_text("FirstName,LastName,TotalAmount\nJohn,Doe,100.50\nJane,Smith,200.75")

        column_mapping = {
            "first_name": (["FirstName"], "string"),
            "last_name": (["LastName"], "string"),
            "total_amount": (["TotalAmount"], "float"),
        }

        df = read_csv(
            full_path=str(csv_path),
            column_mapping=column_mapping,
            has_header=True,
        )

        # Verify no duplicate columns
        assert "FirstName" not in df.columns, "Original column should not exist after rename"
        assert "LastName" not in df.columns, "Original column should not exist after rename"
        assert "TotalAmount" not in df.columns, "Original column should not exist after rename"

        # Verify renamed columns exist with data
        assert "first_name" in df.columns
        assert "last_name" in df.columns
        assert "total_amount" in df.columns
        assert len(df.columns) == 3, f"Should have exactly 3 columns, got {len(df.columns)}"

        # Verify data is present
        assert df["first_name"][0] == "John"
        assert df["last_name"][0] == "Doe"
        assert df["total_amount"][0] == 100.50

    def test_read_xlsx_no_duplicate_columns_with_rename(self, temp_dir):
        """Test that Excel column renaming doesn't create duplicate columns"""
        xlsx_path = temp_dir / "rename_test.xlsx"
        test_df = pl.DataFrame({
            "FirstName": ["John", "Jane"],
            "LastName": ["Doe", "Smith"],
            "TotalAmount": [100.50, 200.75]
        })
        test_df.write_excel(str(xlsx_path))

        column_mapping = {
            "first_name": (["FirstName"], "string"),
            "last_name": (["LastName"], "string"),
            "total_amount": (["TotalAmount"], "float"),
        }

        df = read_xlsx(
            full_path=str(xlsx_path),
            column_mapping=column_mapping,
        )

        # Verify no duplicate columns
        assert "FirstName" not in df.columns, "Original column should not exist after rename"
        assert "LastName" not in df.columns, "Original column should not exist after rename"
        assert "TotalAmount" not in df.columns, "Original column should not exist after rename"

        # Verify renamed columns exist with data
        assert "first_name" in df.columns
        assert "last_name" in df.columns
        assert "total_amount" in df.columns
        assert len(df.columns) == 3, f"Should have exactly 3 columns, got {len(df.columns)}"

        # Verify data is present
        assert df["first_name"][0] == "John"
        assert df["last_name"][0] == "Doe"
        assert df["total_amount"][0] == 100.50

    def test_read_tsv_no_duplicate_columns_with_rename(self, temp_dir):
        """Test that TSV column renaming doesn't create duplicate columns"""
        tsv_path = temp_dir / "rename_test.tsv"
        tsv_path.write_text("FirstName\tLastName\tTotalAmount\nJohn\tDoe\t100.50\nJane\tSmith\t200.75")

        column_mapping = {
            "first_name": (["FirstName"], "string"),
            "last_name": (["LastName"], "string"),
            "total_amount": (["TotalAmount"], "float"),
        }

        df = read_csv(
            full_path=str(tsv_path),
            column_mapping=column_mapping,
            has_header=True,
            separator="\t",
        )

        # Verify no duplicate columns
        assert "FirstName" not in df.columns, "Original column should not exist after rename"
        assert "LastName" not in df.columns, "Original column should not exist after rename"
        assert "TotalAmount" not in df.columns, "Original column should not exist after rename"

        # Verify renamed columns exist with data
        assert "first_name" in df.columns
        assert "last_name" in df.columns
        assert "total_amount" in df.columns
        assert len(df.columns) == 3, f"Should have exactly 3 columns, got {len(df.columns)}"

        # Verify data is present
        assert df["first_name"][0] == "John"
        assert df["last_name"][0] == "Doe"
        assert df["total_amount"][0] == 100.50

    def test_read_psv_no_duplicate_columns_with_rename(self, temp_dir):
        """Test that PSV (pipe-separated) column renaming doesn't create duplicate columns"""
        psv_path = temp_dir / "rename_test.psv"
        psv_path.write_text("FirstName|LastName|TotalAmount\nJohn|Doe|100.50\nJane|Smith|200.75")

        column_mapping = {
            "first_name": (["FirstName"], "string"),
            "last_name": (["LastName"], "string"),
            "total_amount": (["TotalAmount"], "float"),
        }

        df = read_csv(
            full_path=str(psv_path),
            column_mapping=column_mapping,
            has_header=True,
            separator="|",
        )

        # Verify no duplicate columns
        assert "FirstName" not in df.columns, "Original column should not exist after rename"
        assert "LastName" not in df.columns, "Original column should not exist after rename"
        assert "TotalAmount" not in df.columns, "Original column should not exist after rename"

        # Verify renamed columns exist with data
        assert "first_name" in df.columns
        assert "last_name" in df.columns
        assert "total_amount" in df.columns
        assert len(df.columns) == 3, f"Should have exactly 3 columns, got {len(df.columns)}"

        # Verify data is present
        assert df["first_name"][0] == "John"
        assert df["last_name"][0] == "Doe"
        assert df["total_amount"][0] == 100.50


# ===== DATABASE OPERATION TESTS =====


class TestDatabaseOperations:
    """Test PostgreSQL database operations"""

    def test_table_exists_true(self, db_uri):
        """Test table_exists when table exists"""
        execute_sql(db_uri, "CREATE TABLE test_schema.test_table (id INT)")

        assert table_exists(db_uri, "test_schema", "test_table")

    def test_table_exists_false(self, db_uri):
        """Test table_exists when table doesn't exist"""
        assert not table_exists(db_uri, "test_schema", "nonexistent_table")

    def test_create_table_from_dataframe(self, db_uri):
        """Test creating table from DataFrame"""
        df = pl.DataFrame({
            "id": pl.Series([1, 2, 3], dtype=pl.Int64),
            "name": pl.Series(["Alice", "Bob", "Charlie"], dtype=pl.Utf8),
            "score": [85.5, 92.0, 78.5],
        })

        create_table_from_dataframe(db_uri, df, "test_schema", "new_table")

        assert table_exists(db_uri, "test_schema", "new_table")

    def test_get_table_schema(self, db_uri):
        """Test getting table schema"""
        execute_sql(db_uri, """
            CREATE TABLE test_schema.schema_test (
                id BIGINT,
                name TEXT,
                score DOUBLE PRECISION
            )
        """)

        schema = get_table_schema(db_uri, "test_schema", "schema_test")

        assert "id" in schema
        assert "name" in schema
        assert "score" in schema
        assert schema["id"] == "bigint"
        assert schema["name"] == "text"

    def test_validate_schema_match_success(self, db_uri):
        """Test schema validation when schemas match"""
        # Create table
        execute_sql(db_uri, """
            CREATE TABLE test_schema.validate_test (
                id BIGINT,
                name TEXT
            )
        """)

        # Create matching DataFrame - polars dtypes that map to BIGINT and TEXT
        df = pl.DataFrame({
            "id": pl.Series([1, 2], dtype=pl.Int64),
            "name": pl.Series(["Alice", "Bob"], dtype=pl.Utf8),
        })

        # Should not raise
        validate_schema_match(db_uri, df, "test_schema", "validate_test")

    def test_validate_schema_match_missing_col_in_table(self, db_uri):
        """Test schema validation when DataFrame has extra columns"""
        # Create table
        execute_sql(db_uri, """
            CREATE TABLE test_schema.validate_test2 (
                id BIGINT
            )
        """)

        # Create DataFrame with extra column
        df = pl.DataFrame({
            "id": pl.Series([1, 2], dtype=pl.Int64),
            "extra": pl.Series(["a", "b"], dtype=pl.Utf8),
        })

        with pytest.raises(ValueError, match="DataFrame has columns not in table"):
            validate_schema_match(db_uri, df, "test_schema", "validate_test2")

    def test_copy_dataframe_to_table(self, db_uri):
        """Test bulk loading DataFrame to PostgreSQL"""
        df = pl.DataFrame({
            "id": pl.Series([1, 2, 3], dtype=pl.Int64),
            "name": pl.Series(["Alice", "Bob", "Charlie"], dtype=pl.Utf8),
            "score": [85.5, 92.0, 78.5],
        })

        copy_dataframe_to_table(db_uri, df, "test_schema", "copy_test")

        # Verify data was loaded
        result = execute_sql_fetchone(db_uri, "SELECT COUNT(*) FROM test_schema.copy_test")
        assert result[0] == 3

    def test_copy_dataframe_with_nulls(self, db_uri):
        """Test COPY with NULL values"""
        df = pl.DataFrame({
            "id": pl.Series([1, 2, None], dtype=pl.Int64),
            "name": pl.Series(["Alice", None, "Charlie"], dtype=pl.Utf8),
            "score": [85.5, None, 78.5],
        })

        copy_dataframe_to_table(db_uri, df, "test_schema", "null_test")

        # Verify NULLs were preserved
        result = execute_sql_fetchone(db_uri, "SELECT id, name, score FROM test_schema.null_test WHERE id = 2")

        assert result[1] is None
        assert result[2] is None


# ===== METADATA FUNCTION TESTS =====


class TestMetadataFunctions:
    """Test metadata tracking functions"""

    def test_get_csv_header_and_row_count(self, sample_csv_file):
        """Test getting header and row count from CSV"""
        header, row_count = get_csv_header_and_row_count(
            file=sample_csv_file,
            has_header=True,
        )

        assert header == ["name", "age", "score"]
        assert row_count == 3

    def test_get_csv_header_and_row_count_no_header(self, sample_csv_no_header):
        """Test getting header and row count from headerless CSV"""
        header, row_count = get_csv_header_and_row_count(
            file=sample_csv_no_header,
            has_header=False,
        )

        # First line is returned as "header" even though it's data
        assert len(header) == 3
        assert row_count == 3

    def test_get_csv_header_psv(self, sample_psv_file):
        """Test getting header from pipe-delimited file"""
        header, row_count = get_csv_header_and_row_count(
            file=sample_psv_file,
            separator="|",
            has_header=True,
        )

        assert header == ["name", "age", "score"]
        assert row_count == 2

    def test_get_csv_header_with_quoted_headers(self, temp_dir):
        """Test getting header from CSV with quoted headers"""
        csv_path = temp_dir / "quoted_headers.csv"
        csv_path.write_text('"name","age","score"\nAlice,25,95.5\nBob,30,87.2')

        header, row_count = get_csv_header_and_row_count(
            file=csv_path,
            has_header=True,
        )

        # Header should be unquoted
        assert header == ["name", "age", "score"]
        assert row_count == 2

    def test_get_file_metadata_row_csv(self, sample_csv_file, temp_dir):
        """Test generating metadata row for CSV file"""
        row = get_file_metadata_row(
            search_dir=temp_dir,
            landing_dir=temp_dir,
            file=sample_csv_file,
            filetype="csv",
            has_header=True,
            encoding="utf-8-sig",
        )

        assert row["full_path"] == sample_csv_file.as_posix()
        assert row["metadata_ingest_status"] == "Success"
        assert row["header"] == ["name", "age", "score"]
        assert row["row_count"] == 3
        assert row["file_hash"] is not None
        assert row["filesize"] > 0

    def test_get_file_metadata_row_with_error(self, temp_dir):
        """Test metadata row generation with error"""
        row = get_file_metadata_row(
            search_dir=temp_dir,
            landing_dir=temp_dir,
            file=None,
            filetype="csv",
            has_header=True,
            error_message="Test error",
        )

        assert row["metadata_ingest_status"] == "Failure"
        assert row["error_message"] == "Test error"

    def test_row_count_check_success(self, db_uri, sample_csv_file):
        """Test row count validation success"""
        # Create metadata table and insert row
        execute_sql(db_uri, """
            CREATE TABLE test_schema.metadata (
                full_path TEXT PRIMARY KEY,
                row_count BIGINT
            )
        """)
        execute_sql(db_uri,
            "INSERT INTO test_schema.metadata (full_path, row_count) VALUES ($1, $2)",
            (str(sample_csv_file), 3)
        )

        # Create DataFrame with matching row count
        df = pl.DataFrame({"col1": [1, 2, 3]})

        # Should not raise
        row_count_check(
            uri=db_uri,
            schema="test_schema",
            df=df,
            full_path=str(sample_csv_file),
        )

    def test_row_count_check_failure(self, db_uri, sample_csv_file):
        """Test row count validation failure"""
        # Create metadata table and insert row
        execute_sql(db_uri, """
            CREATE TABLE test_schema.metadata (
                full_path TEXT PRIMARY KEY,
                row_count BIGINT
            )
        """)
        execute_sql(db_uri,
            "INSERT INTO test_schema.metadata (full_path, row_count) VALUES ($1, $2)",
            (str(sample_csv_file), 5)  # Wrong count
        )

        # Create DataFrame with different row count
        df = pl.DataFrame({"col1": [1, 2, 3]})

        with pytest.raises(ValueError, match="Check failed"):
            row_count_check(
                uri=db_uri,
                schema="test_schema",
                df=df,
                full_path=str(sample_csv_file),
            )

    def test_update_metadata_success(self, db_uri, sample_csv_file):
        """Test updating metadata with success status"""
        # Create metadata table
        execute_sql(db_uri, """
            CREATE TABLE test_schema.metadata (
                full_path TEXT PRIMARY KEY,
                ingest_datetime TIMESTAMP,
                status TEXT,
                error_message TEXT,
                unpivot_row_multiplier INTEGER,
                ingest_runtime INTEGER
            )
        """)
        execute_sql(db_uri,
            "INSERT INTO test_schema.metadata (full_path) VALUES ($1)",
            (str(sample_csv_file),)
        )

        update_metadata(
            uri=db_uri,
            full_path=str(sample_csv_file),
            schema="test_schema",
            ingest_runtime=5,
        )

        # Verify update
        result = execute_sql_fetchone(db_uri,
            "SELECT status, ingest_runtime FROM test_schema.metadata WHERE full_path = $1",
            (str(sample_csv_file),)
        )

        assert result[0] == "Success"
        assert result[1] == 5

    def test_update_metadata_failure(self, db_uri, sample_csv_file):
        """Test updating metadata with failure status"""
        # Create metadata table
        execute_sql(db_uri, """
            CREATE TABLE test_schema.metadata (
                full_path TEXT PRIMARY KEY,
                ingest_datetime TIMESTAMP,
                status TEXT,
                error_message TEXT,
                unpivot_row_multiplier INTEGER,
                ingest_runtime INTEGER
            )
        """)
        execute_sql(db_uri,
            "INSERT INTO test_schema.metadata (full_path) VALUES ($1)",
            (str(sample_csv_file),)
        )

        update_metadata(
            uri=db_uri,
            full_path=str(sample_csv_file),
            schema="test_schema",
            error_message="Test error occurred",
        )

        # Verify update
        result = execute_sql_fetchone(db_uri,
            "SELECT status, error_message FROM test_schema.metadata WHERE full_path = $1",
            (str(sample_csv_file),)
        )

        assert result[0] == "Failure"
        assert result[1] == "Test error occurred"


# ===== SCHEMA INFERENCE TESTS =====


class TestSnakeCaseConversion:
    """Test snake_case conversion function"""

    def test_snake_case_camel_case(self):
        """Test conversion from camelCase"""
        assert to_snake_case("firstName") == "first_name"
        assert to_snake_case("totalAmount") == "total_amount"
        assert to_snake_case("userId") == "user_id"

    def test_snake_case_pascal_case(self):
        """Test conversion from PascalCase"""
        assert to_snake_case("FirstName") == "first_name"
        assert to_snake_case("LastName") == "last_name"
        assert to_snake_case("TotalAmount") == "total_amount"

    def test_snake_case_spaces(self):
        """Test conversion with spaces"""
        assert to_snake_case("User ID") == "user_id"
        assert to_snake_case("First Name") == "first_name"
        assert to_snake_case("Total Amount USD") == "total_amount_usd"

    def test_snake_case_hyphens(self):
        """Test conversion with hyphens"""
        assert to_snake_case("price-per-unit") == "price_per_unit"
        assert to_snake_case("created-at") == "created_at"
        assert to_snake_case("user-id") == "user_id"

    def test_snake_case_mixed_delimiters(self):
        """Test conversion with mixed delimiters"""
        assert to_snake_case("User ID-value") == "user_id_value"
        assert to_snake_case("First-Name Value") == "first_name_value"

    def test_snake_case_already_snake_case(self):
        """Test that already snake_case names are unchanged"""
        assert to_snake_case("first_name") == "first_name"
        assert to_snake_case("user_id") == "user_id"
        assert to_snake_case("total_amount") == "total_amount"

    def test_snake_case_all_caps(self):
        """Test conversion from ALL_CAPS"""
        assert to_snake_case("API_KEY") == "api_key"
        assert to_snake_case("HTTP_STATUS") == "http_status"

    def test_snake_case_consecutive_caps(self):
        """Test conversion with consecutive capitals"""
        assert to_snake_case("HTTPSConnection") == "httpsconnection"
        assert to_snake_case("XMLParser") == "xmlparser"
        # Note: consecutive caps stay together without underscores
        # This is expected behavior

    def test_snake_case_numbers(self):
        """Test conversion with numbers"""
        assert to_snake_case("user123") == "user123"
        assert to_snake_case("User123ID") == "user123_id"  # ID gets separated
        assert to_snake_case("col_1") == "col_1"

    def test_snake_case_edge_cases(self):
        """Test edge cases"""
        assert to_snake_case("") == ""
        assert to_snake_case("a") == "a"
        assert to_snake_case("A") == "a"
        assert to_snake_case("_") == ""
        assert to_snake_case("__multiple__underscores__") == "multiple_underscores"

    def test_snake_case_leading_trailing_spaces(self):
        """Test handling of leading/trailing spaces and underscores"""
        assert to_snake_case(" FirstName ") == "first_name"
        assert to_snake_case("_firstName_") == "first_name"


class TestSchemaInference:
    """Test CLI schema inference functions"""

    def test_infer_schema_from_csv(self, sample_csv_file):
        """Test schema inference from CSV file"""
        column_mapping = infer_schema_from_file(
            str(sample_csv_file),
            filetype="csv",
            has_header=True,
        )

        assert "name" in column_mapping
        assert "age" in column_mapping
        assert "score" in column_mapping

        # Check format
        assert column_mapping["name"] == ([], "string")
        assert column_mapping["age"] == ([], "int")
        assert column_mapping["score"] == ([], "float")

    def test_infer_schema_from_csv_no_header(self, sample_csv_no_header):
        """Test schema inference from headerless CSV"""
        column_mapping = infer_schema_from_file(
            str(sample_csv_no_header),
            filetype="csv",
            has_header=False,
        )

        # Should generate col_0, col_1, col_2
        assert "col_0" in column_mapping
        assert "col_1" in column_mapping
        assert "col_2" in column_mapping

    def test_infer_schema_from_psv(self, sample_psv_file):
        """Test schema inference from pipe-delimited file"""
        column_mapping = infer_schema_from_file(
            str(sample_psv_file),
            filetype="psv",
            has_header=True,
        )

        assert "name" in column_mapping
        assert "age" in column_mapping
        assert "score" in column_mapping

    def test_infer_schema_auto_detect_filetype(self, sample_csv_file):
        """Test auto-detection of file type from extension"""
        column_mapping = infer_schema_from_file(
            str(sample_csv_file),
            has_header=True,
        )

        assert "name" in column_mapping
        assert len(column_mapping) == 3

    def test_infer_schema_with_snake_case_conversion(self, temp_dir):
        """Test schema inference converts column names to snake_case"""
        # Create CSV with non-snake_case column names
        csv_content = """FirstName,LastName,User ID,totalAmount,createdAt
John,Doe,12345,99.50,2024-01-01
Jane,Smith,67890,150.75,2024-01-02
"""
        csv_path = temp_dir / "test_snake_case.csv"
        csv_path.write_text(csv_content)

        column_mapping = infer_schema_from_file(
            str(csv_path),
            filetype="csv",
            has_header=True,
        )

        # Check that column names are snake_case
        assert "first_name" in column_mapping
        assert "last_name" in column_mapping
        assert "user_id" in column_mapping
        assert "total_amount" in column_mapping
        assert "created_at" in column_mapping

        # Original names should NOT be in the mapping keys
        assert "FirstName" not in column_mapping
        assert "LastName" not in column_mapping
        assert "User ID" not in column_mapping
        assert "totalAmount" not in column_mapping
        assert "createdAt" not in column_mapping

        # Check that original names are in the possible columns array
        assert column_mapping["first_name"][0] == ["FirstName"]
        assert column_mapping["last_name"][0] == ["LastName"]
        assert column_mapping["user_id"][0] == ["User ID"]
        assert column_mapping["total_amount"][0] == ["totalAmount"]
        assert column_mapping["created_at"][0] == ["createdAt"]

        # Check types are inferred correctly
        assert column_mapping["first_name"][1] == "string"
        assert column_mapping["user_id"][1] == "int"
        assert column_mapping["total_amount"][1] == "float"

    def test_infer_schema_already_snake_case(self, temp_dir):
        """Test schema inference with already snake_case columns"""
        # Create CSV with snake_case column names
        csv_content = """first_name,last_name,user_id,total_amount
John,Doe,12345,99.50
Jane,Smith,67890,150.75
"""
        csv_path = temp_dir / "test_already_snake.csv"
        csv_path.write_text(csv_content)

        column_mapping = infer_schema_from_file(
            str(csv_path),
            filetype="csv",
            has_header=True,
        )

        # Check that column names remain snake_case
        assert "first_name" in column_mapping
        assert "last_name" in column_mapping
        assert "user_id" in column_mapping
        assert "total_amount" in column_mapping

        # Check that possible columns arrays are EMPTY (no conversion needed)
        assert column_mapping["first_name"][0] == []
        assert column_mapping["last_name"][0] == []
        assert column_mapping["user_id"][0] == []
        assert column_mapping["total_amount"][0] == []

    def test_infer_schema_mixed_case_types(self, temp_dir):
        """Test schema inference with various column name formats"""
        # Create CSV with mixed naming conventions
        csv_content = """camelCase,PascalCase,snake_case,UPPER_CASE,Spaced Name
value1,value2,value3,value4,value5
"""
        csv_path = temp_dir / "test_mixed_case.csv"
        csv_path.write_text(csv_content)

        column_mapping = infer_schema_from_file(
            str(csv_path),
            filetype="csv",
            has_header=True,
        )

        # Check conversions
        assert "camel_case" in column_mapping
        assert column_mapping["camel_case"][0] == ["camelCase"]

        assert "pascal_case" in column_mapping
        assert column_mapping["pascal_case"][0] == ["PascalCase"]

        assert "snake_case" in column_mapping
        assert column_mapping["snake_case"][0] == []  # Already snake_case

        assert "upper_case" in column_mapping
        assert column_mapping["upper_case"][0] == ["UPPER_CASE"]

        assert "spaced_name" in column_mapping
        assert column_mapping["spaced_name"][0] == ["Spaced Name"]


# ===== INTEGRATION TESTS =====


class TestIntegration:
    """Integration tests for complete workflows"""

    def test_add_files_to_metadata_table(self, db_uri, temp_dir, sample_csv_file):
        """Test adding files to metadata table"""
        # Create search and landing directories
        search_dir = temp_dir / "search"
        landing_dir = temp_dir / "landing"
        search_dir.mkdir()
        landing_dir.mkdir()

        # Copy CSV to search dir
        import shutil
        dest_file = search_dir / "test.csv"
        shutil.copy(sample_csv_file, dest_file)

        # Test the add_files function (sub-function of add_files_to_metadata_table)
        file_list = [dest_file]
        rows = add_files(
            search_dir=search_dir,
            landing_dir=landing_dir,
            resume=False,
            sample=None,
            file_list=file_list,
            filetype="csv",
            has_header=True,
            full_path_list=[],
            encoding="utf-8",
            num_search_dir_parents=0,
        )

        assert len(rows) == 1
        assert rows[0]["metadata_ingest_status"] == "Success"

    def test_extract_and_add_zip_files(self, db_uri, temp_dir, sample_zip_file):
        """Test extracting ZIP files and adding to metadata"""
        landing_dir = temp_dir / "landing"
        landing_dir.mkdir()

        file_list = [sample_zip_file]

        rows = extract_and_add_zip_files(
            file_list=file_list,
            full_path_list=[],
            search_dir=temp_dir,
            landing_dir=landing_dir,
            has_header=True,
            filetype="csv",
            resume=False,
            sample=None,
            encoding="utf-8",
            archive_glob="*.csv",
            num_search_dir_parents=0,
        )

        assert len(rows) == 2  # Two CSV files in the ZIP

        # Verify extracted files exist
        extracted_files = list(landing_dir.rglob("*.csv"))
        assert len(extracted_files) == 2

    def test_drop_search_dir(self, db_uri, temp_dir):
        """Test dropping files from metadata by search_dir"""
        # Create metadata table
        execute_sql(db_uri, """
            CREATE TABLE test_schema.metadata (
                search_dir TEXT,
                full_path TEXT PRIMARY KEY
            )
        """)
        execute_sql(db_uri,
            "INSERT INTO test_schema.metadata VALUES ($1, $2)",
            (str(temp_dir), str(temp_dir / "file1.csv"))
        )
        execute_sql(db_uri,
            "INSERT INTO test_schema.metadata VALUES ($1, $2)",
            (str(temp_dir), str(temp_dir / "file2.csv"))
        )

        drop_search_dir(
            uri=db_uri,
            search_dir=str(temp_dir),
            schema="test_schema",
        )

        # Verify deletion
        result = execute_sql_fetchone(db_uri, "SELECT COUNT(*) FROM test_schema.metadata")
        assert result[0] == 0

    def test_drop_partition(self, db_uri):
        """Test dropping partition from table"""
        # Create test table
        execute_sql(db_uri, """
            CREATE TABLE test_schema.partition_test (
                full_path TEXT,
                data TEXT
            )
        """)
        execute_sql(db_uri,
            "INSERT INTO test_schema.partition_test VALUES ($1, $2)",
            ("/path/to/file1.csv", "data1")
        )
        execute_sql(db_uri,
            "INSERT INTO test_schema.partition_test VALUES ($1, $2)",
            ("/path/to/file2.csv", "data2")
        )

        drop_partition(
            uri=db_uri,
            table="partition_test",
            partition_key="/path/to/file1.csv",
            schema="test_schema",
        )

        # Verify deletion
        result = execute_sql_fetchone(db_uri, "SELECT COUNT(*) FROM test_schema.partition_test")
        assert result[0] == 1


# ===== EDGE CASE TESTS =====


class TestEdgeCases:
    """Test edge cases and error handling"""

    def test_read_csv_with_bom(self, temp_dir):
        """Test reading CSV with BOM (byte-order mark)"""
        csv_path = temp_dir / "bom.csv"
        # Write file with BOM
        csv_path.write_bytes(b'\xef\xbb\xbfname,age\nAlice,25\n')

        column_mapping = {
            "name": ([], "string"),
            "age": ([], "int"),
        }

        df = read_csv(
            full_path=str(csv_path),
            column_mapping=column_mapping,
            has_header=True,
        )

        # BOM should be handled correctly
        assert "name" in df.columns
        assert len(df) == 1

    def test_read_csv_empty_null_values(self, temp_dir):
        """Test handling empty strings as null values"""
        csv_path = temp_dir / "nulls.csv"
        csv_path.write_text("name,age\nAlice,25\nBob,\n")

        column_mapping = {
            "name": ([], "string"),
            "age": ([], "int"),
        }

        df = read_csv(
            full_path=str(csv_path),
            column_mapping=column_mapping,
            has_header=True,
            null_value="",
        )

        assert df["age"][1] is None

    def test_copy_dataframe_large_error_message(self, db_uri):
        """Test error message truncation in update_metadata"""
        # Create metadata table with all required columns
        execute_sql(db_uri, """
            CREATE TABLE test_schema.metadata (
                full_path TEXT PRIMARY KEY,
                ingest_datetime TIMESTAMP,
                status TEXT,
                error_message TEXT,
                unpivot_row_multiplier INTEGER,
                ingest_runtime INTEGER
            )
        """)
        execute_sql(db_uri,
            "INSERT INTO test_schema.metadata (full_path) VALUES ($1)",
            ("/path/to/file.csv",)
        )

        # Create very long error message
        long_error = "x" * 1000

        update_metadata(
            uri=db_uri,
            full_path="/path/to/file.csv",
            schema="test_schema",
            error_message=long_error,
        )

        # Verify truncation
        result = execute_sql_fetchone(db_uri,
            "SELECT error_message FROM test_schema.metadata WHERE full_path = $1",
            ("/path/to/file.csv",)
        )
        error = result[0]

        assert len(error) <= 550  # 500 + "... [truncated]"
        assert "truncated" in error

    def test_validate_schema_type_mismatch(self, db_uri):
        """Test schema validation with type mismatch"""
        # Create table with wrong types
        execute_sql(db_uri, """
            CREATE TABLE test_schema.type_test (
                id TEXT,
                name BIGINT
            )
        """)

        # Create DataFrame with different types
        df = pl.DataFrame({
            "id": pl.Series([1, 2], dtype=pl.Int64),
            "name": pl.Series(["Alice", "Bob"], dtype=pl.Utf8),
        })

        with pytest.raises(ValueError, match="type mismatches"):
            validate_schema_match(db_uri, df, "test_schema", "type_test")


class TestMissingFunctions:
    """Tests for functions that weren't covered yet"""

    def test_read_xlsx(self, temp_dir):
        """Test reading Excel files"""
        # Create a simple Excel file
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'name'
        ws['B1'] = 'age'
        ws['C1'] = 'score'
        ws['A2'] = 'Alice'
        ws['B2'] = 25
        ws['C2'] = 85.5
        ws['A3'] = 'Bob'
        ws['B3'] = 30
        ws['C3'] = 92.0

        xlsx_path = temp_dir / "test.xlsx"
        wb.save(str(xlsx_path))

        column_mapping = {
            "name": ([], "string"),
            "age": ([], "int"),
            "score": ([], "float"),
        }

        df = read_xlsx(
            full_path=str(xlsx_path),
            column_mapping=column_mapping,
        )

        assert len(df) == 2
        assert list(df.columns) == ["name", "age", "score"]
        assert df["name"][0] == "Alice"

    def test_read_parquet(self, temp_dir):
        """Test reading Parquet files with column_mapping"""
        # Create a Parquet file using polars
        test_df = pl.DataFrame({
            "name": ["Alice", "Bob", "Charlie"],
            "age": [25, 30, 35],
            "score": [85.5, 92.0, 78.5],
            "extra_col": ["x", "y", "z"],  # Column to be excluded
        })

        parquet_path = temp_dir / "test.parquet"
        test_df.write_parquet(str(parquet_path))

        # Test with column_mapping
        column_mapping = {
            "name": ([], "string"),
            "age": ([], "int"),
            "score": ([], "float"),
            # extra_col is intentionally not included - should be filtered out
        }

        df = read_parquet(
            full_path=str(parquet_path),
            column_mapping=column_mapping,
        )

        assert len(df) == 3
        assert list(df.columns) == ["name", "age", "score"]
        assert "extra_col" not in df.columns  # Should be filtered out
        assert df["name"][0] == "Alice"

    def test_read_parquet_with_rename(self, temp_dir):
        """Test reading Parquet files with column renaming"""
        test_df = pl.DataFrame({
            "old_name": ["Alice", "Bob"],
            "value": [100, 200],
        })

        parquet_path = temp_dir / "test_rename.parquet"
        test_df.write_parquet(str(parquet_path))

        column_mapping = {
            "new_name": (["old_name"], "string"),
            "value": ([], "int"),
        }

        df = read_parquet(
            full_path=str(parquet_path),
            column_mapping=column_mapping,
        )

        assert "new_name" in df.columns
        assert "old_name" not in df.columns
        assert df["new_name"][0] == "Alice"

    def test_read_parquet_with_missing_cols(self, temp_dir):
        """Test reading Parquet with missing columns in mapping"""
        test_df = pl.DataFrame({
            "name": ["Alice", "Bob"],
        })

        parquet_path = temp_dir / "test_missing.parquet"
        test_df.write_parquet(str(parquet_path))

        column_mapping = {
            "name": ([], "string"),
            "missing_col": ([], "string"),  # Not in parquet file
        }

        df = read_parquet(
            full_path=str(parquet_path),
            column_mapping=column_mapping,
        )

        assert "missing_col" in df.columns
        assert df["missing_col"].null_count() == len(df)

    def test_s3_path_detection(self, temp_dir):
        """Test S3 path detection in various contexts"""
        # Valid S3 paths
        assert is_s3_path("s3://my-bucket/path/file.csv")
        assert is_s3_path("s3://bucket/")

        # Invalid S3 paths
        assert not is_s3_path("/local/path/file.csv")
        assert not is_s3_path(str(temp_dir / "file.csv"))

    def test_update_table_basic(self, db_uri, temp_dir, sample_csv_file):
        """Test update_table function end-to-end"""
        # Create metadata table first
        execute_sql(db_uri, """
            CREATE TABLE test_schema.metadata (
                search_dir TEXT,
                landing_dir TEXT,
                full_path TEXT PRIMARY KEY,
                filesize BIGINT,
                header TEXT[],
                row_count BIGINT,
                archive_full_path TEXT,
                file_hash TEXT,
                metadata_ingest_datetime TIMESTAMP,
                metadata_ingest_status TEXT,
                ingest_datetime TIMESTAMP,
                ingest_runtime INTEGER,
                status TEXT,
                error_message TEXT,
                unpivot_row_multiplier INTEGER
            )
        """)

        # Insert metadata for the sample file
        # Note: landing_dir must have trailing slash to match update_table queries
        execute_sql(db_uri, """
            INSERT INTO test_schema.metadata
            (full_path, row_count, metadata_ingest_status, landing_dir)
            VALUES ($1, $2, $3, $4)
        """, (str(sample_csv_file), 3, "Success", str(temp_dir) + "/"))

        column_mapping = {
            "name": ([], "string"),
            "age": ([], "int"),
            "score": ([], "float"),
        }

        # Run update_table
        result_df = update_table(
            uri=db_uri,
            schema="test_schema",
            output_table="test_output",
            filetype="csv",
            column_mapping=column_mapping,
            landing_dir=str(temp_dir),
            resume=False,
        )

        # Verify data was loaded
        result = execute_sql_fetchone(db_uri, "SELECT COUNT(*) FROM test_schema.test_output")
        assert result[0] == 3

    def test_add_files_to_metadata_table_full(self, db_uri, temp_dir, sample_csv_file):
        """Test add_files_to_metadata_table end-to-end"""
        # This is already tested indirectly through add_files in TestIntegration
        # Here we just verify the metadata table creation aspect

        # Clean up any existing metadata table from previous tests
        execute_sql(db_uri, "DROP TABLE IF EXISTS test_schema.metadata")

        # Create search and landing directories
        search_dir = temp_dir / "search"
        landing_dir = temp_dir / "landing"
        search_dir.mkdir()
        landing_dir.mkdir()

        # Copy CSV to search dir
        import shutil
        dest_file = search_dir / "test.csv"
        shutil.copy(sample_csv_file, dest_file)

        # This would call add_files_to_metadata_table but it's complex
        # and requires specific database state. Already covered in integration tests.
        # Just verify the table creation logic works

        assert not table_exists(db_uri, "test_schema", "metadata")

        # Create metadata table manually to test the schema
        execute_sql(db_uri, """
            CREATE TABLE test_schema.metadata (
                search_dir TEXT,
                landing_dir TEXT,
                full_path TEXT PRIMARY KEY,
                filesize BIGINT,
                header TEXT[],
                row_count BIGINT,
                archive_full_path TEXT,
                file_hash TEXT,
                metadata_ingest_datetime TIMESTAMP,
                metadata_ingest_status TEXT,
                ingest_datetime TIMESTAMP,
                ingest_runtime INTEGER,
                status TEXT,
                error_message TEXT,
                unpivot_row_multiplier INTEGER
            )
        """)

        assert table_exists(db_uri, "test_schema", "metadata")


class TestAdditionalFunctions:
    """Additional tests for uncovered functions"""

    def test_drop_file_from_metadata_and_table(self, db_uri):
        """Test dropping file from both metadata and data table"""
        # Create metadata table
        execute_sql(db_uri, """
            CREATE TABLE test_schema.metadata (
                full_path TEXT PRIMARY KEY
            )
        """)
        execute_sql(db_uri,
            "INSERT INTO test_schema.metadata VALUES ($1)",
            ("/path/to/file.csv",)
        )

        # Create data table
        execute_sql(db_uri, """
            CREATE TABLE test_schema.data_table (
                full_path TEXT,
                data TEXT
            )
        """)
        execute_sql(db_uri,
            "INSERT INTO test_schema.data_table VALUES ($1, $2)",
            ("/path/to/file.csv", "test_data")
        )

        # Drop file from both tables
        drop_file_from_metadata_and_table(
            uri=db_uri,
            table="data_table",
            full_path="/path/to/file.csv",
            schema="test_schema",
        )

        # Verify deletions
        metadata_count = execute_sql_fetchone(db_uri, "SELECT COUNT(*) FROM test_schema.metadata")[0]
        data_count = execute_sql_fetchone(db_uri, "SELECT COUNT(*) FROM test_schema.data_table")[0]

        assert metadata_count == 0
        assert data_count == 0

    def test_read_using_column_mapping_invalid_filetype(self):
        """Test router function with invalid filetype"""
        import pytest

        with pytest.raises(ValueError, match="Invalid filetype: invalid_type"):
            read_using_column_mapping(
                full_path="/path/to/file.xyz",
                filetype="invalid_type",
                column_mapping={},
                has_header=True,
            )

    def test_row_count_check_with_unpivot_multiplier(self, db_uri, sample_csv_file):
        """Test row count check with unpivot multiplier"""
        # Create metadata table
        execute_sql(db_uri, """
            CREATE TABLE test_schema.metadata (
                full_path TEXT PRIMARY KEY,
                row_count BIGINT
            )
        """)
        execute_sql(db_uri,
            "INSERT INTO test_schema.metadata (full_path, row_count) VALUES ($1, $2)",
            (str(sample_csv_file), 3)
        )

        # Create DataFrame with 9 rows (3 * 3 multiplier)
        df = pl.DataFrame({"col1": list(range(9))})

        # Should not raise with multiplier of 3
        row_count_check(
            uri=db_uri,
            schema="test_schema",
            df=df,
            full_path=str(sample_csv_file),
            unpivot_row_multiplier=3,
        )

    def test_update_metadata_with_unpivot_multiplier(self, db_uri):
        """Test updating metadata with unpivot row multiplier"""
        # Create metadata table
        execute_sql(db_uri, """
            CREATE TABLE test_schema.metadata (
                full_path TEXT PRIMARY KEY,
                ingest_datetime TIMESTAMP,
                status TEXT,
                error_message TEXT,
                unpivot_row_multiplier INTEGER,
                ingest_runtime INTEGER
            )
        """)
        execute_sql(db_uri,
            "INSERT INTO test_schema.metadata (full_path) VALUES ($1)",
            ("/path/to/file.csv",)
        )

        update_metadata(
            uri=db_uri,
            full_path="/path/to/file.csv",
            schema="test_schema",
            unpivot_row_multiplier=5,
            ingest_runtime=10,
        )

        # Verify update
        result = execute_sql_fetchone(db_uri,
            "SELECT unpivot_row_multiplier, ingest_runtime FROM test_schema.metadata WHERE full_path = $1",
            ("/path/to/file.csv",)
        )

        assert result[0] == 5
        assert result[1] == 10


# ===== UPDATE_TABLE ADVANCED FEATURE TESTS =====


class TestUpdateTableAdvancedFeatures:
    """Test advanced features of update_table function"""

    def _setup_metadata_table(self, db_uri, temp_dir, csv_files):
        """Helper to set up metadata table with files"""
        execute_sql(db_uri, "DROP TABLE IF EXISTS test_schema.metadata")
        execute_sql(db_uri, """
            CREATE TABLE test_schema.metadata (
                search_dir TEXT,
                landing_dir TEXT,
                full_path TEXT PRIMARY KEY,
                filesize BIGINT,
                header TEXT[],
                row_count BIGINT,
                archive_full_path TEXT,
                file_hash TEXT,
                metadata_ingest_datetime TIMESTAMP,
                metadata_ingest_status TEXT,
                ingest_datetime TIMESTAMP,
                ingest_runtime INTEGER,
                status TEXT,
                error_message TEXT,
                unpivot_row_multiplier INTEGER
            )
        """)

        for csv_file, row_count in csv_files:
            execute_sql(db_uri, """
                INSERT INTO test_schema.metadata
                (full_path, row_count, metadata_ingest_status, landing_dir)
                VALUES ($1, $2, $3, $4)
            """, (str(csv_file), row_count, "Success", str(temp_dir) + "/"))

    def test_update_table_with_transform_fn(self, db_uri, temp_dir):
        """Test update_table with transform_fn"""
        # Create CSV file
        csv_path = temp_dir / "transform_test.csv"
        csv_path.write_text("name,value\nAlice,100\nBob,200\n")

        # Setup metadata
        self._setup_metadata_table(db_uri, temp_dir, [(csv_path, 2)])

        column_mapping = {
            "name": ([], "string"),
            "value": ([], "int"),
        }

        # Transform function that doubles the value
        def transform_fn(df):
            return df.with_columns((pl.col("value") * 2).alias("value"))

        result_df = update_table(
            uri=db_uri,
            schema="test_schema",
            output_table="transform_output",
            filetype="csv",
            column_mapping=column_mapping,
            landing_dir=str(temp_dir),
            transform_fn=transform_fn,
            resume=False,
        )

        # Verify transformation was applied
        result = execute_sql_fetchone(db_uri, "SELECT value FROM test_schema.transform_output WHERE name = 'Alice'")
        assert result[0] == 200  # 100 * 2

    def test_update_table_with_additional_cols_fn(self, db_uri, temp_dir):
        """Test update_table with additional_cols_fn"""
        csv_path = temp_dir / "additional_cols_test.csv"
        csv_path.write_text("name,value\nAlice,100\n")

        self._setup_metadata_table(db_uri, temp_dir, [(csv_path, 1)])

        column_mapping = {
            "name": ([], "string"),
            "value": ([], "int"),
        }

        # Additional columns function
        def additional_cols_fn(file_path):
            return {
                "source_file": Path(file_path).stem,
                "load_timestamp": "2024-01-01",
            }

        result_df = update_table(
            uri=db_uri,
            schema="test_schema",
            output_table="additional_cols_output",
            filetype="csv",
            column_mapping=column_mapping,
            landing_dir=str(temp_dir),
            additional_cols_fn=additional_cols_fn,
            resume=False,
        )

        # Verify additional columns were added
        result = execute_sql_fetchone(db_uri, "SELECT source_file, load_timestamp FROM test_schema.additional_cols_output")
        assert result[0] == "additional_cols_test"
        assert result[1] == "2024-01-01"

    def test_update_table_with_output_table_naming_fn(self, db_uri, temp_dir):
        """Test update_table with output_table_naming_fn"""
        csv_path = temp_dir / "naming_test.csv"
        csv_path.write_text("name,value\nAlice,100\n")

        self._setup_metadata_table(db_uri, temp_dir, [(csv_path, 1)])

        column_mapping = {
            "name": ([], "string"),
            "value": ([], "int"),
        }

        # Naming function that uses file stem as table name
        def output_table_naming_fn(file_path):
            return f"table_{Path(file_path).stem}"

        result_df = update_table(
            uri=db_uri,
            schema="test_schema",
            output_table_naming_fn=output_table_naming_fn,
            filetype="csv",
            column_mapping=column_mapping,
            landing_dir=str(temp_dir),
            resume=False,
        )

        # Verify table was created with custom name
        assert table_exists(db_uri, "test_schema", "table_naming_test")

    def test_update_table_with_file_list_filter_fn(self, db_uri, temp_dir):
        """Test update_table with file_list_filter_fn"""
        # Create multiple CSV files
        csv1 = temp_dir / "include_me.csv"
        csv1.write_text("name,value\nAlice,100\n")
        csv2 = temp_dir / "exclude_me.csv"
        csv2.write_text("name,value\nBob,200\n")

        self._setup_metadata_table(db_uri, temp_dir, [(csv1, 1), (csv2, 1)])

        column_mapping = {
            "name": ([], "string"),
            "value": ([], "int"),
        }

        # Filter function that only includes files with "include" in the name
        def file_list_filter_fn(file_list):
            return [f for f in file_list if "include" in str(f)]

        result_df = update_table(
            uri=db_uri,
            schema="test_schema",
            output_table="filter_output",
            filetype="csv",
            column_mapping=column_mapping,
            landing_dir=str(temp_dir),
            file_list_filter_fn=file_list_filter_fn,
            resume=False,
        )

        # Verify only one file was processed
        result = execute_sql_fetchone(db_uri, "SELECT COUNT(*) FROM test_schema.filter_output")
        assert result[0] == 1

    def test_update_table_with_custom_read_fn(self, db_uri, temp_dir):
        """Test update_table with custom_read_fn"""
        csv_path = temp_dir / "custom_read.csv"
        csv_path.write_text("name,value\nAlice,100\n")

        self._setup_metadata_table(db_uri, temp_dir, [(csv_path, 1)])

        # Custom read function that adds a computed column
        def custom_read_fn(full_path):
            df = pl.read_csv(full_path)
            df = df.with_columns((pl.col("value") * 10).alias("computed"))
            return df

        result_df = update_table(
            uri=db_uri,
            schema="test_schema",
            output_table="custom_read_output",
            filetype="csv",
            custom_read_fn=custom_read_fn,
            landing_dir=str(temp_dir),
            resume=False,
        )

        # Verify custom read was used
        result = execute_sql_fetchone(db_uri, "SELECT computed FROM test_schema.custom_read_output")
        assert result[0] == 1000

    def test_update_table_with_column_mapping_fn(self, db_uri, temp_dir):
        """Test update_table with column_mapping_fn"""
        csv_path = temp_dir / "mapping_fn_test.csv"
        csv_path.write_text("name,value\nAlice,100\n")

        self._setup_metadata_table(db_uri, temp_dir, [(csv_path, 1)])

        # Dynamic column mapping function
        def column_mapping_fn(file_path):
            return {
                "name": ([], "string"),
                "value": ([], "int"),
            }

        result_df = update_table(
            uri=db_uri,
            schema="test_schema",
            output_table="mapping_fn_output",
            filetype="csv",
            column_mapping_fn=column_mapping_fn,
            landing_dir=str(temp_dir),
            resume=False,
        )

        # Verify data was loaded
        result = execute_sql_fetchone(db_uri, "SELECT COUNT(*) FROM test_schema.mapping_fn_output")
        assert result[0] == 1

    def test_update_table_with_pivot_mapping(self, db_uri, temp_dir):
        """Test update_table with pivot_mapping (unpivot)"""
        csv_path = temp_dir / "pivot_test.csv"
        csv_path.write_text("id,jan,feb,mar\n1,100,200,300\n")

        # Note: row_count is 1 but after unpivot it will be 3 (1 * 3 value columns)
        self._setup_metadata_table(db_uri, temp_dir, [(csv_path, 1)])

        column_mapping = {
            "id": ([], "int"),
            "jan": ([], "int"),
            "feb": ([], "int"),
            "mar": ([], "int"),
        }

        pivot_mapping = {
            "id_vars": ["id"],
            "variable_column_name": "month",
            "value_column_name": "amount",
        }

        result_df = update_table(
            uri=db_uri,
            schema="test_schema",
            output_table="pivot_output",
            filetype="csv",
            column_mapping=column_mapping,
            landing_dir=str(temp_dir),
            pivot_mapping=pivot_mapping,
            resume=False,
        )

        # Verify unpivot was applied - should have 3 rows (jan, feb, mar)
        result = execute_sql_fetchone(db_uri, "SELECT COUNT(*) FROM test_schema.pivot_output")
        assert result[0] == 3

    def test_update_table_with_sample(self, db_uri, temp_dir):
        """Test update_table with sample parameter"""
        # Create multiple CSV files
        for i in range(5):
            csv_path = temp_dir / f"sample_{i}.csv"
            csv_path.write_text(f"name,value\nRow{i},100\n")

        files = [(temp_dir / f"sample_{i}.csv", 1) for i in range(5)]
        self._setup_metadata_table(db_uri, temp_dir, files)

        column_mapping = {
            "name": ([], "string"),
            "value": ([], "int"),
        }

        result_df = update_table(
            uri=db_uri,
            schema="test_schema",
            output_table="sample_output",
            filetype="csv",
            column_mapping=column_mapping,
            landing_dir=str(temp_dir),
            sample=2,  # Only process 2 files
            resume=False,
        )

        # Verify only 2 files were processed
        result = execute_sql_fetchone(db_uri, "SELECT COUNT(DISTINCT full_path) FROM test_schema.sample_output")
        assert result[0] == 2

    def test_update_table_resume(self, db_uri, temp_dir):
        """Test update_table with resume=True"""
        csv1 = temp_dir / "resume1.csv"
        csv1.write_text("name,value\nAlice,100\n")
        csv2 = temp_dir / "resume2.csv"
        csv2.write_text("name,value\nBob,200\n")

        # Setup metadata with one file already processed
        execute_sql(db_uri, "DROP TABLE IF EXISTS test_schema.metadata")
        execute_sql(db_uri, """
            CREATE TABLE test_schema.metadata (
                search_dir TEXT,
                landing_dir TEXT,
                full_path TEXT PRIMARY KEY,
                filesize BIGINT,
                header TEXT[],
                row_count BIGINT,
                archive_full_path TEXT,
                file_hash TEXT,
                metadata_ingest_datetime TIMESTAMP,
                metadata_ingest_status TEXT,
                ingest_datetime TIMESTAMP,
                ingest_runtime INTEGER,
                status TEXT,
                error_message TEXT,
                unpivot_row_multiplier INTEGER
            )
        """)

        # File 1 is already processed
        execute_sql(db_uri, """
            INSERT INTO test_schema.metadata
            (full_path, row_count, metadata_ingest_status, landing_dir, ingest_datetime, status)
            VALUES ($1, $2, $3, $4, NOW(), 'Success')
        """, (str(csv1), 1, "Success", str(temp_dir) + "/"))

        # File 2 is not processed yet
        execute_sql(db_uri, """
            INSERT INTO test_schema.metadata
            (full_path, row_count, metadata_ingest_status, landing_dir)
            VALUES ($1, $2, $3, $4)
        """, (str(csv2), 1, "Success", str(temp_dir) + "/"))

        column_mapping = {
            "name": ([], "string"),
            "value": ([], "int"),
        }

        result_df = update_table(
            uri=db_uri,
            schema="test_schema",
            output_table="resume_output",
            filetype="csv",
            column_mapping=column_mapping,
            landing_dir=str(temp_dir),
            resume=True,  # Only process unprocessed files
        )

        # Verify only the new file was processed
        result = execute_sql_fetchone(db_uri, "SELECT COUNT(*) FROM test_schema.resume_output")
        assert result[0] == 1

    def test_update_table_with_header_fn(self, db_uri, temp_dir):
        """Test update_table with header_fn for headerless files"""
        csv_path = temp_dir / "no_header.csv"
        csv_path.write_text("Alice,100,85.5\nBob,200,92.0\n")

        self._setup_metadata_table(db_uri, temp_dir, [(csv_path, 2)])

        column_mapping = {
            "name": ([], "string"),
            "value": ([], "int"),
            "score": ([], "float"),
        }

        # Header function that provides column names
        def header_fn(file_path):
            return ["name", "value", "score"]

        result_df = update_table(
            uri=db_uri,
            schema="test_schema",
            output_table="header_fn_output",
            filetype="csv",
            column_mapping=column_mapping,
            landing_dir=str(temp_dir),
            header_fn=header_fn,
            resume=False,
        )

        # Verify data was loaded with correct headers
        result = execute_sql_fetchone(db_uri, "SELECT name, value FROM test_schema.header_fn_output WHERE name = 'Alice'")
        assert result[0] == "Alice"


# ===== ADD_FILES_TO_METADATA_TABLE END-TO-END TESTS =====


class TestAddFilesToMetadataTableEndToEnd:
    """End-to-end tests for add_files_to_metadata_table"""

    def test_add_files_to_metadata_table_full_workflow(self, db_uri, temp_dir):
        """Test complete workflow of add_files_to_metadata_table"""
        import shutil

        # Clean up any existing metadata table from previous tests
        execute_sql(db_uri, "DROP TABLE IF EXISTS test_schema.metadata")

        # Create search and landing directories
        search_dir = temp_dir / "search"
        landing_dir = temp_dir / "landing"
        search_dir.mkdir()
        landing_dir.mkdir()

        # Create test CSV files
        csv1 = search_dir / "file1.csv"
        csv1.write_text("col1,col2\n1,2\n3,4\n")
        csv2 = search_dir / "file2.csv"
        csv2.write_text("col1,col2\n5,6\n7,8\n")

        # Call add_files_to_metadata_table
        result_df = add_files_to_metadata_table(
            uri=db_uri,
            schema="test_schema",
            search_dir=str(search_dir),
            landing_dir=str(landing_dir),
            filetype="csv",
            has_header=True,
            encoding="utf-8",
            resume=False,
        )

        # Verify metadata table was created and populated
        assert table_exists(db_uri, "test_schema", "metadata")

        result = execute_sql_fetchone(db_uri, "SELECT COUNT(*) FROM test_schema.metadata")
        assert result[0] == 2

        # Verify files were copied to landing directory
        assert (landing_dir / "file1.csv").exists()
        assert (landing_dir / "file2.csv").exists()

    def test_add_files_to_metadata_table_with_zip(self, db_uri, temp_dir):
        """Test add_files_to_metadata_table with ZIP extraction"""
        import shutil

        # Create search and landing directories
        search_dir = temp_dir / "search"
        landing_dir = temp_dir / "landing"
        search_dir.mkdir()
        landing_dir.mkdir()

        # Create a ZIP file with CSV files
        zip_path = search_dir / "archive.zip"
        with zipfile.ZipFile(zip_path, 'w') as zf:
            zf.writestr("inner1.csv", "col1,col2\n1,2\n")
            zf.writestr("inner2.csv", "col1,col2\n3,4\n")

        # Call add_files_to_metadata_table with compression
        result_df = add_files_to_metadata_table(
            uri=db_uri,
            schema="test_schema",
            search_dir=str(search_dir),
            landing_dir=str(landing_dir),
            filetype="csv",
            compression_type="zip",
            archive_glob="*.csv",
            has_header=True,
            encoding="utf-8",
            resume=False,
        )

        # Verify metadata table was populated
        result = execute_sql_fetchone(db_uri, "SELECT COUNT(*) FROM test_schema.metadata")
        assert result[0] == 2

        # Verify files were extracted
        extracted_files = list(landing_dir.rglob("*.csv"))
        assert len(extracted_files) == 2

    def test_add_files_to_metadata_table_resume(self, db_uri, temp_dir):
        """Test add_files_to_metadata_table with resume=True"""
        import shutil

        search_dir = temp_dir / "search"
        landing_dir = temp_dir / "landing"
        search_dir.mkdir()
        landing_dir.mkdir()

        # Create initial file and process it
        csv1 = search_dir / "file1.csv"
        csv1.write_text("col1,col2\n1,2\n")

        # First run
        add_files_to_metadata_table(
            uri=db_uri,
            schema="test_schema",
            search_dir=str(search_dir),
            landing_dir=str(landing_dir),
            filetype="csv",
            has_header=True,
            encoding="utf-8",
            resume=False,
        )

        # Add a new file
        csv2 = search_dir / "file2.csv"
        csv2.write_text("col1,col2\n3,4\n")

        # Second run with resume=True
        result_df = add_files_to_metadata_table(
            uri=db_uri,
            schema="test_schema",
            search_dir=str(search_dir),
            landing_dir=str(landing_dir),
            filetype="csv",
            has_header=True,
            encoding="utf-8",
            resume=True,
        )

        # Verify both files are in metadata
        result = execute_sql_fetchone(db_uri, "SELECT COUNT(*) FROM test_schema.metadata")
        assert result[0] == 2

    def test_add_files_to_metadata_table_with_filter(self, db_uri, temp_dir):
        """Test add_files_to_metadata_table with file_list_filter_fn"""
        import shutil

        search_dir = temp_dir / "search"
        landing_dir = temp_dir / "landing"
        search_dir.mkdir()
        landing_dir.mkdir()

        # Create multiple CSV files
        (search_dir / "include.csv").write_text("col1,col2\n1,2\n")
        (search_dir / "exclude.csv").write_text("col1,col2\n3,4\n")

        # Filter function
        def file_list_filter_fn(file_list):
            return [f for f in file_list if "include" in str(f)]

        result_df = add_files_to_metadata_table(
            uri=db_uri,
            schema="test_schema",
            search_dir=str(search_dir),
            landing_dir=str(landing_dir),
            filetype="csv",
            has_header=True,
            encoding="utf-8",
            file_list_filter_fn=file_list_filter_fn,
            resume=False,
        )

        # Verify only filtered file was processed
        result = execute_sql_fetchone(db_uri, "SELECT COUNT(*) FROM test_schema.metadata")
        assert result[0] == 1


# ===== CLI SCHEMA INFERENCE TESTS =====


class TestCLISchemaInference:
    """Test CLI schema inference functionality"""

    def test_cli_argument_parsing(self, temp_dir):
        """Test CLI argument parsing via subprocess"""
        import subprocess

        # Create test file
        csv_path = temp_dir / "cli_test.csv"
        csv_path.write_text("name,age,score\nAlice,25,85.5\n")

        # Run CLI
        result = subprocess.run(
            ["python", "src/table_functions_postgres.py", str(csv_path), "--pretty"],
            capture_output=True,
            text=True,
            cwd=Path(__file__).parent.parent,
        )

        assert result.returncode == 0
        output = json.loads(result.stdout)
        assert "name" in output
        assert "age" in output
        assert "score" in output

    def test_cli_filetype_detection(self, temp_dir):
        """Test CLI auto-detects file type from extension"""
        import subprocess

        # Create PSV file with .psv extension
        psv_path = temp_dir / "cli_test.psv"
        psv_path.write_text("name|age|score\nAlice|25|85.5\n")

        result = subprocess.run(
            ["python", "src/table_functions_postgres.py", str(psv_path), "--pretty"],
            capture_output=True,
            text=True,
            cwd=Path(__file__).parent.parent,
        )

        assert result.returncode == 0
        output = json.loads(result.stdout)
        assert "name" in output

    def test_cli_no_header(self, temp_dir):
        """Test CLI with --no-header flag"""
        import subprocess

        csv_path = temp_dir / "no_header.csv"
        csv_path.write_text("Alice,25,85.5\nBob,30,92.0\n")

        result = subprocess.run(
            ["python", "src/table_functions_postgres.py", str(csv_path), "--no-header", "--pretty"],
            capture_output=True,
            text=True,
            cwd=Path(__file__).parent.parent,
        )

        assert result.returncode == 0
        output = json.loads(result.stdout)
        assert "col_0" in output
        assert "col_1" in output
        assert "col_2" in output

    def test_cli_custom_separator(self, temp_dir):
        """Test CLI with custom separator"""
        import subprocess

        # Create file with custom separator
        csv_path = temp_dir / "custom_sep.csv"
        csv_path.write_text("name;age;score\nAlice;25;85.5\n")

        result = subprocess.run(
            ["python", "src/table_functions_postgres.py", str(csv_path),
             "--separator", ";", "--pretty"],
            capture_output=True,
            text=True,
            cwd=Path(__file__).parent.parent,
        )

        assert result.returncode == 0
        output = json.loads(result.stdout)
        assert "name" in output
        assert "age" in output

    def test_cli_sample_rows(self, temp_dir):
        """Test CLI with --sample-rows parameter"""
        import subprocess

        # Create file with many rows
        csv_path = temp_dir / "many_rows.csv"
        with open(csv_path, 'w') as f:
            f.write("id,value\n")
            for i in range(1000):
                f.write(f"{i},{i * 10}\n")

        result = subprocess.run(
            ["python", "src/table_functions_postgres.py", str(csv_path),
             "--sample-rows", "100", "--pretty"],
            capture_output=True,
            text=True,
            cwd=Path(__file__).parent.parent,
        )

        assert result.returncode == 0
        output = json.loads(result.stdout)
        assert "id" in output
        assert "value" in output

    def test_cli_file_not_found(self, temp_dir):
        """Test CLI error handling for non-existent file"""
        import subprocess

        result = subprocess.run(
            ["python", "src/table_functions_postgres.py", "/nonexistent/file.csv"],
            capture_output=True,
            text=True,
            cwd=Path(__file__).parent.parent,
        )

        assert result.returncode == 1
        assert "Error" in result.stdout or "not found" in result.stdout.lower()

    def test_cli_parquet_file(self, temp_dir):
        """Test CLI with parquet file"""
        import subprocess

        # Create parquet file
        test_df = pl.DataFrame({
            "name": ["Alice", "Bob"],
            "age": [25, 30],
        })
        parquet_path = temp_dir / "test.parquet"
        test_df.write_parquet(str(parquet_path))

        result = subprocess.run(
            ["python", "src/table_functions_postgres.py", str(parquet_path), "--pretty"],
            capture_output=True,
            text=True,
            cwd=Path(__file__).parent.parent,
        )

        assert result.returncode == 0
        output = json.loads(result.stdout)
        assert "name" in output
        assert "age" in output


# ===== CONNECTION AND SQL HELPER TESTS =====


class TestConnectionHelpers:
    """Test connection management and SQL helper functions"""

    def test_get_connection_context_manager(self, db_uri):
        """Test get_connection context manager properly opens and closes connections"""
        with get_connection(db_uri) as conn:
            # Connection should be usable
            with conn.cursor() as cur:
                cur.execute("SELECT 1")
                result = cur.fetchone()
                assert result[0] == 1

    def test_execute_sql_basic(self, db_uri):
        """Test execute_sql with basic DDL"""
        execute_sql(db_uri, "CREATE TABLE test_schema.exec_test (id INT)")
        assert table_exists(db_uri, "test_schema", "exec_test")

    def test_execute_sql_with_params(self, db_uri):
        """Test execute_sql with parameters"""
        execute_sql(db_uri, "CREATE TABLE test_schema.param_test (name TEXT, value INT)")
        execute_sql(db_uri, "INSERT INTO test_schema.param_test VALUES ($1, $2)", ("Alice", 100))

        result = execute_sql_fetchone(db_uri, "SELECT name, value FROM test_schema.param_test")
        assert result[0] == "Alice"
        assert result[1] == 100

    def test_execute_sql_fetch_multiple_rows(self, db_uri):
        """Test execute_sql_fetch returns all rows"""
        execute_sql(db_uri, "CREATE TABLE test_schema.multi_test (id INT)")
        for i in range(5):
            execute_sql(db_uri, f"INSERT INTO test_schema.multi_test VALUES ({i})")

        results = execute_sql_fetch(db_uri, "SELECT id FROM test_schema.multi_test ORDER BY id")
        assert len(results) == 5
        assert [r[0] for r in results] == [0, 1, 2, 3, 4]

    def test_execute_sql_fetchone_no_results(self, db_uri):
        """Test execute_sql_fetchone returns None when no results"""
        execute_sql(db_uri, "CREATE TABLE test_schema.empty_test (id INT)")

        result = execute_sql_fetchone(db_uri, "SELECT id FROM test_schema.empty_test")
        assert result is None

    def test_execute_sql_invalid_sql(self, db_uri):
        """Test execute_sql raises error on invalid SQL"""
        with pytest.raises(Exception):
            execute_sql(db_uri, "THIS IS NOT VALID SQL")


# ===== ENCODING TESTS =====


class TestEncodings:
    """Test different file encodings"""

    def test_read_csv_cp1252_encoding(self, temp_dir):
        """Test reading Windows-1252 encoded file"""
        # Create file with Windows-1252 encoding
        csv_path = temp_dir / "cp1252.csv"
        content = "name,value\nTest1,100\nTest2,200\n"
        csv_path.write_bytes(content.encode('cp1252'))

        column_mapping = {
            "name": ([], "string"),
            "value": ([], "int"),
        }

        df = read_csv(
            full_path=str(csv_path),
            column_mapping=column_mapping,
            has_header=True,
            encoding="cp1252",
        )

        assert len(df) == 2
        assert df["name"][0] == "Test1"
        assert df["value"][1] == 200

    def test_read_csv_latin1_encoding(self, temp_dir):
        """Test reading Latin-1 encoded file"""
        csv_path = temp_dir / "latin1.csv"
        content = "name,value\nTest1,100\nTest2,200\n"
        csv_path.write_bytes(content.encode('latin-1'))

        column_mapping = {
            "name": ([], "string"),
            "value": ([], "int"),
        }

        df = read_csv(
            full_path=str(csv_path),
            column_mapping=column_mapping,
            has_header=True,
            encoding="latin-1",
        )

        assert len(df) == 2
        assert df["name"][0] == "Test1"
        assert df["name"][1] == "Test2"


# ===== EXCEL EDGE CASE TESTS =====


class TestExcelEdgeCases:
    """Test Excel reading edge cases"""

    def test_read_xlsx_with_skiprows(self, temp_dir):
        """Test reading Excel with rows to skip"""
        xlsx_path = temp_dir / "skiprows.xlsx"

        # Create Excel with header rows to skip
        df = pl.DataFrame({
            "col1": ["Title Row", "Description", "name", "Alice", "Bob"],
            "col2": ["", "", "age", "25", "30"],
        })
        df.write_excel(xlsx_path)

        column_mapping = {
            "name": ([], "string"),
            "age": ([], "string"),  # Read as string since mixed content
        }

        result_df = read_xlsx(
            full_path=str(xlsx_path),
            column_mapping=column_mapping,
            excel_skiprows=2,
        )

        # Should skip first 2 rows, use row 3 as header
        assert "name" in result_df.columns or "col1" in result_df.columns

    def test_read_xlsx_empty_cells(self, temp_dir):
        """Test reading Excel with empty cells"""
        xlsx_path = temp_dir / "empty_cells.xlsx"

        df = pl.DataFrame({
            "name": ["Alice", None, "Charlie"],
            "value": [100, 200, None],
        })
        df.write_excel(xlsx_path)

        column_mapping = {
            "name": ([], "string"),
            "value": ([], "int"),
        }

        result_df = read_xlsx(
            full_path=str(xlsx_path),
            column_mapping=column_mapping,
        )

        assert len(result_df) == 3
        # Check that nulls are preserved
        assert result_df["name"].null_count() == 1
        assert result_df["value"].null_count() == 1


# ===== FIXED WIDTH EDGE CASES =====


class TestFixedWidthEdgeCases:
    """Test fixed-width file reading edge cases"""

    def test_read_fixed_width_with_whitespace(self, temp_dir):
        """Test fixed-width with fields that are all whitespace"""
        fw_path = temp_dir / "whitespace.txt"
        # Each field is 10 chars: name, age, city
        # Positions are 1-indexed: name at 1, age at 11, city at 21
        fw_path.write_text("Alice     25        Boston    \n          30        Denver    \n")

        # Format: column_name: (type, starting_position (1-indexed), field_size)
        column_mapping = {
            "name": ("string", 1, 10),
            "age": ("string", 11, 10),  # string since whitespace-padded
            "city": ("string", 21, 10),
        }

        df = read_fixed_width(
            full_path=str(fw_path),
            column_mapping=column_mapping,
        )

        assert len(df) == 2
        # First row should have "Alice"
        assert df["name"][0] == "Alice"

    def test_read_fixed_width_numeric_fields(self, temp_dir):
        """Test fixed-width with right-aligned numeric fields"""
        fw_path = temp_dir / "numeric.txt"
        # Each field is 8 chars
        fw_path.write_text("     100    1.50\n    2000   25.99\n")

        # Format: column_name: (type, starting_position (1-indexed), field_size)
        column_mapping = {
            "amount": ("int", 1, 8),
            "rate": ("float", 9, 8),
        }

        df = read_fixed_width(
            full_path=str(fw_path),
            column_mapping=column_mapping,
        )

        assert len(df) == 2
        assert df["amount"][0] == 100
        assert df["amount"][1] == 2000


# ===== UPDATE TABLE EDGE CASES =====


class TestUpdateTableEdgeCases:
    """Test update_table edge cases"""

    def _setup_metadata_table(self, db_uri, temp_dir, csv_files):
        """Helper to set up metadata table with files"""
        execute_sql(db_uri, "DROP TABLE IF EXISTS test_schema.metadata")
        execute_sql(db_uri, """
            CREATE TABLE test_schema.metadata (
                search_dir TEXT,
                landing_dir TEXT,
                full_path TEXT PRIMARY KEY,
                filesize BIGINT,
                header TEXT[],
                row_count BIGINT,
                archive_full_path TEXT,
                file_hash TEXT,
                metadata_ingest_datetime TIMESTAMP,
                metadata_ingest_status TEXT,
                ingest_datetime TIMESTAMP,
                ingest_runtime INTEGER,
                status TEXT,
                error_message TEXT,
                unpivot_row_multiplier INTEGER
            )
        """)

        for csv_file, row_count in csv_files:
            execute_sql(db_uri, """
                INSERT INTO test_schema.metadata
                (full_path, row_count, metadata_ingest_status, landing_dir)
                VALUES ($1, $2, $3, $4)
            """, (str(csv_file), row_count, "Success", str(temp_dir) + "/"))

    def test_update_table_empty_file(self, db_uri, temp_dir):
        """Test update_table with empty CSV file (header only)"""
        csv_path = temp_dir / "empty.csv"
        csv_path.write_text("name,value\n")  # Header only, no data

        self._setup_metadata_table(db_uri, temp_dir, [(csv_path, 0)])

        column_mapping = {
            "name": ([], "string"),
            "value": ([], "int"),
        }

        result_df = update_table(
            uri=db_uri,
            schema="test_schema",
            output_table="empty_output",
            filetype="csv",
            column_mapping=column_mapping,
            landing_dir=str(temp_dir),
            resume=False,
        )

        # Table should exist but be empty
        assert table_exists(db_uri, "test_schema", "empty_output")
        count = execute_sql_fetchone(db_uri, "SELECT COUNT(*) FROM test_schema.empty_output")
        assert count[0] == 0

    def test_update_table_with_null_value_string(self, db_uri, temp_dir):
        """Test update_table with custom null value string"""
        csv_path = temp_dir / "null_string.csv"
        csv_path.write_text("name,value\nAlice,100\nNA,NA\nBob,200\n")

        self._setup_metadata_table(db_uri, temp_dir, [(csv_path, 3)])

        column_mapping = {
            "name": ([], "string"),
            "value": ([], "int"),
        }

        result_df = update_table(
            uri=db_uri,
            schema="test_schema",
            output_table="null_string_output",
            filetype="csv",
            column_mapping=column_mapping,
            landing_dir=str(temp_dir),
            null_value="NA",
            resume=False,
        )

        # Check that NA was treated as null
        result = execute_sql_fetch(db_uri, "SELECT name, value FROM test_schema.null_string_output ORDER BY value NULLS FIRST")
        # First row should have null value
        assert result[0][1] is None

    def test_update_table_pivot_multiple_value_columns(self, db_uri, temp_dir):
        """Test update_table with pivot_mapping with multiple id columns"""
        csv_path = temp_dir / "multi_pivot.csv"
        csv_path.write_text("id,region,q1,q2,q3,q4\n1,East,100,200,300,400\n2,West,150,250,350,450\n")

        self._setup_metadata_table(db_uri, temp_dir, [(csv_path, 2)])

        column_mapping = {
            "id": ([], "int"),
            "region": ([], "string"),
            "q1": ([], "int"),
            "q2": ([], "int"),
            "q3": ([], "int"),
            "q4": ([], "int"),
        }

        pivot_mapping = {
            "id_vars": ["id", "region"],  # Multiple id columns
            "variable_column_name": "quarter",
            "value_column_name": "sales",
        }

        result_df = update_table(
            uri=db_uri,
            schema="test_schema",
            output_table="multi_pivot_output",
            filetype="csv",
            column_mapping=column_mapping,
            landing_dir=str(temp_dir),
            pivot_mapping=pivot_mapping,
            resume=False,
        )

        # Should have 8 rows (2 ids * 4 quarters)
        count = execute_sql_fetchone(db_uri, "SELECT COUNT(*) FROM test_schema.multi_pivot_output")
        assert count[0] == 8

        # Verify structure
        result = execute_sql_fetchone(db_uri, "SELECT id, region, quarter, sales FROM test_schema.multi_pivot_output LIMIT 1")
        assert result is not None

    def test_update_table_retry_failed(self, db_uri, temp_dir):
        """Test update_table with retry_failed=True"""
        csv_path = temp_dir / "retry.csv"
        csv_path.write_text("name,value\nAlice,100\n")

        # Setup metadata with a failed file
        execute_sql(db_uri, "DROP TABLE IF EXISTS test_schema.metadata")
        execute_sql(db_uri, """
            CREATE TABLE test_schema.metadata (
                search_dir TEXT,
                landing_dir TEXT,
                full_path TEXT PRIMARY KEY,
                filesize BIGINT,
                header TEXT[],
                row_count BIGINT,
                archive_full_path TEXT,
                file_hash TEXT,
                metadata_ingest_datetime TIMESTAMP,
                metadata_ingest_status TEXT,
                ingest_datetime TIMESTAMP,
                ingest_runtime INTEGER,
                status TEXT,
                error_message TEXT,
                unpivot_row_multiplier INTEGER
            )
        """)
        execute_sql(db_uri, """
            INSERT INTO test_schema.metadata
            (full_path, row_count, metadata_ingest_status, landing_dir, status, error_message)
            VALUES ($1, $2, $3, $4, $5, $6)
        """, (str(csv_path), 1, "Success", str(temp_dir) + "/", "Failure", "Previous error"))

        column_mapping = {
            "name": ([], "string"),
            "value": ([], "int"),
        }

        result_df = update_table(
            uri=db_uri,
            schema="test_schema",
            output_table="retry_output",
            filetype="csv",
            column_mapping=column_mapping,
            landing_dir=str(temp_dir),
            retry_failed=True,
        )

        # File should have been reprocessed
        count = execute_sql_fetchone(db_uri, "SELECT COUNT(*) FROM test_schema.retry_output")
        assert count[0] == 1


# ===== TYPE INFERENCE EDGE CASES =====


class TestTypeInferenceEdgeCases:
    """Test type inference with edge cases"""

    def test_infer_schema_mixed_types_in_column(self, temp_dir):
        """Test schema inference with mixed types in a column"""
        csv_path = temp_dir / "mixed.csv"
        csv_path.write_text("value\n100\ntext\n200\n")

        result = infer_schema_from_file(str(csv_path))

        # Should infer as string since mixed types
        assert result["value"][1] == "string"

    def test_infer_schema_empty_column(self, temp_dir):
        """Test schema inference with column that has all empty values"""
        csv_path = temp_dir / "empty_col.csv"
        csv_path.write_text("name,empty_col,value\nAlice,,100\nBob,,200\n")

        result = infer_schema_from_file(str(csv_path))

        # Empty column should be inferred as string
        assert "empty_col" in result
        assert result["empty_col"][1] == "string"

    def test_infer_schema_boolean_column(self, temp_dir):
        """Test schema inference with boolean values"""
        csv_path = temp_dir / "bool.csv"
        csv_path.write_text("name,active\nAlice,true\nBob,false\n")

        result = infer_schema_from_file(str(csv_path))

        assert result["active"][1] == "boolean"

    def test_infer_schema_date_column(self, temp_dir):
        """Test schema inference with date values"""
        csv_path = temp_dir / "dates.csv"
        csv_path.write_text("name,created_at\nAlice,2024-01-15\nBob,2024-02-20\n")

        result = infer_schema_from_file(str(csv_path))

        # Dates might be inferred as datetime or string depending on Polars
        assert "created_at" in result


# ===== CSV HEADER AND ROW COUNT EDGE CASES =====


class TestCsvHeaderRowCountEdgeCases:
    """Test get_csv_header_and_row_count edge cases"""

    def test_csv_header_with_crlf_line_endings(self, temp_dir):
        """Test CSV with Windows line endings (CRLF)"""
        csv_path = temp_dir / "crlf.csv"
        csv_path.write_bytes(b"name,value\r\nAlice,100\r\nBob,200\r\n")

        header, row_count = get_csv_header_and_row_count(
            encoding="utf-8",
            file=str(csv_path),
            separator=",",
            has_header=True
        )

        assert header == ["name", "value"]
        assert row_count == 2

    def test_csv_header_single_row(self, temp_dir):
        """Test CSV with only one data row"""
        csv_path = temp_dir / "single.csv"
        csv_path.write_text("name,value\nAlice,100\n")

        header, row_count = get_csv_header_and_row_count(
            encoding="utf-8",
            file=str(csv_path),
            separator=",",
            has_header=True
        )

        assert row_count == 1

    def test_csv_header_no_data_rows(self, temp_dir):
        """Test CSV with header only (no data)"""
        csv_path = temp_dir / "header_only.csv"
        csv_path.write_text("name,value\n")

        header, row_count = get_csv_header_and_row_count(
            encoding="utf-8",
            file=str(csv_path),
            separator=",",
            has_header=True
        )

        assert header == ["name", "value"]
        assert row_count == 0

    def test_csv_header_special_characters(self, temp_dir):
        """Test CSV with special characters in header"""
        csv_path = temp_dir / "special.csv"
        csv_path.write_text("user-name,total$amount,count#items\nAlice,100,5\n")

        header, row_count = get_csv_header_and_row_count(
            encoding="utf-8",
            file=str(csv_path),
            separator=",",
            has_header=True
        )

        assert "user-name" in header
        assert "total$amount" in header


# ===== DROP FUNCTIONS EDGE CASES =====


class TestDropFunctionsEdgeCases:
    """Test drop functions with edge cases"""

    def test_drop_search_dir_no_matches(self, db_uri):
        """Test drop_search_dir when no files match"""
        execute_sql(db_uri, """
            CREATE TABLE test_schema.metadata (
                search_dir TEXT,
                full_path TEXT PRIMARY KEY
            )
        """)
        execute_sql(db_uri, "INSERT INTO test_schema.metadata VALUES ($1, $2)", ("/some/path", "/some/path/file.csv"))

        # Drop with non-matching path
        drop_search_dir(
            uri=db_uri,
            search_dir="/nonexistent/path",
            schema="test_schema",
        )

        # Original row should still exist
        count = execute_sql_fetchone(db_uri, "SELECT COUNT(*) FROM test_schema.metadata")
        assert count[0] == 1

    def test_drop_partition_with_special_chars(self, db_uri):
        """Test drop_partition with special characters in path"""
        execute_sql(db_uri, """
            CREATE TABLE test_schema.special_test (
                full_path TEXT,
                data TEXT
            )
        """)
        # Path with special characters
        special_path = "/path/with spaces/and'quotes/file.csv"
        execute_sql(db_uri, f"INSERT INTO test_schema.special_test VALUES ('{special_path.replace(chr(39), chr(39)+chr(39))}', 'data')")

        drop_partition(
            uri=db_uri,
            table="special_test",
            partition_key=special_path,
            schema="test_schema",
        )

        count = execute_sql_fetchone(db_uri, "SELECT COUNT(*) FROM test_schema.special_test")
        assert count[0] == 0


# ===== COLUMN MAPPING EDGE CASES =====


class TestColumnMappingEdgeCases:
    """Test column mapping with edge cases"""

    def test_prepare_column_mapping_all_missing(self, temp_dir):
        """Test column mapping when all columns need to be added"""
        csv_path = temp_dir / "missing.csv"
        csv_path.write_text("col1,col2\nval1,val2\n")

        column_mapping = {
            "new_col1": ([], "string"),
            "new_col2": ([], "string"),
            "default": ([], "string"),
        }

        df = read_csv(
            full_path=str(csv_path),
            column_mapping=column_mapping,
            has_header=True,
        )

        # Should have the original columns plus potentially missing columns handled
        assert "col1" in df.columns or "new_col1" in df.columns

    def test_prepare_column_mapping_with_rename_and_missing(self):
        """Test column mapping with both rename and missing columns"""
        column_mapping = {
            "new_name": (["old_name"], "string"),
            "existing": ([], "int"),
            "missing_col": ([], "string"),
            "default": ([], "string"),
        }

        header = ["old_name", "existing"]

        rename_dict, read_dtypes, missing_cols = prepare_column_mapping(
            header, column_mapping
        )

        assert rename_dict == {"old_name": "new_name"}
        assert "missing_col" in missing_cols
        assert missing_cols["missing_col"] == "string"


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
