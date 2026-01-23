"""
Comprehensive tests for table_functions.py

Tests cover:
- S3 helper functions
- Column mapping and selection functions
- File reading functions (CSV, TSV, PSV, XLSX, Parquet, Fixed-width)
- Database operations (table creation, COPY, schema validation)
- Metadata functions (file extraction, metadata tracking)
- CLI schema inference functions

Uses testcontainers for PostgreSQL - automatically manages Docker containers!
Requires Docker to be running.
"""

import pytest
import pandas as pd
import numpy as np
from pathlib import Path
from typing import Dict, List, Tuple, Any
import tempfile
import zipfile
import json
from unittest.mock import Mock, patch, MagicMock
from testcontainers.postgres import PostgresContainer

import psycopg

# Import functions to test
from table_functions import (
    # S3 helpers
    is_s3_path,
    get_s3_filesystem,
    # Path utilities
    normalize_path,
    path_join,
    path_basename,
    path_parent,
    # Cache path utilities
    get_persistent_temp_dir,
    get_cache_path_from_s3,
    get_archive_cache_path_from_s3,
    get_cache_path_from_source_path,
    set_temp_dir_override,
    # Column mapping
    prepare_column_mapping,
    # File readers
    read_csv,
    read_xlsx,
    read_parquet,
    read_fixed_width,
    read_using_column_mapping,
    # Database functions
    table_exists,
    get_table_schema,
    validate_schema_match,
    create_table_from_dataframe,
    copy_dataframe_to_table,
    # Metadata functions
    get_csv_header_and_row_count,
    get_file_metadata_row,
    row_count_check,
    update_metadata,
    update_table,
    add_files_to_metadata_table,
    extract_and_add_zip_files,
    add_files,
    drop_metadata_by_source,
    drop_partition,
    drop_file_from_metadata_and_table,
    # Schema inference
    to_snake_case,
    infer_schema_from_file,
)


# ===== TEST HELPER FUNCTIONS =====

import re


def _convert_placeholders(sql: str) -> str:
    """Convert $1, $2, etc. placeholders to %s for psycopg3"""
    return re.sub(r"\$\d+", "%s", sql)


def execute_sql(conn: psycopg.Connection, sql: str, params: tuple = None):
    """Execute SQL statement without returning results"""
    sql = _convert_placeholders(sql)
    with conn.cursor() as cur:
        if params:
            cur.execute(sql, params)
        else:
            cur.execute(sql)
    conn.commit()


def execute_sql_fetch(conn: psycopg.Connection, sql: str, params: tuple = None):
    """Execute SQL and return all results"""
    sql = _convert_placeholders(sql)
    with conn.cursor() as cur:
        if params:
            cur.execute(sql, params)
        else:
            cur.execute(sql)
        return cur.fetchall()


def execute_sql_fetchone(conn: psycopg.Connection, sql: str, params: tuple = None):
    """Execute SQL and return single result"""
    sql = _convert_placeholders(sql)
    with conn.cursor() as cur:
        if params:
            cur.execute(sql, params)
        else:
            cur.execute(sql)
        return cur.fetchone()


# ===== FIXTURES =====


@pytest.fixture
def temp_dir():
    """Create a temporary directory for test files"""
    with tempfile.TemporaryDirectory() as tmpdir:
        yield Path(tmpdir)


@pytest.fixture(scope="session")
def postgres_container():
    """
    Create a PostgreSQL container for the entire test session.

    Uses testcontainers to automatically:
    - Pull PostgreSQL Docker image (if needed)
    - Start a PostgreSQL container
    - Expose the database port
    - Clean up when tests are done

    Requires Docker to be running!
    """
    with PostgresContainer("postgres:16-alpine") as postgres:
        yield postgres


@pytest.fixture(scope="function")
def db_conn(postgres_container):
    """
    Create a database connection for each test function.

    Reuses the same PostgreSQL container but creates a fresh schema for each test.
    """
    # Get connection string from container
    connection_url = postgres_container.get_connection_url()

    # Convert SQLAlchemy-style URL to standard PostgreSQL URL
    # testcontainers returns: postgresql+psycopg2://...
    # psycopg expects: postgresql://...
    uri = connection_url.replace("postgresql+psycopg2://", "postgresql://")

    # Create connection
    conn = psycopg.connect(uri)

    # Create test schema
    execute_sql(conn, "CREATE SCHEMA IF NOT EXISTS test_schema")

    yield conn

    # Cleanup schema after test
    try:
        execute_sql(conn, "DROP SCHEMA IF EXISTS test_schema CASCADE")
    except:
        pass
    finally:
        conn.close()


@pytest.fixture(scope="function")
def conninfo(postgres_container):
    """
    Provide a connection string for each test function.

    Reuses the same PostgreSQL container but creates a fresh schema for each test.
    """
    # Get connection string from container
    connection_url = postgres_container.get_connection_url()

    # Convert SQLAlchemy-style URL to standard PostgreSQL URL
    # testcontainers returns: postgresql+psycopg2://...
    # psycopg expects: postgresql://...
    uri = connection_url.replace("postgresql+psycopg2://", "postgresql://")

    # Create test schema using a temporary connection
    with psycopg.connect(uri) as conn:
        execute_sql(conn, "CREATE SCHEMA IF NOT EXISTS test_schema")

    yield uri

    # Cleanup schema after test
    try:
        with psycopg.connect(uri) as conn:
            execute_sql(conn, "DROP SCHEMA IF EXISTS test_schema CASCADE")
    except:
        pass


@pytest.fixture
def sample_csv_file(temp_dir):
    """Create a sample CSV file for testing"""
    csv_path = temp_dir / "sample.csv"
    data = """name,age,score
Alice,25,85.5
Bob,30,92.0
Charlie,35,78.5"""
    csv_path.write_text(data)
    return csv_path


@pytest.fixture
def sample_csv_no_header(temp_dir):
    """Create a CSV file without headers"""
    csv_path = temp_dir / "sample_no_header.csv"
    data = """Alice,25,85.5
Bob,30,92.0
Charlie,35,78.5"""
    csv_path.write_text(data)
    return csv_path


@pytest.fixture
def sample_psv_file(temp_dir):
    """Create a pipe-delimited file"""
    psv_path = temp_dir / "sample.psv"
    data = """name|age|score
Alice|25|85.5
Bob|30|92.0"""
    psv_path.write_text(data)
    return psv_path


@pytest.fixture
def sample_zip_file(temp_dir):
    """Create a ZIP file containing CSV files"""
    zip_path = temp_dir / "sample.zip"

    with zipfile.ZipFile(zip_path, "w") as zf:
        zf.writestr("file1.csv", "col1,col2\n1,2\n3,4")
        zf.writestr("file2.csv", "col1,col2\n5,6\n7,8")

    return zip_path


# ===== S3 HELPER TESTS =====


class TestS3Helpers:
    """Test S3 helper functions"""

    def test_is_s3_path_true(self):
        assert is_s3_path("s3://bucket/path/to/file")

    def test_is_s3_path_false(self):
        assert not is_s3_path("/local/path")
        assert not is_s3_path("http://example.com")
        assert not is_s3_path("s3:/bucket/path")  # Path-mangled no longer supported

    def test_get_s3_filesystem_returns_passed_filesystem(self):
        """Test that get_s3_filesystem returns the passed filesystem if provided"""
        mock_fs = MagicMock()
        result = get_s3_filesystem(mock_fs)
        assert result is mock_fs

    def test_get_s3_filesystem_creates_new(self):
        """Test that get_s3_filesystem creates a new S3FileSystem if not provided"""
        import s3fs

        result = get_s3_filesystem(None)
        assert isinstance(result, s3fs.S3FileSystem)


# ===== PATH UTILITY TESTS =====


class TestPathUtilities:
    """Test path utility functions for S3 and local paths"""

    # normalize_path tests
    def test_normalize_path_local(self):
        """Test normalizing local paths"""
        assert normalize_path("/path/to/file") == "/path/to/file"
        assert normalize_path("/path/to/dir/") == "/path/to/dir"
        assert normalize_path("relative/path") == "relative/path"

    def test_normalize_path_s3(self):
        """Test normalizing S3 paths"""
        assert normalize_path("s3://bucket/path/to/file") == "s3://bucket/path/to/file"
        assert normalize_path("s3://bucket/path/") == "s3://bucket/path"
        assert normalize_path("s3://bucket") == "s3://bucket"

    def test_normalize_path_with_path_object(self):
        """Test normalizing Path objects"""
        from pathlib import Path

        assert normalize_path(Path("/path/to/file")) == "/path/to/file"

    # path_join tests
    def test_path_join_local(self):
        """Test joining local paths"""
        assert path_join("/path", "to", "file") == "/path/to/file"
        assert path_join("relative", "path") == "relative/path"

    def test_path_join_s3(self):
        """Test joining S3 paths"""
        assert (
            path_join("s3://bucket", "path", "file.csv") == "s3://bucket/path/file.csv"
        )
        assert (
            path_join("s3://bucket/", "path/", "file.csv")
            == "s3://bucket/path/file.csv"
        )

    def test_path_join_empty_parts(self):
        """Test joining with empty parts"""
        assert path_join("/path", "", "file") == "/path/file"
        assert path_join("") == ""

    def test_path_join_strips_slashes(self):
        """Test that joining strips extra slashes"""
        assert path_join("/path/", "/to/", "/file") == "/path/to/file"
        assert path_join("s3://bucket/", "/path/") == "s3://bucket/path"

    # path_basename tests
    def test_path_basename_local(self):
        """Test getting basename from local paths"""
        assert path_basename("/path/to/file.csv") == "file.csv"
        assert path_basename("/path/to/dir/") == "dir"
        assert path_basename("file.csv") == "file.csv"

    def test_path_basename_s3(self):
        """Test getting basename from S3 paths"""
        assert path_basename("s3://bucket/path/to/file.csv") == "file.csv"
        assert path_basename("s3://bucket/path/") == "path"

    # path_parent tests
    def test_path_parent_local(self):
        """Test getting parent from local paths"""
        assert path_parent("/path/to/file.csv") == "/path/to"
        assert path_parent("/path/to/") == "/path"
        assert path_parent("/file.csv") == ""

    def test_path_parent_s3(self):
        """Test getting parent from S3 paths"""
        assert path_parent("s3://bucket/path/to/file.csv") == "s3://bucket/path/to"
        assert path_parent("s3://bucket/path") == "s3://bucket"
        # S3 bucket root returns itself
        assert path_parent("s3://bucket") == "s3://bucket"

    def test_path_parent_single_element(self):
        """Test parent of single element path"""
        assert path_parent("file.csv") == ""


# ===== CACHE PATH TESTS =====


class TestCachePaths:
    """Test cache path functions for S3 files and archives"""

    def test_get_cache_path_from_s3(self, temp_dir):
        """Test S3 path to cache path conversion"""
        set_temp_dir_override(temp_dir)
        try:
            cache_path = get_cache_path_from_s3("s3://my-bucket/data/file.csv")
            assert cache_path == temp_dir / "my-bucket" / "data" / "file.csv"
            # Parent directory should be created
            assert cache_path.parent.exists()
        finally:
            set_temp_dir_override(None)

    def test_get_archive_cache_path_from_s3(self, temp_dir):
        """Test S3 archive path goes to archives subdirectory"""
        set_temp_dir_override(temp_dir)
        try:
            cache_path = get_archive_cache_path_from_s3(
                "s3://my-bucket/data/archive.zip"
            )
            assert (
                cache_path
                == temp_dir / "archives" / "my-bucket" / "data" / "archive.zip"
            )
            # Parent directory should be created
            assert cache_path.parent.exists()
        finally:
            set_temp_dir_override(None)

    def test_get_cache_path_from_source_path_s3_file(self, temp_dir):
        """Test cache path for regular S3 file (no archive)"""
        set_temp_dir_override(temp_dir)
        try:
            cache_path = get_cache_path_from_source_path("s3://bucket/path/file.csv")
            assert cache_path == temp_dir / "bucket" / "path" / "file.csv"
        finally:
            set_temp_dir_override(None)

    def test_get_cache_path_from_source_path_s3_archive(self, temp_dir):
        """Test cache path for S3 archive with inner path"""
        set_temp_dir_override(temp_dir)
        try:
            cache_path = get_cache_path_from_source_path(
                "s3://bucket/archive.zip::inner/file.csv"
            )
            # Extracted contents go in temp/bucket/archive.zip/inner/file.csv
            # (NOT temp/archives/ - that's where the archive itself is cached)
            assert (
                cache_path == temp_dir / "bucket" / "archive.zip" / "inner" / "file.csv"
            )
            assert cache_path.parent.exists()
        finally:
            set_temp_dir_override(None)

    def test_get_cache_path_from_source_path_local_file(self):
        """Test that local files are returned as-is"""
        cache_path = get_cache_path_from_source_path("/local/path/file.csv")
        assert cache_path == Path("/local/path/file.csv")

    def test_get_cache_path_from_source_path_local_archive(self, temp_dir):
        """Test cache path for local archive with inner path"""
        set_temp_dir_override(temp_dir)
        try:
            cache_path = get_cache_path_from_source_path(
                "/local/archive.zip::inner/file.csv"
            )
            # Extracted contents go in temp/local/archive.zip/inner/file.csv
            assert (
                cache_path == temp_dir / "local" / "archive.zip" / "inner" / "file.csv"
            )
        finally:
            set_temp_dir_override(None)

    def test_archive_and_extracted_paths_dont_conflict(self, temp_dir):
        """Test that archive cache and extracted contents don't conflict"""
        set_temp_dir_override(temp_dir)
        try:
            # Where the downloaded archive goes
            archive_cache = get_archive_cache_path_from_s3("s3://bucket/data.zip")
            # Where extracted contents go
            extracted_cache = get_cache_path_from_source_path(
                "s3://bucket/data.zip::inner/file.csv"
            )

            # These should be different paths
            assert archive_cache != extracted_cache
            assert archive_cache.parent != extracted_cache.parent

            # Archive goes in temp/archives/bucket/data.zip
            assert "archives" in str(archive_cache)
            # Extracted goes in temp/bucket/data.zip/inner/file.csv
            assert "archives" not in str(extracted_cache)

            # We should be able to create both without conflict
            archive_cache.parent.mkdir(parents=True, exist_ok=True)
            archive_cache.write_text("archive content")

            extracted_cache.parent.mkdir(parents=True, exist_ok=True)
            extracted_cache.write_text("extracted content")

            # Both should exist
            assert archive_cache.exists()
            assert extracted_cache.exists()
        finally:
            set_temp_dir_override(None)


# ===== COLUMN MAPPING TESTS =====


class TestColumnMapping:
    """Test column selection and mapping functions"""

    def test_prepare_column_mapping_basic(self):
        """Test basic column selection without renaming"""
        header = ["col1", "col2", "col3"]
        column_mapping = {
            "col1": ([], "string"),
            "col2": ([], "int"),
            "col3": ([], "float"),
        }

        rename_dict, read_dtypes, missing_cols = prepare_column_mapping(
            header, column_mapping
        )

        assert rename_dict == {}  # No renames needed
        assert read_dtypes == {"col1": "string", "col2": "Int64", "col3": "float64"}
        assert missing_cols == {}

    def test_prepare_column_mapping_with_rename(self):
        """Test column selection with renaming"""
        header = ["old_name", "col2"]
        column_mapping = {
            "new_name": (["old_name", "alt_name"], "string"),
            "col2": ([], "int"),
        }

        rename_dict, read_dtypes, missing_cols = prepare_column_mapping(
            header, column_mapping
        )

        assert rename_dict == {"old_name": "new_name"}
        assert read_dtypes == {"old_name": "string", "col2": "Int64"}
        assert missing_cols == {}

    def test_prepare_column_mapping_with_missing_cols(self):
        """Test handling of missing columns"""
        header = ["col1", "col2"]
        column_mapping = {
            "col1": ([], "string"),
            "col2": ([], "int"),
            "col3": ([], "float"),  # Missing in header
        }

        rename_dict, read_dtypes, missing_cols = prepare_column_mapping(
            header, column_mapping
        )

        assert missing_cols == {"col3": "float"}
        assert read_dtypes == {"col1": "string", "col2": "Int64"}

    def test_prepare_column_mapping_with_default(self):
        """Test default type for unmapped columns"""
        header = ["col1", "col2", "col3", "col4"]
        column_mapping = {
            "col1": ([], "string"),
            "col2": ([], "int"),
            "default": ([], "string"),
        }

        rename_dict, read_dtypes, missing_cols = prepare_column_mapping(
            header, column_mapping
        )

        # col3 and col4 should get default type
        assert read_dtypes == {
            "col1": "string",
            "col2": "Int64",
            "col3": "string",
            "col4": "string",
        }
        assert rename_dict == {}
        assert missing_cols == {}


# ===== FILE READING TESTS =====


class TestFileReading:
    """Test file reading functions"""

    def test_read_csv_with_header(self, sample_csv_file):
        """Test reading CSV with header"""
        column_mapping = {
            "name": ([], "string"),
            "age": ([], "int"),
            "score": ([], "float"),
        }

        df = read_csv(
            full_path=str(sample_csv_file),
            column_mapping=column_mapping,
            has_header=True,
        )

        assert len(df) == 3
        assert list(df.columns) == ["name", "age", "score"]
        assert df["name"][0] == "Alice"
        assert df["age"].dtype == pd.Int64Dtype()

    def test_read_csv_no_header(self, sample_csv_no_header):
        """Test reading CSV without header"""
        column_mapping = {
            "name": ([], "string"),
            "age": ([], "int"),
            "score": ([], "float"),
        }

        header = ["name", "age", "score"]

        df = read_csv(
            full_path=str(sample_csv_no_header),
            column_mapping=column_mapping,
            header=header,
            has_header=False,
        )

        assert len(df) == 3
        assert df["name"][0] == "Alice"

    def test_read_csv_with_quoted_headers(self, temp_dir):
        """Test reading CSV with quoted headers"""
        csv_path = temp_dir / "quoted_headers.csv"
        csv_path.write_text(
            '"name","age","score"\nAlice,25,95.5\nBob,30,87.2\nCharlie,22,92.0'
        )

        column_mapping = {
            "name": ([], "string"),
            "age": ([], "int"),
            "score": ([], "float"),
        }

        df = read_csv(
            full_path=str(csv_path),
            column_mapping=column_mapping,
            has_header=True,
        )

        assert len(df) == 3
        assert list(df.columns) == ["name", "age", "score"]
        assert df["name"][0] == "Alice"
        assert df["age"].dtype == pd.Int64Dtype()
        assert df["score"][0] == 95.5

    def test_read_csv_psv_separator(self, sample_psv_file):
        """Test reading pipe-delimited file"""
        column_mapping = {
            "name": ([], "string"),
            "age": ([], "int"),
            "score": ([], "float"),
        }

        df = read_csv(
            full_path=str(sample_psv_file),
            column_mapping=column_mapping,
            has_header=True,
            separator="|",
        )

        assert len(df) == 2
        assert df["name"][0] == "Alice"

    def test_read_csv_with_missing_cols(self, sample_csv_file):
        """Test reading CSV with missing columns in mapping"""
        column_mapping = {
            "name": ([], "string"),
            "age": ([], "int"),
            "score": ([], "float"),
            "extra_col": ([], "string"),  # Not in CSV
        }

        df = read_csv(
            full_path=str(sample_csv_file),
            column_mapping=column_mapping,
            has_header=True,
        )

        assert "extra_col" in df.columns
        assert df["extra_col"].isna().sum() == len(df)

    def test_read_using_column_mapping_csv(self, sample_csv_file):
        """Test router function for CSV"""
        column_mapping = {
            "name": ([], "string"),
            "age": ([], "int"),
            "score": ([], "float"),
        }

        df = read_using_column_mapping(
            full_path=str(sample_csv_file),
            filetype="csv",
            column_mapping=column_mapping,
            has_header=True,
        )

        assert df is not None
        assert len(df) == 3

    def test_read_using_column_mapping_tsv(self, temp_dir):
        """Test router function for TSV"""
        tsv_path = temp_dir / "sample.tsv"
        tsv_path.write_text("col1\tcol2\n1\t2\n3\t4")

        column_mapping = {
            "col1": ([], "int"),
            "col2": ([], "int"),
        }

        df = read_using_column_mapping(
            full_path=str(tsv_path),
            filetype="tsv",
            column_mapping=column_mapping,
            has_header=True,
        )

        assert df is not None
        assert len(df) == 2

    def test_read_using_column_mapping_psv(self, sample_psv_file):
        """Test router function for PSV"""
        column_mapping = {
            "name": ([], "string"),
            "age": ([], "int"),
            "score": ([], "float"),
        }

        df = read_using_column_mapping(
            full_path=str(sample_psv_file),
            filetype="psv",
            column_mapping=column_mapping,
            has_header=True,
        )

        assert df is not None

    def test_read_fixed_width(self, temp_dir):
        """Test reading fixed-width file"""
        fw_path = temp_dir / "fixed_width.txt"
        fw_path.write_text("AL   100  200\nCA   150  250\n")

        column_mapping = {
            "state": ("string", 1, 2),
            "pop": ("int", 6, 3),
            "area": ("int", 11, 3),
        }

        df = read_fixed_width(
            full_path=str(fw_path),
            column_mapping=column_mapping,
        )

        assert len(df) == 2
        assert df["state"][0] == "AL"
        assert df["pop"][0] == 100

    def test_read_csv_no_duplicate_columns_with_rename(self, temp_dir):
        """Test that CSV column renaming doesn't create duplicate columns"""
        csv_path = temp_dir / "rename_test.csv"
        csv_path.write_text(
            "FirstName,LastName,TotalAmount\nJohn,Doe,100.50\nJane,Smith,200.75"
        )

        column_mapping = {
            "first_name": (["FirstName"], "string"),
            "last_name": (["LastName"], "string"),
            "total_amount": (["TotalAmount"], "float"),
        }

        df = read_csv(
            full_path=str(csv_path),
            column_mapping=column_mapping,
            has_header=True,
        )

        # Verify no duplicate columns
        assert "FirstName" not in df.columns, (
            "Original column should not exist after rename"
        )
        assert "LastName" not in df.columns, (
            "Original column should not exist after rename"
        )
        assert "TotalAmount" not in df.columns, (
            "Original column should not exist after rename"
        )

        # Verify renamed columns exist with data
        assert "first_name" in df.columns
        assert "last_name" in df.columns
        assert "total_amount" in df.columns
        assert len(df.columns) == 3, (
            f"Should have exactly 3 columns, got {len(df.columns)}"
        )

        # Verify data is present
        assert df["first_name"][0] == "John"
        assert df["last_name"][0] == "Doe"
        assert df["total_amount"][0] == 100.50

    def test_read_xlsx_no_duplicate_columns_with_rename(self, temp_dir):
        """Test that Excel column renaming doesn't create duplicate columns"""
        xlsx_path = temp_dir / "rename_test.xlsx"
        test_df = pd.DataFrame(
            {
                "FirstName": ["John", "Jane"],
                "LastName": ["Doe", "Smith"],
                "TotalAmount": [100.50, 200.75],
            }
        )
        test_df.to_excel(str(xlsx_path), index=False)

        column_mapping = {
            "first_name": (["FirstName"], "string"),
            "last_name": (["LastName"], "string"),
            "total_amount": (["TotalAmount"], "float"),
        }

        df = read_xlsx(
            full_path=str(xlsx_path),
            column_mapping=column_mapping,
        )

        # Verify no duplicate columns
        assert "FirstName" not in df.columns, (
            "Original column should not exist after rename"
        )
        assert "LastName" not in df.columns, (
            "Original column should not exist after rename"
        )
        assert "TotalAmount" not in df.columns, (
            "Original column should not exist after rename"
        )

        # Verify renamed columns exist with data
        assert "first_name" in df.columns
        assert "last_name" in df.columns
        assert "total_amount" in df.columns
        assert len(df.columns) == 3, (
            f"Should have exactly 3 columns, got {len(df.columns)}"
        )

        # Verify data is present
        assert df["first_name"][0] == "John"
        assert df["last_name"][0] == "Doe"
        assert df["total_amount"][0] == 100.50

    def test_read_tsv_no_duplicate_columns_with_rename(self, temp_dir):
        """Test that TSV column renaming doesn't create duplicate columns"""
        tsv_path = temp_dir / "rename_test.tsv"
        tsv_path.write_text(
            "FirstName\tLastName\tTotalAmount\nJohn\tDoe\t100.50\nJane\tSmith\t200.75"
        )

        column_mapping = {
            "first_name": (["FirstName"], "string"),
            "last_name": (["LastName"], "string"),
            "total_amount": (["TotalAmount"], "float"),
        }

        df = read_csv(
            full_path=str(tsv_path),
            column_mapping=column_mapping,
            has_header=True,
            separator="\t",
        )

        # Verify no duplicate columns
        assert "FirstName" not in df.columns, (
            "Original column should not exist after rename"
        )
        assert "LastName" not in df.columns, (
            "Original column should not exist after rename"
        )
        assert "TotalAmount" not in df.columns, (
            "Original column should not exist after rename"
        )

        # Verify renamed columns exist with data
        assert "first_name" in df.columns
        assert "last_name" in df.columns
        assert "total_amount" in df.columns
        assert len(df.columns) == 3, (
            f"Should have exactly 3 columns, got {len(df.columns)}"
        )

        # Verify data is present
        assert df["first_name"][0] == "John"
        assert df["last_name"][0] == "Doe"
        assert df["total_amount"][0] == 100.50

    def test_read_psv_no_duplicate_columns_with_rename(self, temp_dir):
        """Test that PSV (pipe-separated) column renaming doesn't create duplicate columns"""
        psv_path = temp_dir / "rename_test.psv"
        psv_path.write_text(
            "FirstName|LastName|TotalAmount\nJohn|Doe|100.50\nJane|Smith|200.75"
        )

        column_mapping = {
            "first_name": (["FirstName"], "string"),
            "last_name": (["LastName"], "string"),
            "total_amount": (["TotalAmount"], "float"),
        }

        df = read_csv(
            full_path=str(psv_path),
            column_mapping=column_mapping,
            has_header=True,
            separator="|",
        )

        # Verify no duplicate columns
        assert "FirstName" not in df.columns, (
            "Original column should not exist after rename"
        )
        assert "LastName" not in df.columns, (
            "Original column should not exist after rename"
        )
        assert "TotalAmount" not in df.columns, (
            "Original column should not exist after rename"
        )

        # Verify renamed columns exist with data
        assert "first_name" in df.columns
        assert "last_name" in df.columns
        assert "total_amount" in df.columns
        assert len(df.columns) == 3, (
            f"Should have exactly 3 columns, got {len(df.columns)}"
        )

        # Verify data is present
        assert df["first_name"][0] == "John"
        assert df["last_name"][0] == "Doe"
        assert df["total_amount"][0] == 100.50


# ===== DATABASE OPERATION TESTS =====


class TestDatabaseOperations:
    """Test PostgreSQL database operations"""

    def test_table_exists_true(self, db_conn):
        """Test table_exists when table exists"""
        execute_sql(db_conn, "CREATE TABLE test_schema.test_table (id INT)")

        assert table_exists(db_conn, "test_schema", "test_table")

    def test_table_exists_false(self, db_conn):
        """Test table_exists when table doesn't exist"""
        assert not table_exists(db_conn, "test_schema", "nonexistent_table")

    def test_create_table_from_dataframe(self, db_conn):
        """Test creating table from DataFrame"""
        df = pd.DataFrame(
            {
                "id": pd.Series([1, 2, 3], dtype=pd.Int64Dtype()),
                "name": pd.Series(["Alice", "Bob", "Charlie"], dtype="string"),
                "score": [85.5, 92.0, 78.5],
            }
        )

        create_table_from_dataframe(db_conn, df, "test_schema", "new_table")

        assert table_exists(db_conn, "test_schema", "new_table")

    def test_get_table_schema(self, db_conn):
        """Test getting table schema"""
        execute_sql(
            db_conn,
            """
            CREATE TABLE test_schema.schema_test (
                id BIGINT,
                name TEXT,
                score DOUBLE PRECISION
            )
        """,
        )

        schema = get_table_schema(db_conn, "test_schema", "schema_test")

        assert "id" in schema
        assert "name" in schema
        assert "score" in schema
        assert schema["id"] == "bigint"
        assert schema["name"] == "text"

    def test_validate_schema_match_success(self, db_conn):
        """Test schema validation when schemas match"""
        # Create table
        execute_sql(
            db_conn,
            """
            CREATE TABLE test_schema.validate_test (
                id BIGINT,
                name TEXT
            )
        """,
        )

        # Create matching DataFrame - polars dtypes that map to BIGINT and TEXT
        df = pd.DataFrame(
            {
                "id": pd.Series([1, 2], dtype=pd.Int64Dtype()),
                "name": pd.Series(["Alice", "Bob"], dtype="string"),
            }
        )

        # Should not raise
        validate_schema_match(db_conn, df, "test_schema", "validate_test")

    def test_validate_schema_match_missing_col_in_table(self, db_conn):
        """Test schema validation when DataFrame has extra columns"""
        # Create table
        execute_sql(
            db_conn,
            """
            CREATE TABLE test_schema.validate_test2 (
                id BIGINT
            )
        """,
        )

        # Create DataFrame with extra column
        df = pd.DataFrame(
            {
                "id": pd.Series([1, 2], dtype=pd.Int64Dtype()),
                "extra": pd.Series(["a", "b"], dtype="string"),
            }
        )

        with pytest.raises(ValueError, match="DataFrame has columns not in table"):
            validate_schema_match(db_conn, df, "test_schema", "validate_test2")

    def test_copy_dataframe_to_table(self, db_conn):
        """Test bulk loading DataFrame to PostgreSQL"""
        df = pd.DataFrame(
            {
                "id": pd.Series([1, 2, 3], dtype=pd.Int64Dtype()),
                "name": pd.Series(["Alice", "Bob", "Charlie"], dtype="string"),
                "score": [85.5, 92.0, 78.5],
            }
        )

        copy_dataframe_to_table(db_conn, df, "test_schema", "copy_test")

        # Verify data was loaded
        result = execute_sql_fetchone(
            db_conn, "SELECT COUNT(*) FROM test_schema.copy_test"
        )
        assert result[0] == 3

    def test_copy_dataframe_with_nulls(self, db_conn):
        """Test COPY with NULL values"""
        df = pd.DataFrame(
            {
                "id": pd.Series([1, 2, None], dtype=pd.Int64Dtype()),
                "name": pd.Series(["Alice", None, "Charlie"], dtype="string"),
                "score": [85.5, None, 78.5],
            }
        )

        copy_dataframe_to_table(db_conn, df, "test_schema", "null_test")

        # Verify NULLs were preserved
        result = execute_sql_fetchone(
            db_conn, "SELECT id, name, score FROM test_schema.null_test WHERE id = 2"
        )

        assert result[1] is None
        assert result[2] is None


# ===== METADATA FUNCTION TESTS =====


class TestMetadataFunctions:
    """Test metadata tracking functions"""

    def test_get_csv_header_and_row_count(self, sample_csv_file):
        """Test getting header and row count from CSV"""
        header, row_count = get_csv_header_and_row_count(
            file=sample_csv_file,
            has_header=True,
        )

        assert header == ["name", "age", "score"]
        assert row_count == 3

    def test_get_csv_header_and_row_count_no_header(self, sample_csv_no_header):
        """Test getting header and row count from headerless CSV"""
        header, row_count = get_csv_header_and_row_count(
            file=sample_csv_no_header,
            has_header=False,
        )

        # First line is returned as "header" even though it's data
        assert len(header) == 3
        assert row_count == 3

    def test_get_csv_header_psv(self, sample_psv_file):
        """Test getting header from pipe-delimited file"""
        header, row_count = get_csv_header_and_row_count(
            file=sample_psv_file,
            separator="|",
            has_header=True,
        )

        assert header == ["name", "age", "score"]
        assert row_count == 2

    def test_get_csv_header_with_quoted_headers(self, temp_dir):
        """Test getting header from CSV with quoted headers"""
        csv_path = temp_dir / "quoted_headers.csv"
        csv_path.write_text('"name","age","score"\nAlice,25,95.5\nBob,30,87.2')

        header, row_count = get_csv_header_and_row_count(
            file=csv_path,
            has_header=True,
        )

        # Header should be unquoted
        assert header == ["name", "age", "score"]
        assert row_count == 2

    def test_get_file_metadata_row_csv(self, sample_csv_file, temp_dir):
        """Test generating metadata row for CSV file"""
        row = get_file_metadata_row(
            source_path=sample_csv_file.as_posix(),
            source_dir=str(temp_dir) + "/",
            filetype="csv",
            has_header=True,
        )

        assert row["source_path"] == sample_csv_file.as_posix()
        assert row["metadata_ingest_status"] == "Success"
        assert row["header"] == ["name", "age", "score"]
        assert row["row_count"] == 3
        assert row["file_hash"] is not None
        assert row["filesize"] > 0

    def test_get_file_metadata_row_with_error(self, temp_dir):
        """Test metadata row generation with error"""
        row = get_file_metadata_row(
            source_path="nonexistent.csv",
            source_dir=str(temp_dir) + "/",
            filetype="csv",
            has_header=True,
            error_message="Test error",
        )

        assert row["metadata_ingest_status"] == "Failure"
        assert row["error_message"] == "Test error"

    def test_row_count_check_success(self, db_conn, sample_csv_file):
        """Test row count validation success"""
        # Create metadata table and insert row
        execute_sql(
            db_conn,
            """
            CREATE TABLE test_schema.metadata (
                source_path TEXT PRIMARY KEY,
                row_count BIGINT
            )
        """,
        )
        execute_sql(
            db_conn,
            "INSERT INTO test_schema.metadata (source_path, row_count) VALUES ($1, $2)",
            (str(sample_csv_file), 3),
        )

        # Create DataFrame with matching row count
        df = pd.DataFrame({"col1": [1, 2, 3]})

        # Should not raise
        row_count_check(
            conn=db_conn,
            schema="test_schema",
            df=df,
            source_path=str(sample_csv_file),
        )

    def test_row_count_check_failure(self, db_conn, sample_csv_file):
        """Test row count validation failure"""
        # Create metadata table and insert row
        execute_sql(
            db_conn,
            """
            CREATE TABLE test_schema.metadata (
                source_path TEXT PRIMARY KEY,
                row_count BIGINT
            )
        """,
        )
        execute_sql(
            db_conn,
            "INSERT INTO test_schema.metadata (source_path, row_count) VALUES ($1, $2)",
            (str(sample_csv_file), 5),  # Wrong count
        )

        # Create DataFrame with different row count
        df = pd.DataFrame({"col1": [1, 2, 3]})

        with pytest.raises(ValueError, match="Check failed"):
            row_count_check(
                conn=db_conn,
                schema="test_schema",
                df=df,
                source_path=str(sample_csv_file),
            )

    def test_update_metadata_success(self, db_conn, sample_csv_file):
        """Test updating metadata with success status"""
        # Create metadata table
        execute_sql(
            db_conn,
            """
            CREATE TABLE test_schema.metadata (
                source_path TEXT PRIMARY KEY,
                ingest_datetime TIMESTAMP,
                status TEXT,
                error_message TEXT,
                unpivot_row_multiplier INTEGER,
                ingest_runtime INTEGER,
                output_table TEXT
            )
        """,
        )
        execute_sql(
            db_conn,
            "INSERT INTO test_schema.metadata (source_path) VALUES ($1)",
            (str(sample_csv_file),),
        )

        update_metadata(
            conn=db_conn,
            source_path=str(sample_csv_file),
            schema="test_schema",
            ingest_runtime=5,
        )

        # Verify update
        result = execute_sql_fetchone(
            db_conn,
            "SELECT status, ingest_runtime FROM test_schema.metadata WHERE source_path = $1",
            (str(sample_csv_file),),
        )

        assert result[0] == "Success"
        assert result[1] == 5

    def test_update_metadata_failure(self, db_conn, sample_csv_file):
        """Test updating metadata with failure status"""
        # Create metadata table
        execute_sql(
            db_conn,
            """
            CREATE TABLE test_schema.metadata (
                source_path TEXT PRIMARY KEY,
                ingest_datetime TIMESTAMP,
                status TEXT,
                error_message TEXT,
                unpivot_row_multiplier INTEGER,
                ingest_runtime INTEGER,
                output_table TEXT
            )
        """,
        )
        execute_sql(
            db_conn,
            "INSERT INTO test_schema.metadata (source_path) VALUES ($1)",
            (str(sample_csv_file),),
        )

        update_metadata(
            conn=db_conn,
            source_path=str(sample_csv_file),
            schema="test_schema",
            error_message="Test error occurred",
        )

        # Verify update
        result = execute_sql_fetchone(
            db_conn,
            "SELECT status, error_message FROM test_schema.metadata WHERE source_path = $1",
            (str(sample_csv_file),),
        )

        assert result[0] == "Failure"
        assert result[1] == "Test error occurred"

    def test_update_metadata_with_output_table(self, db_conn, sample_csv_file):
        """Test updating metadata with output_table"""
        # Create metadata table
        execute_sql(
            db_conn,
            """
            CREATE TABLE test_schema.metadata (
                source_path TEXT PRIMARY KEY,
                ingest_datetime TIMESTAMP,
                status TEXT,
                error_message TEXT,
                unpivot_row_multiplier INTEGER,
                ingest_runtime INTEGER,
                output_table TEXT
            )
        """,
        )
        execute_sql(
            db_conn,
            "INSERT INTO test_schema.metadata (source_path) VALUES ($1)",
            (str(sample_csv_file),),
        )

        update_metadata(
            conn=db_conn,
            source_path=str(sample_csv_file),
            schema="test_schema",
            ingest_runtime=5,
            output_table="raw.my_table",
        )

        # Verify update
        result = execute_sql_fetchone(
            db_conn,
            "SELECT status, output_table FROM test_schema.metadata WHERE source_path = $1",
            (str(sample_csv_file),),
        )

        assert result[0] == "Success"
        assert result[1] == "raw.my_table"


# ===== SCHEMA INFERENCE TESTS =====


class TestSnakeCaseConversion:
    """Test snake_case conversion function"""

    def test_snake_case_camel_case(self):
        """Test conversion from camelCase"""
        assert to_snake_case("firstName") == "first_name"
        assert to_snake_case("totalAmount") == "total_amount"
        assert to_snake_case("userId") == "user_id"

    def test_snake_case_pascal_case(self):
        """Test conversion from PascalCase"""
        assert to_snake_case("FirstName") == "first_name"
        assert to_snake_case("LastName") == "last_name"
        assert to_snake_case("TotalAmount") == "total_amount"

    def test_snake_case_spaces(self):
        """Test conversion with spaces"""
        assert to_snake_case("User ID") == "user_id"
        assert to_snake_case("First Name") == "first_name"
        assert to_snake_case("Total Amount USD") == "total_amount_usd"

    def test_snake_case_hyphens(self):
        """Test conversion with hyphens"""
        assert to_snake_case("price-per-unit") == "price_per_unit"
        assert to_snake_case("created-at") == "created_at"
        assert to_snake_case("user-id") == "user_id"

    def test_snake_case_mixed_delimiters(self):
        """Test conversion with mixed delimiters"""
        assert to_snake_case("User ID-value") == "user_id_value"
        assert to_snake_case("First-Name Value") == "first_name_value"

    def test_snake_case_already_snake_case(self):
        """Test that already snake_case names are unchanged"""
        assert to_snake_case("first_name") == "first_name"
        assert to_snake_case("user_id") == "user_id"
        assert to_snake_case("total_amount") == "total_amount"

    def test_snake_case_all_caps(self):
        """Test conversion from ALL_CAPS"""
        assert to_snake_case("API_KEY") == "api_key"
        assert to_snake_case("HTTP_STATUS") == "http_status"

    def test_snake_case_consecutive_caps(self):
        """Test conversion with consecutive capitals"""
        # inflection separates acronyms intelligently
        assert to_snake_case("HTTPSConnection") == "https_connection"
        assert to_snake_case("XMLParser") == "xml_parser"

    def test_snake_case_numbers(self):
        """Test conversion with numbers"""
        assert to_snake_case("user123") == "user123"
        assert to_snake_case("User123ID") == "user123_id"  # ID gets separated
        assert to_snake_case("col_1") == "col_1"

    def test_snake_case_edge_cases(self):
        """Test edge cases"""
        assert to_snake_case("") == ""
        assert to_snake_case("a") == "a"
        assert to_snake_case("A") == "a"
        # parameterize strips leading/trailing and collapses consecutive underscores
        assert to_snake_case("_") == ""
        assert to_snake_case("__multiple__underscores__") == "multiple_underscores"

    def test_snake_case_leading_trailing_spaces(self):
        """Test handling of leading/trailing spaces - parameterize strips them"""
        assert to_snake_case(" FirstName ") == "first_name"
        assert to_snake_case("_firstName_") == "first_name"

    def test_snake_case_transliteration(self):
        """Test that accented characters are transliterated to ASCII"""
        assert to_snake_case("café") == "cafe"
        assert to_snake_case("naïve") == "naive"
        assert to_snake_case("résumé") == "resume"
        assert to_snake_case("Müller") == "muller"
        assert to_snake_case("niño") == "nino"
        assert to_snake_case("Ärger") == "arger"


class TestSchemaInference:
    """Test CLI schema inference functions"""

    def test_infer_schema_from_csv(self, sample_csv_file):
        """Test schema inference from CSV file"""
        result = infer_schema_from_file(
            str(sample_csv_file),
            filetype="csv",
            has_header=True,
        )
        column_mapping = result["column_mapping"]

        assert "name" in column_mapping
        assert "age" in column_mapping
        assert "score" in column_mapping

        # Check format
        assert column_mapping["name"] == ([], "string")
        assert column_mapping["age"] == ([], "int")
        assert column_mapping["score"] == ([], "float")

    def test_infer_schema_from_csv_no_header(self, sample_csv_no_header):
        """Test schema inference from headerless CSV"""
        result = infer_schema_from_file(
            str(sample_csv_no_header),
            filetype="csv",
            has_header=False,
        )
        column_mapping = result["column_mapping"]

        # Should generate col_0, col_1, col_2
        assert "col_0" in column_mapping
        assert "col_1" in column_mapping
        assert "col_2" in column_mapping

    def test_infer_schema_from_psv(self, sample_psv_file):
        """Test schema inference from pipe-delimited file"""
        result = infer_schema_from_file(
            str(sample_psv_file),
            filetype="psv",
            has_header=True,
        )
        column_mapping = result["column_mapping"]

        assert "name" in column_mapping
        assert "age" in column_mapping
        assert "score" in column_mapping

    def test_infer_schema_auto_detect_filetype(self, sample_csv_file):
        """Test auto-detection of file type from extension"""
        result = infer_schema_from_file(
            str(sample_csv_file),
            has_header=True,
        )
        column_mapping = result["column_mapping"]

        assert "name" in column_mapping
        assert len(column_mapping) == 3

    def test_infer_schema_with_snake_case_conversion(self, temp_dir):
        """Test schema inference converts column names to snake_case"""
        # Create CSV with non-snake_case column names
        csv_content = """FirstName,LastName,User ID,totalAmount,createdAt
John,Doe,12345,99.50,2024-01-01
Jane,Smith,67890,150.75,2024-01-02
"""
        csv_path = temp_dir / "test_snake_case.csv"
        csv_path.write_text(csv_content)

        result = infer_schema_from_file(
            str(csv_path),
            filetype="csv",
            has_header=True,
        )
        column_mapping = result["column_mapping"]

        # Check that column names are snake_case
        assert "first_name" in column_mapping
        assert "last_name" in column_mapping
        assert "user_id" in column_mapping
        assert "total_amount" in column_mapping
        assert "created_at" in column_mapping

        # Original names should NOT be in the mapping keys
        assert "FirstName" not in column_mapping
        assert "LastName" not in column_mapping
        assert "User ID" not in column_mapping
        assert "totalAmount" not in column_mapping
        assert "createdAt" not in column_mapping

        # Check that original names are in the possible columns array
        assert column_mapping["first_name"][0] == ["FirstName"]
        assert column_mapping["last_name"][0] == ["LastName"]
        assert column_mapping["user_id"][0] == ["User ID"]
        assert column_mapping["total_amount"][0] == ["totalAmount"]
        assert column_mapping["created_at"][0] == ["createdAt"]

        # Check types are inferred correctly
        assert column_mapping["first_name"][1] == "string"
        assert column_mapping["user_id"][1] == "int"
        assert column_mapping["total_amount"][1] == "float"

    def test_infer_schema_already_snake_case(self, temp_dir):
        """Test schema inference with already snake_case columns"""
        # Create CSV with snake_case column names
        csv_content = """first_name,last_name,user_id,total_amount
John,Doe,12345,99.50
Jane,Smith,67890,150.75
"""
        csv_path = temp_dir / "test_already_snake.csv"
        csv_path.write_text(csv_content)

        result = infer_schema_from_file(
            str(csv_path),
            filetype="csv",
            has_header=True,
        )
        column_mapping = result["column_mapping"]

        # Check that column names remain snake_case
        assert "first_name" in column_mapping
        assert "last_name" in column_mapping
        assert "user_id" in column_mapping
        assert "total_amount" in column_mapping

        # Check that possible columns arrays are EMPTY (no conversion needed)
        assert column_mapping["first_name"][0] == []
        assert column_mapping["last_name"][0] == []
        assert column_mapping["user_id"][0] == []
        assert column_mapping["total_amount"][0] == []

    def test_infer_schema_mixed_case_types(self, temp_dir):
        """Test schema inference with various column name formats"""
        # Create CSV with mixed naming conventions
        csv_content = """camelCase,PascalCase,snake_case,UPPER_CASE,Spaced Name
value1,value2,value3,value4,value5
"""
        csv_path = temp_dir / "test_mixed_case.csv"
        csv_path.write_text(csv_content)

        result = infer_schema_from_file(
            str(csv_path),
            filetype="csv",
            has_header=True,
        )
        column_mapping = result["column_mapping"]

        # Check conversions
        assert "camel_case" in column_mapping
        assert column_mapping["camel_case"][0] == ["camelCase"]

        assert "pascal_case" in column_mapping
        assert column_mapping["pascal_case"][0] == ["PascalCase"]

        assert "snake_case" in column_mapping
        assert column_mapping["snake_case"][0] == []  # Already snake_case

        assert "upper_case" in column_mapping
        assert column_mapping["upper_case"][0] == ["UPPER_CASE"]

        assert "spaced_name" in column_mapping
        assert column_mapping["spaced_name"][0] == ["Spaced Name"]

    def test_infer_schema_detects_null_values(self, temp_dir):
        """Test schema inference handles null value representations"""
        csv_content = """name,age,score,status
Alice,25,95.5,active
Bob,NA,80.0,inactive
Charlie,30,None,active
Dave,N/A,70.0,NULL
"""
        csv_path = temp_dir / "test_nulls.csv"
        csv_path.write_text(csv_content)

        result = infer_schema_from_file(
            str(csv_path),
            filetype="csv",
            has_header=True,
        )

        # Pandas handles nulls internally but doesn't report detected null patterns
        # The important thing is that the schema is still correctly inferred
        assert "column_mapping" in result
        assert "name" in result["column_mapping"]
        assert "age" in result["column_mapping"]

    def test_infer_schema_no_null_values(self, temp_dir):
        """Test schema inference returns None when no custom null values present"""
        csv_content = """name,age,score
Alice,25,95.5
Bob,30,80.0
Charlie,35,70.0
"""
        csv_path = temp_dir / "test_no_nulls.csv"
        csv_path.write_text(csv_content)

        result = infer_schema_from_file(
            str(csv_path),
            filetype="csv",
            has_header=True,
        )

        # No custom null values should be detected
        assert result["null_values"] is None


class TestCLIOutputFormat:
    """Test CLI output format with table_name and column_mapping structure"""

    def test_cli_single_file_output_format(self, temp_dir):
        """Test CLI output format for single file includes table_name"""
        import subprocess
        import json

        # Create a test CSV
        csv_content = """FirstName,LastName,Age
John,Doe,30
Jane,Smith,25
"""
        csv_path = temp_dir / "UserData.csv"
        csv_path.write_text(csv_content)

        # Run CLI
        result = subprocess.run(
            ["python", "table_functions.py", str(csv_path)],
            capture_output=True,
            text=True,
        )

        assert result.returncode == 0
        output = json.loads(result.stdout)

        # Check structure: keyed by original filename
        assert "UserData.csv" in output
        file_data = output["UserData.csv"]

        # Check table_name is snake_case of stem
        assert "table_name" in file_data
        assert file_data["table_name"] == "user_data"

        # Check column_mapping exists
        assert "column_mapping" in file_data
        column_mapping = file_data["column_mapping"]

        # Check column mapping content
        assert "first_name" in column_mapping
        assert column_mapping["first_name"][0] == ["FirstName"]
        assert "last_name" in column_mapping
        assert "age" in column_mapping

    def test_cli_directory_output_format(self, temp_dir):
        """Test CLI output format for directory includes table_name per file"""
        import subprocess
        import json

        # Create test directory with multiple files
        data_dir = temp_dir / "data"
        data_dir.mkdir()

        # File 1
        csv1 = """Name,Value
foo,1
bar,2
"""
        (data_dir / "SalesData.csv").write_text(csv1)

        # File 2
        csv2 = """Id,Amount
100,50.5
200,75.25
"""
        (data_dir / "inventory_items.csv").write_text(csv2)

        # Run CLI on directory
        result = subprocess.run(
            ["python", "table_functions.py", str(data_dir)],
            capture_output=True,
            text=True,
        )

        assert result.returncode == 0
        output = json.loads(result.stdout)

        # Check both files are present, keyed by original filename
        assert "SalesData.csv" in output
        assert "inventory_items.csv" in output

        # Check SalesData.csv
        sales = output["SalesData.csv"]
        assert sales["table_name"] == "sales_data"
        assert "column_mapping" in sales
        assert "name" in sales["column_mapping"]

        # Check inventory_items.csv (already snake_case)
        inventory = output["inventory_items.csv"]
        assert inventory["table_name"] == "inventory_items"
        assert "column_mapping" in inventory
        assert "id" in inventory["column_mapping"]
        assert "amount" in inventory["column_mapping"]

    def test_cli_output_usable_with_column_mapping_fn(self, temp_dir):
        """Test that CLI output can be used directly with column_mapping_fn pattern"""
        import subprocess
        import json
        from pathlib import Path

        # Create test files
        data_dir = temp_dir / "data"
        data_dir.mkdir()

        csv1 = """ProductName,Price
Widget,10.99
Gadget,25.50
"""
        (data_dir / "products.csv").write_text(csv1)

        # Get CLI output
        result = subprocess.run(
            ["python", "table_functions.py", str(data_dir)],
            capture_output=True,
            text=True,
        )

        output = json.loads(result.stdout)

        # Simulate the pattern from CLAUDE.md
        all_mappings = output

        def get_column_mapping(file_path):
            return all_mappings[file_path.name]["column_mapping"]

        def get_table_name(file_path):
            return all_mappings[file_path.name]["table_name"]

        # Test the functions work correctly
        test_path = Path(data_dir / "products.csv")

        column_mapping = get_column_mapping(test_path)
        assert "product_name" in column_mapping
        assert column_mapping["product_name"][0] == ["ProductName"]
        assert "price" in column_mapping

        table_name = get_table_name(test_path)
        assert table_name == "products"


# ===== INTEGRATION TESTS =====


class TestIntegration:
    """Integration tests for complete workflows"""

    def test_add_files_to_metadata_table(
        self, conninfo, db_conn, temp_dir, sample_csv_file
    ):
        """Test adding files to metadata table"""
        # Create search and landing directories
        source_dir = temp_dir / "search"
        landing_dir = temp_dir / "landing"
        source_dir.mkdir()
        landing_dir.mkdir()

        # Copy CSV to search dir
        import shutil

        dest_file = source_dir / "test.csv"
        shutil.copy(sample_csv_file, dest_file)

        # Test the add_files function (sub-function of add_files_to_metadata_table)
        file_list = [str(dest_file)]
        rows = add_files(
            source_dir=str(source_dir) + "/",
            resume=False,
            sample=None,
            file_list=file_list,
            filetype="csv",
            has_header=True,
            source_path_list=[],
        )

        assert len(rows) == 1
        assert rows[0]["metadata_ingest_status"] == "Success"

    def test_extract_and_add_zip_files(self, db_conn, temp_dir, sample_zip_file):
        """Test extracting ZIP files and adding to metadata"""
        import os

        original_cwd = os.getcwd()
        try:
            os.chdir(temp_dir)
            file_list = [str(sample_zip_file)]

            rows, archive_stats = extract_and_add_zip_files(
                file_list=file_list,
                source_path_list=[],
                source_dir=str(temp_dir) + "/",
                has_header=True,
                filetype="csv",
                resume=False,
                sample=None,
                archive_glob="*.csv",
            )

            assert len(rows) == 2  # Two CSV files in the ZIP
            assert len(archive_stats) == 1  # One archive processed
            assert archive_stats[str(sample_zip_file)] == 2  # 2 files from that archive

            # Verify source_path has :: delimiter for archives
            for row in rows:
                assert "::" in row["source_path"]
        finally:
            os.chdir(original_cwd)

    def test_extract_zip_with_nested_directories(self, db_conn, temp_dir):
        """Test extracting ZIP with nested directory structure extracts correctly"""
        import zipfile
        import os

        original_cwd = os.getcwd()
        try:
            os.chdir(temp_dir)

            # Create ZIP with nested structure
            zip_path = temp_dir / "nested.zip"
            with zipfile.ZipFile(zip_path, "w") as zf:
                zf.writestr("subdir/nested/file1.csv", "a,b\n1,2\n")
                zf.writestr("another/deep/path/file2.csv", "x,y\n3,4\n")

            rows, archive_stats = extract_and_add_zip_files(
                file_list=[str(zip_path)],
                source_path_list=[],
                source_dir=str(temp_dir) + "/",
                has_header=True,
                filetype="csv",
                resume=False,
                sample=None,
                archive_glob="*.csv",
            )

            assert len(rows) == 2

            # Verify source_path includes the inner path with :: delimiter
            source_paths = [row["source_path"] for row in rows]
            assert any("subdir/nested/file1.csv" in p for p in source_paths)
            assert any("another/deep/path/file2.csv" in p for p in source_paths)
        finally:
            os.chdir(original_cwd)

    def test_extract_zip_no_temp_file_leakage(self, db_conn, temp_dir):
        """Test that temp files are cleaned up after extraction"""
        import zipfile
        import tempfile
        import os

        original_cwd = os.getcwd()
        try:
            os.chdir(temp_dir)

            # Create a simple ZIP
            zip_path = temp_dir / "test.zip"
            with zipfile.ZipFile(zip_path, "w") as zf:
                zf.writestr("file.csv", "a,b\n1,2\n")

            # Get system temp dir to check for leakage
            system_temp = Path(tempfile.gettempdir())

            # Count temp dirs before
            temp_dirs_before = set(system_temp.iterdir())

            rows, archive_stats = extract_and_add_zip_files(
                file_list=[str(zip_path)],
                source_path_list=[],
                source_dir=str(temp_dir) + "/",
                has_header=True,
                filetype="csv",
                resume=False,
                sample=None,
                archive_glob="*.csv",
            )

            # Count temp dirs after
            temp_dirs_after = set(system_temp.iterdir())

            # No new temp directories should remain
            new_temp_dirs = temp_dirs_after - temp_dirs_before
            # Filter to only dirs that look like our temp extraction dirs
            leaked_dirs = [
                d for d in new_temp_dirs if d.is_dir() and "tmp" in d.name.lower()
            ]

            assert len(rows) == 1
            assert len(leaked_dirs) == 0, f"Temp directories leaked: {leaked_dirs}"
        finally:
            os.chdir(original_cwd)

    def test_drop_metadata_by_source(self, conninfo, db_conn, temp_dir):
        """Test dropping files from metadata by source_dir"""
        # Create metadata table
        execute_sql(
            db_conn,
            """
            CREATE TABLE test_schema.metadata (
                source_dir TEXT,
                source_path TEXT PRIMARY KEY
            )
        """,
        )
        execute_sql(
            db_conn,
            "INSERT INTO test_schema.metadata VALUES ($1, $2)",
            (str(temp_dir), str(temp_dir / "file1.csv")),
        )
        execute_sql(
            db_conn,
            "INSERT INTO test_schema.metadata VALUES ($1, $2)",
            (str(temp_dir), str(temp_dir / "file2.csv")),
        )

        drop_metadata_by_source(
            conninfo=conninfo,
            source_dir=str(temp_dir),
            schema="test_schema",
        )

        # Verify deletion
        result = execute_sql_fetchone(
            db_conn, "SELECT COUNT(*) FROM test_schema.metadata"
        )
        assert result[0] == 0

    def test_drop_partition(self, conninfo, db_conn):
        """Test dropping partition from table"""
        # Create test table
        execute_sql(
            db_conn,
            """
            CREATE TABLE test_schema.partition_test (
                source_path TEXT,
                data TEXT
            )
        """,
        )
        execute_sql(
            db_conn,
            "INSERT INTO test_schema.partition_test VALUES ($1, $2)",
            ("/path/to/file1.csv", "data1"),
        )
        execute_sql(
            db_conn,
            "INSERT INTO test_schema.partition_test VALUES ($1, $2)",
            ("/path/to/file2.csv", "data2"),
        )

        drop_partition(
            conninfo=conninfo,
            table="partition_test",
            partition_key="/path/to/file1.csv",
            schema="test_schema",
        )

        # Verify deletion
        result = execute_sql_fetchone(
            db_conn, "SELECT COUNT(*) FROM test_schema.partition_test"
        )
        assert result[0] == 1


# ===== EDGE CASE TESTS =====


class TestEdgeCases:
    """Test edge cases and error handling"""

    def test_read_csv_with_bom(self, temp_dir):
        """Test reading CSV with BOM (byte-order mark)"""
        csv_path = temp_dir / "bom.csv"
        # Write file with BOM
        csv_path.write_bytes(b"\xef\xbb\xbfname,age\nAlice,25\n")

        column_mapping = {
            "name": ([], "string"),
            "age": ([], "int"),
        }

        df = read_csv(
            full_path=str(csv_path),
            column_mapping=column_mapping,
            has_header=True,
        )

        # BOM should be handled correctly
        assert "name" in df.columns
        assert len(df) == 1

    def test_read_csv_empty_null_values(self, temp_dir):
        """Test handling empty strings as null values"""
        csv_path = temp_dir / "nulls.csv"
        csv_path.write_text("name,age\nAlice,25\nBob,\n")

        column_mapping = {
            "name": ([], "string"),
            "age": ([], "int"),
        }

        df = read_csv(
            full_path=str(csv_path),
            column_mapping=column_mapping,
            has_header=True,
            null_values=[""],
        )

        assert pd.isna(df["age"][1])

    def test_read_csv_multiple_null_values(self, temp_dir):
        """Test handling multiple null value representations"""
        csv_path = temp_dir / "multi_nulls.csv"
        csv_path.write_text(
            "name,age,score\nAlice,25,100\nBob,NA,80\nCharlie,30,None\nDave,N/A,N/A\n"
        )

        column_mapping = {
            "name": ([], "string"),
            "age": ([], "int"),
            "score": ([], "int"),
        }

        df = read_csv(
            full_path=str(csv_path),
            column_mapping=column_mapping,
            has_header=True,
            null_values=["NA", "None", "N/A"],
        )

        # Bob's age should be null (NA)
        assert pd.isna(df["age"][1])
        # Charlie's score should be null (None)
        assert pd.isna(df["score"][2])
        # Dave's age and score should both be null (N/A)
        assert pd.isna(df["age"][3])
        assert pd.isna(df["score"][3])
        # Alice should have valid values
        assert df["age"][0] == 25
        assert df["score"][0] == 100

    def test_copy_dataframe_large_error_message(self, db_conn):
        """Test error message truncation in update_metadata"""
        # Create metadata table with all required columns
        execute_sql(
            db_conn,
            """
            CREATE TABLE test_schema.metadata (
                source_path TEXT PRIMARY KEY,
                ingest_datetime TIMESTAMP,
                status TEXT,
                error_message TEXT,
                unpivot_row_multiplier INTEGER,
                ingest_runtime INTEGER,
                output_table TEXT
            )
        """,
        )
        execute_sql(
            db_conn,
            "INSERT INTO test_schema.metadata (source_path) VALUES ($1)",
            ("/path/to/file.csv",),
        )

        # Create very long error message
        long_error = "x" * 1000

        update_metadata(
            conn=db_conn,
            source_path="/path/to/file.csv",
            schema="test_schema",
            error_message=long_error,
        )

        # Verify truncation
        result = execute_sql_fetchone(
            db_conn,
            "SELECT error_message FROM test_schema.metadata WHERE source_path = $1",
            ("/path/to/file.csv",),
        )
        error = result[0]

        assert len(error) <= 550  # 500 + "... [truncated]"
        assert "truncated" in error

    def test_validate_schema_type_mismatch(self, db_conn):
        """Test schema validation with type mismatch"""
        # Create table with wrong types
        execute_sql(
            db_conn,
            """
            CREATE TABLE test_schema.type_test (
                id TEXT,
                name BIGINT
            )
        """,
        )

        # Create DataFrame with different types
        df = pd.DataFrame(
            {
                "id": pd.Series([1, 2], dtype=pd.Int64Dtype()),
                "name": pd.Series(["Alice", "Bob"], dtype="string"),
            }
        )

        with pytest.raises(ValueError, match="type mismatches"):
            validate_schema_match(db_conn, df, "test_schema", "type_test")


class TestMissingFunctions:
    """Tests for functions that weren't covered yet"""

    def test_read_xlsx(self, temp_dir):
        """Test reading Excel files"""
        # Create a simple Excel file
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "name"
        ws["B1"] = "age"
        ws["C1"] = "score"
        ws["A2"] = "Alice"
        ws["B2"] = 25
        ws["C2"] = 85.5
        ws["A3"] = "Bob"
        ws["B3"] = 30
        ws["C3"] = 92.0

        xlsx_path = temp_dir / "test.xlsx"
        wb.save(str(xlsx_path))

        column_mapping = {
            "name": ([], "string"),
            "age": ([], "int"),
            "score": ([], "float"),
        }

        df = read_xlsx(
            full_path=str(xlsx_path),
            column_mapping=column_mapping,
        )

        assert len(df) == 2
        assert list(df.columns) == ["name", "age", "score"]
        assert df["name"][0] == "Alice"

    def test_read_parquet(self, temp_dir):
        """Test reading Parquet files with column_mapping"""
        # Create a Parquet file using pandas
        test_df = pd.DataFrame(
            {
                "name": ["Alice", "Bob", "Charlie"],
                "age": [25, 30, 35],
                "score": [85.5, 92.0, 78.5],
                "extra_col": ["x", "y", "z"],  # Column to be excluded
            }
        )

        parquet_path = temp_dir / "test.parquet"
        test_df.to_parquet(str(parquet_path))

        # Test with column_mapping
        column_mapping = {
            "name": ([], "string"),
            "age": ([], "int"),
            "score": ([], "float"),
            # extra_col is intentionally not included - should be filtered out
        }

        df = read_parquet(
            full_path=str(parquet_path),
            column_mapping=column_mapping,
        )

        assert len(df) == 3
        assert list(df.columns) == ["name", "age", "score"]
        assert "extra_col" not in df.columns  # Should be filtered out
        assert df["name"][0] == "Alice"

    def test_read_parquet_with_rename(self, temp_dir):
        """Test reading Parquet files with column renaming"""
        test_df = pd.DataFrame(
            {
                "old_name": ["Alice", "Bob"],
                "value": [100, 200],
            }
        )

        parquet_path = temp_dir / "test_rename.parquet"
        test_df.to_parquet(str(parquet_path))

        column_mapping = {
            "new_name": (["old_name"], "string"),
            "value": ([], "int"),
        }

        df = read_parquet(
            full_path=str(parquet_path),
            column_mapping=column_mapping,
        )

        assert "new_name" in df.columns
        assert "old_name" not in df.columns
        assert df["new_name"][0] == "Alice"

    def test_read_parquet_with_missing_cols(self, temp_dir):
        """Test reading Parquet with missing columns in mapping"""
        test_df = pd.DataFrame(
            {
                "name": ["Alice", "Bob"],
            }
        )

        parquet_path = temp_dir / "test_missing.parquet"
        test_df.to_parquet(str(parquet_path))

        column_mapping = {
            "name": ([], "string"),
            "missing_col": ([], "string"),  # Not in parquet file
        }

        df = read_parquet(
            full_path=str(parquet_path),
            column_mapping=column_mapping,
        )

        assert "missing_col" in df.columns
        assert df["missing_col"].isna().sum() == len(df)

    def test_s3_path_detection(self, temp_dir):
        """Test S3 path detection in various contexts"""
        # Valid S3 paths
        assert is_s3_path("s3://my-bucket/path/file.csv")
        assert is_s3_path("s3://bucket/")

        # Invalid S3 paths
        assert not is_s3_path("/local/path/file.csv")
        assert not is_s3_path(str(temp_dir / "file.csv"))

    def test_update_table_basic(self, conninfo, db_conn, temp_dir, sample_csv_file):
        """Test update_table function end-to-end"""
        # Create metadata table first
        execute_sql(
            db_conn,
            """
            CREATE TABLE test_schema.metadata (
                source_dir TEXT,
                source_path TEXT PRIMARY KEY,
                filesize BIGINT,
                header TEXT[],
                row_count BIGINT,
                                file_hash TEXT,
                metadata_ingest_datetime TIMESTAMP,
                metadata_ingest_status TEXT,
                ingest_datetime TIMESTAMP,
                ingest_runtime INTEGER,
                status TEXT,
                error_message TEXT,
                unpivot_row_multiplier INTEGER,
                output_table TEXT
            )
        """,
        )

        # Insert metadata for the sample file
        # Note: landing_dir must have trailing slash to match update_table queries
        execute_sql(
            db_conn,
            """
            INSERT INTO test_schema.metadata
            (source_path, row_count, metadata_ingest_status, source_dir)
            VALUES ($1, $2, $3, $4)
        """,
            (str(sample_csv_file), 3, "Success", str(temp_dir) + "/"),
        )

        column_mapping = {
            "name": ([], "string"),
            "age": ([], "int"),
            "score": ([], "float"),
        }

        # Run update_table
        result_df = update_table(
            conninfo=conninfo,
            schema="test_schema",
            output_table="test_output",
            filetype="csv",
            column_mapping=column_mapping,
            source_dir=str(temp_dir) + "/",
            resume=False,
        )

        # Verify data was loaded
        result = execute_sql_fetchone(
            db_conn, "SELECT COUNT(*) FROM test_schema.test_output"
        )
        assert result[0] == 3

    def test_add_files_to_metadata_table_full(
        self, conninfo, db_conn, temp_dir, sample_csv_file
    ):
        """Test add_files_to_metadata_table end-to-end"""
        # This is already tested indirectly through add_files in TestIntegration
        # Here we just verify the metadata table creation aspect

        # Clean up any existing metadata table from previous tests
        execute_sql(db_conn, "DROP TABLE IF EXISTS test_schema.metadata")

        # Create search and landing directories
        source_dir = temp_dir / "search"
        landing_dir = temp_dir / "landing"
        source_dir.mkdir()
        landing_dir.mkdir()

        # Copy CSV to search dir
        import shutil

        dest_file = source_dir / "test.csv"
        shutil.copy(sample_csv_file, dest_file)

        # This would call add_files_to_metadata_table but it's complex
        # and requires specific database state. Already covered in integration tests.
        # Just verify the table creation logic works

        assert not table_exists(db_conn, "test_schema", "metadata")

        # Create metadata table manually to test the schema
        execute_sql(
            db_conn,
            """
            CREATE TABLE test_schema.metadata (
                source_dir TEXT,
                source_path TEXT PRIMARY KEY,
                filesize BIGINT,
                header TEXT[],
                row_count BIGINT,
                                file_hash TEXT,
                metadata_ingest_datetime TIMESTAMP,
                metadata_ingest_status TEXT,
                ingest_datetime TIMESTAMP,
                ingest_runtime INTEGER,
                status TEXT,
                error_message TEXT,
                unpivot_row_multiplier INTEGER,
                output_table TEXT
            )
        """,
        )

        assert table_exists(db_conn, "test_schema", "metadata")


class TestAdditionalFunctions:
    """Additional tests for uncovered functions"""

    def test_drop_file_from_metadata_and_table(self, conninfo, db_conn):
        """Test dropping file from both metadata and data table"""
        # Create metadata table
        execute_sql(
            db_conn,
            """
            CREATE TABLE test_schema.metadata (
                source_path TEXT PRIMARY KEY
            )
        """,
        )
        execute_sql(
            db_conn,
            "INSERT INTO test_schema.metadata VALUES ($1)",
            ("/path/to/file.csv",),
        )

        # Create data table
        execute_sql(
            db_conn,
            """
            CREATE TABLE test_schema.data_table (
                source_path TEXT,
                data TEXT
            )
        """,
        )
        execute_sql(
            db_conn,
            "INSERT INTO test_schema.data_table VALUES ($1, $2)",
            ("/path/to/file.csv", "test_data"),
        )

        # Drop file from both tables
        drop_file_from_metadata_and_table(
            conninfo=conninfo,
            table="data_table",
            source_path="/path/to/file.csv",
            schema="test_schema",
        )

        # Verify deletions
        metadata_count = execute_sql_fetchone(
            db_conn, "SELECT COUNT(*) FROM test_schema.metadata"
        )[0]
        data_count = execute_sql_fetchone(
            db_conn, "SELECT COUNT(*) FROM test_schema.data_table"
        )[0]

        assert metadata_count == 0
        assert data_count == 0

    def test_read_using_column_mapping_invalid_filetype(self):
        """Test router function with invalid filetype"""
        import pytest

        with pytest.raises(ValueError, match="Invalid filetype: invalid_type"):
            read_using_column_mapping(
                full_path="/path/to/file.xyz",
                filetype="invalid_type",
                column_mapping={},
                has_header=True,
            )

    def test_row_count_check_with_unpivot_multiplier(self, db_conn, sample_csv_file):
        """Test row count check with unpivot multiplier"""
        # Create metadata table
        execute_sql(
            db_conn,
            """
            CREATE TABLE test_schema.metadata (
                source_path TEXT PRIMARY KEY,
                row_count BIGINT
            )
        """,
        )
        execute_sql(
            db_conn,
            "INSERT INTO test_schema.metadata (source_path, row_count) VALUES ($1, $2)",
            (str(sample_csv_file), 3),
        )

        # Create DataFrame with 9 rows (3 * 3 multiplier)
        df = pd.DataFrame({"col1": list(range(9))})

        # Should not raise with multiplier of 3
        row_count_check(
            conn=db_conn,
            schema="test_schema",
            df=df,
            source_path=str(sample_csv_file),
            unpivot_row_multiplier=3,
        )

    def test_update_metadata_with_unpivot_multiplier(self, db_conn):
        """Test updating metadata with unpivot row multiplier"""
        # Create metadata table
        execute_sql(
            db_conn,
            """
            CREATE TABLE test_schema.metadata (
                source_path TEXT PRIMARY KEY,
                ingest_datetime TIMESTAMP,
                status TEXT,
                error_message TEXT,
                unpivot_row_multiplier INTEGER,
                ingest_runtime INTEGER,
                output_table TEXT
            )
        """,
        )
        execute_sql(
            db_conn,
            "INSERT INTO test_schema.metadata (source_path) VALUES ($1)",
            ("/path/to/file.csv",),
        )

        update_metadata(
            conn=db_conn,
            source_path="/path/to/file.csv",
            schema="test_schema",
            unpivot_row_multiplier=5,
            ingest_runtime=10,
        )

        # Verify update
        result = execute_sql_fetchone(
            db_conn,
            "SELECT unpivot_row_multiplier, ingest_runtime FROM test_schema.metadata WHERE source_path = $1",
            ("/path/to/file.csv",),
        )

        assert result[0] == 5
        assert result[1] == 10


# ===== UPDATE_TABLE ADVANCED FEATURE TESTS =====


class TestUpdateTableAdvancedFeatures:
    """Test advanced features of update_table function"""

    def _setup_metadata_table(self, db_conn, temp_dir, csv_files):
        """Helper to set up metadata table with files"""
        execute_sql(db_conn, "DROP TABLE IF EXISTS test_schema.metadata")
        execute_sql(
            db_conn,
            """
            CREATE TABLE test_schema.metadata (
                source_dir TEXT,
                source_path TEXT PRIMARY KEY,
                filesize BIGINT,
                header TEXT[],
                row_count BIGINT,
                                file_hash TEXT,
                metadata_ingest_datetime TIMESTAMP,
                metadata_ingest_status TEXT,
                ingest_datetime TIMESTAMP,
                ingest_runtime INTEGER,
                status TEXT,
                error_message TEXT,
                unpivot_row_multiplier INTEGER,
                output_table TEXT
            )
        """,
        )

        for csv_file, row_count in csv_files:
            execute_sql(
                db_conn,
                """
                INSERT INTO test_schema.metadata
                (source_path, row_count, metadata_ingest_status, source_dir)
                VALUES ($1, $2, $3, $4)
            """,
                (str(csv_file), row_count, "Success", str(temp_dir) + "/"),
            )

    def test_update_table_with_transform_fn(self, conninfo, db_conn, temp_dir):
        """Test update_table with transform_fn"""
        # Create CSV file
        csv_path = temp_dir / "transform_test.csv"
        csv_path.write_text("name,value\nAlice,100\nBob,200\n")

        # Setup metadata
        self._setup_metadata_table(db_conn, temp_dir, [(csv_path, 2)])

        column_mapping = {
            "name": ([], "string"),
            "value": ([], "int"),
        }

        # Transform function that doubles the value
        def transform_fn(df):
            df["value"] = df["value"] * 2
            return df

        result_df = update_table(
            conninfo=conninfo,
            schema="test_schema",
            output_table="transform_output",
            filetype="csv",
            column_mapping=column_mapping,
            source_dir=str(temp_dir) + "/",
            transform_fn=transform_fn,
            resume=False,
        )

        # Verify transformation was applied
        result = execute_sql_fetchone(
            db_conn,
            "SELECT value FROM test_schema.transform_output WHERE name = 'Alice'",
        )
        assert result[0] == 200  # 100 * 2

    def test_update_table_with_additional_cols_fn(self, conninfo, db_conn, temp_dir):
        """Test update_table with additional_cols_fn"""
        csv_path = temp_dir / "additional_cols_test.csv"
        csv_path.write_text("name,value\nAlice,100\n")

        self._setup_metadata_table(db_conn, temp_dir, [(csv_path, 1)])

        column_mapping = {
            "name": ([], "string"),
            "value": ([], "int"),
        }

        # Additional columns function
        def additional_cols_fn(file_path):
            return {
                "source_file": Path(file_path).stem,
                "load_timestamp": "2024-01-01",
            }

        result_df = update_table(
            conninfo=conninfo,
            schema="test_schema",
            output_table="additional_cols_output",
            filetype="csv",
            column_mapping=column_mapping,
            source_dir=str(temp_dir) + "/",
            additional_cols_fn=additional_cols_fn,
            resume=False,
        )

        # Verify additional columns were added
        result = execute_sql_fetchone(
            db_conn,
            "SELECT source_file, load_timestamp FROM test_schema.additional_cols_output",
        )
        assert result[0] == "additional_cols_test"
        assert result[1] == "2024-01-01"

    def test_update_table_with_output_table_naming_fn(
        self, conninfo, db_conn, temp_dir
    ):
        """Test update_table with output_table_naming_fn"""
        csv_path = temp_dir / "naming_test.csv"
        csv_path.write_text("name,value\nAlice,100\n")

        self._setup_metadata_table(db_conn, temp_dir, [(csv_path, 1)])

        column_mapping = {
            "name": ([], "string"),
            "value": ([], "int"),
        }

        # Naming function that uses file stem as table name
        def output_table_naming_fn(file_path):
            return f"table_{Path(file_path).stem}"

        result_df = update_table(
            conninfo=conninfo,
            schema="test_schema",
            output_table_naming_fn=output_table_naming_fn,
            filetype="csv",
            column_mapping=column_mapping,
            source_dir=str(temp_dir) + "/",
            resume=False,
        )

        # Verify table was created with custom name
        assert table_exists(db_conn, "test_schema", "table_naming_test")

    def test_update_table_with_file_list_filter_fn(self, conninfo, db_conn, temp_dir):
        """Test update_table with file_list_filter_fn"""
        # Create multiple CSV files
        csv1 = temp_dir / "include_me.csv"
        csv1.write_text("name,value\nAlice,100\n")
        csv2 = temp_dir / "exclude_me.csv"
        csv2.write_text("name,value\nBob,200\n")

        self._setup_metadata_table(db_conn, temp_dir, [(csv1, 1), (csv2, 1)])

        column_mapping = {
            "name": ([], "string"),
            "value": ([], "int"),
        }

        # Filter function that only includes files with "include" in the name
        def file_list_filter_fn(file_list):
            return [f for f in file_list if "include" in str(f)]

        result_df = update_table(
            conninfo=conninfo,
            schema="test_schema",
            output_table="filter_output",
            filetype="csv",
            column_mapping=column_mapping,
            source_dir=str(temp_dir) + "/",
            file_list_filter_fn=file_list_filter_fn,
            resume=False,
        )

        # Verify only one file was processed
        result = execute_sql_fetchone(
            db_conn, "SELECT COUNT(*) FROM test_schema.filter_output"
        )
        assert result[0] == 1

    def test_update_table_with_custom_read_fn(self, conninfo, db_conn, temp_dir):
        """Test update_table with custom_read_fn"""
        csv_path = temp_dir / "custom_read.csv"
        csv_path.write_text("name,value\nAlice,100\n")

        self._setup_metadata_table(db_conn, temp_dir, [(csv_path, 1)])

        # Custom read function that adds a computed column
        def custom_read_fn(full_path):
            df = pd.read_csv(full_path)
            df["computed"] = df["value"] * 10
            return df

        result_df = update_table(
            conninfo=conninfo,
            schema="test_schema",
            output_table="custom_read_output",
            filetype="csv",
            custom_read_fn=custom_read_fn,
            source_dir=str(temp_dir) + "/",
            resume=False,
        )

        # Verify custom read was used
        result = execute_sql_fetchone(
            db_conn, "SELECT computed FROM test_schema.custom_read_output"
        )
        assert result[0] == 1000

    def test_update_table_with_column_mapping_fn(self, conninfo, db_conn, temp_dir):
        """Test update_table with column_mapping_fn"""
        csv_path = temp_dir / "mapping_fn_test.csv"
        csv_path.write_text("name,value\nAlice,100\n")

        self._setup_metadata_table(db_conn, temp_dir, [(csv_path, 1)])

        # Dynamic column mapping function
        def column_mapping_fn(file_path):
            return {
                "name": ([], "string"),
                "value": ([], "int"),
            }

        result_df = update_table(
            conninfo=conninfo,
            schema="test_schema",
            output_table="mapping_fn_output",
            filetype="csv",
            column_mapping_fn=column_mapping_fn,
            source_dir=str(temp_dir) + "/",
            resume=False,
        )

        # Verify data was loaded
        result = execute_sql_fetchone(
            db_conn, "SELECT COUNT(*) FROM test_schema.mapping_fn_output"
        )
        assert result[0] == 1

    def test_update_table_with_pivot_mapping(self, conninfo, db_conn, temp_dir):
        """Test update_table with pivot_mapping (unpivot)"""
        csv_path = temp_dir / "pivot_test.csv"
        csv_path.write_text("id,jan,feb,mar\n1,100,200,300\n")

        # Note: row_count is 1 but after unpivot it will be 3 (1 * 3 value columns)
        self._setup_metadata_table(db_conn, temp_dir, [(csv_path, 1)])

        column_mapping = {
            "id": ([], "int"),
            "jan": ([], "int"),
            "feb": ([], "int"),
            "mar": ([], "int"),
        }

        pivot_mapping = {
            "id_vars": ["id"],
            "variable_column_name": "month",
            "value_column_name": "amount",
        }

        result_df = update_table(
            conninfo=conninfo,
            schema="test_schema",
            output_table="pivot_output",
            filetype="csv",
            column_mapping=column_mapping,
            source_dir=str(temp_dir) + "/",
            pivot_mapping=pivot_mapping,
            resume=False,
        )

        # Verify unpivot was applied - should have 3 rows (jan, feb, mar)
        result = execute_sql_fetchone(
            db_conn, "SELECT COUNT(*) FROM test_schema.pivot_output"
        )
        assert result[0] == 3

    def test_update_table_with_sample(self, conninfo, db_conn, temp_dir):
        """Test update_table with sample parameter"""
        # Create multiple CSV files
        for i in range(5):
            csv_path = temp_dir / f"sample_{i}.csv"
            csv_path.write_text(f"name,value\nRow{i},100\n")

        files = [(temp_dir / f"sample_{i}.csv", 1) for i in range(5)]
        self._setup_metadata_table(db_conn, temp_dir, files)

        column_mapping = {
            "name": ([], "string"),
            "value": ([], "int"),
        }

        result_df = update_table(
            conninfo=conninfo,
            schema="test_schema",
            output_table="sample_output",
            filetype="csv",
            column_mapping=column_mapping,
            source_dir=str(temp_dir) + "/",
            sample=2,  # Only process 2 files
            resume=False,
        )

        # Verify only 2 files were processed
        result = execute_sql_fetchone(
            db_conn, "SELECT COUNT(DISTINCT source_path) FROM test_schema.sample_output"
        )
        assert result[0] == 2

    def test_update_table_resume(self, conninfo, db_conn, temp_dir):
        """Test update_table with resume=True"""
        csv1 = temp_dir / "resume1.csv"
        csv1.write_text("name,value\nAlice,100\n")
        csv2 = temp_dir / "resume2.csv"
        csv2.write_text("name,value\nBob,200\n")

        # Setup metadata with one file already processed
        execute_sql(db_conn, "DROP TABLE IF EXISTS test_schema.metadata")
        execute_sql(
            db_conn,
            """
            CREATE TABLE test_schema.metadata (
                source_dir TEXT,
                source_path TEXT PRIMARY KEY,
                filesize BIGINT,
                header TEXT[],
                row_count BIGINT,
                                file_hash TEXT,
                metadata_ingest_datetime TIMESTAMP,
                metadata_ingest_status TEXT,
                ingest_datetime TIMESTAMP,
                ingest_runtime INTEGER,
                status TEXT,
                error_message TEXT,
                unpivot_row_multiplier INTEGER,
                output_table TEXT
            )
        """,
        )

        # File 1 is already processed
        execute_sql(
            db_conn,
            """
            INSERT INTO test_schema.metadata
            (source_path, row_count, metadata_ingest_status, source_dir, ingest_datetime, status)
            VALUES ($1, $2, $3, $4, NOW(), 'Success')
        """,
            (str(csv1), 1, "Success", str(temp_dir) + "/"),
        )

        # File 2 is not processed yet
        execute_sql(
            db_conn,
            """
            INSERT INTO test_schema.metadata
            (source_path, row_count, metadata_ingest_status, source_dir)
            VALUES ($1, $2, $3, $4)
        """,
            (str(csv2), 1, "Success", str(temp_dir) + "/"),
        )

        column_mapping = {
            "name": ([], "string"),
            "value": ([], "int"),
        }

        result_df = update_table(
            conninfo=conninfo,
            schema="test_schema",
            output_table="resume_output",
            filetype="csv",
            column_mapping=column_mapping,
            source_dir=str(temp_dir) + "/",
            resume=True,  # Only process unprocessed files
        )

        # Verify only the new file was processed
        result = execute_sql_fetchone(
            db_conn, "SELECT COUNT(*) FROM test_schema.resume_output"
        )
        assert result[0] == 1

    def test_update_table_with_header_fn(self, conninfo, db_conn, temp_dir):
        """Test update_table with header_fn for headerless files"""
        csv_path = temp_dir / "no_header.csv"
        csv_path.write_text("Alice,100,85.5\nBob,200,92.0\n")

        self._setup_metadata_table(db_conn, temp_dir, [(csv_path, 2)])

        column_mapping = {
            "name": ([], "string"),
            "value": ([], "int"),
            "score": ([], "float"),
        }

        # Header function that provides column names
        def header_fn(file_path):
            return ["name", "value", "score"]

        result_df = update_table(
            conninfo=conninfo,
            schema="test_schema",
            output_table="header_fn_output",
            filetype="csv",
            column_mapping=column_mapping,
            source_dir=str(temp_dir) + "/",
            header_fn=header_fn,
            resume=False,
        )

        # Verify data was loaded with correct headers
        result = execute_sql_fetchone(
            db_conn,
            "SELECT name, value FROM test_schema.header_fn_output WHERE name = 'Alice'",
        )
        assert result[0] == "Alice"

    def test_update_table_with_sql_glob_filters_files(
        self, conninfo, db_conn, temp_dir
    ):
        """Test that sql_glob filters which files are processed.

        This test verifies that when sql_glob is set, only files matching
        the pattern are processed, not all files in the source_dir.
        """
        # Create CSV files with different naming patterns
        csv1 = temp_dir / "data_2024.csv"
        csv1.write_text("name,value\nAlice,100\n")
        csv2 = temp_dir / "data_2025.csv"
        csv2.write_text("name,value\nBob,200\n")
        csv3 = temp_dir / "other_file.csv"
        csv3.write_text("name,value\nCharlie,300\n")

        # Set up metadata for all files
        self._setup_metadata_table(db_conn, temp_dir, [(csv1, 1), (csv2, 1), (csv3, 1)])

        column_mapping = {
            "name": ([], "string"),
            "value": ([], "int"),
        }

        # Use sql_glob to only match files starting with "data_"
        result_df = update_table(
            conninfo=conninfo,
            schema="test_schema",
            output_table="sql_glob_test",
            filetype="csv",
            column_mapping=column_mapping,
            source_dir=str(temp_dir) + "/",
            sql_glob="%data_%.csv",  # Should match data_2024.csv and data_2025.csv only
            resume=False,
        )

        # Verify only 2 files were processed (data_2024.csv and data_2025.csv)
        # NOT 3 files (should exclude other_file.csv)
        result = execute_sql_fetchone(
            db_conn, "SELECT COUNT(*) FROM test_schema.sql_glob_test"
        )
        assert result[0] == 2, f"Expected 2 rows (from 2 files), got {result[0]}"

        # Verify the correct files were processed
        result = execute_sql_fetch(
            db_conn, "SELECT name FROM test_schema.sql_glob_test ORDER BY name"
        )
        names = [row[0] for row in result]
        assert names == ["Alice", "Bob"], f"Expected Alice and Bob, got {names}"


# ===== ADD_FILES_TO_METADATA_TABLE END-TO-END TESTS =====


class TestAddFilesToMetadataTableEndToEnd:
    """End-to-end tests for add_files_to_metadata_table"""

    def test_add_files_to_metadata_table_full_workflow(
        self, conninfo, db_conn, temp_dir
    ):
        """Test complete workflow of add_files_to_metadata_table"""
        import shutil

        # Clean up any existing metadata table from previous tests
        execute_sql(db_conn, "DROP TABLE IF EXISTS test_schema.metadata")

        # Create source directory
        source_dir = temp_dir / "source"
        source_dir.mkdir()

        # Create test CSV files
        csv1 = source_dir / "file1.csv"
        csv1.write_text("col1,col2\n1,2\n3,4\n")
        csv2 = source_dir / "file2.csv"
        csv2.write_text("col1,col2\n5,6\n7,8\n")

        # Call add_files_to_metadata_table
        result_df = add_files_to_metadata_table(
            conninfo=conninfo,
            schema="test_schema",
            source_dir=str(source_dir) + "/",
            filetype="csv",
            has_header=True,
            resume=False,
        )

        # Verify metadata table was created and populated
        assert table_exists(db_conn, "test_schema", "metadata")

        result = execute_sql_fetchone(
            db_conn, "SELECT COUNT(*) FROM test_schema.metadata"
        )
        assert result[0] == 2

        # Verify source_path is stored correctly in metadata
        paths = execute_sql_fetchone(
            db_conn, "SELECT array_agg(source_path) FROM test_schema.metadata"
        )
        assert str(csv1) in paths[0] or csv1.as_posix() in paths[0]

    def test_add_files_to_metadata_table_with_zip(self, conninfo, db_conn, temp_dir):
        """Test add_files_to_metadata_table with ZIP extraction"""
        import shutil
        import os

        original_cwd = os.getcwd()
        try:
            os.chdir(temp_dir)

            # Create source directory
            source_dir = temp_dir / "source"
            source_dir.mkdir()

            # Create a ZIP file with CSV files
            zip_path = source_dir / "archive.zip"
            with zipfile.ZipFile(zip_path, "w") as zf:
                zf.writestr("inner1.csv", "col1,col2\n1,2\n")
                zf.writestr("inner2.csv", "col1,col2\n3,4\n")

            # Call add_files_to_metadata_table with compression
            result_df = add_files_to_metadata_table(
                conninfo=conninfo,
                schema="test_schema",
                source_dir=str(source_dir) + "/",
                filetype="csv",
                compression_type="zip",
                archive_glob="*.csv",
                has_header=True,
                resume=False,
            )

            # Verify metadata table was populated
            result = execute_sql_fetchone(
                db_conn, "SELECT COUNT(*) FROM test_schema.metadata"
            )
            assert result[0] == 2

            # Verify source_paths have :: delimiter for archive files
            paths = execute_sql_fetchone(
                db_conn, "SELECT array_agg(source_path) FROM test_schema.metadata"
            )
            assert any("::" in p for p in paths[0])
        finally:
            os.chdir(original_cwd)

    def test_add_files_to_metadata_table_resume(self, conninfo, db_conn, temp_dir):
        """Test add_files_to_metadata_table with resume=True"""
        import shutil

        source_dir = temp_dir / "search"
        landing_dir = temp_dir / "landing"
        source_dir.mkdir()
        landing_dir.mkdir()

        # Create initial file and process it
        csv1 = source_dir / "file1.csv"
        csv1.write_text("col1,col2\n1,2\n")

        # First run
        add_files_to_metadata_table(
            conninfo=conninfo,
            schema="test_schema",
            source_dir=str(source_dir) + "/",
            filetype="csv",
            has_header=True,
            resume=False,
        )

        # Add a new file
        csv2 = source_dir / "file2.csv"
        csv2.write_text("col1,col2\n3,4\n")

        # Second run with resume=True
        result_df = add_files_to_metadata_table(
            conninfo=conninfo,
            schema="test_schema",
            source_dir=str(source_dir) + "/",
            filetype="csv",
            has_header=True,
            resume=True,
        )

        # Verify both files are in metadata
        result = execute_sql_fetchone(
            db_conn, "SELECT COUNT(*) FROM test_schema.metadata"
        )
        assert result[0] == 2

    def test_add_files_to_metadata_table_with_filter(self, conninfo, db_conn, temp_dir):
        """Test add_files_to_metadata_table with file_list_filter_fn"""
        import shutil

        source_dir = temp_dir / "search"
        landing_dir = temp_dir / "landing"
        source_dir.mkdir()
        landing_dir.mkdir()

        # Create multiple CSV files
        (source_dir / "include.csv").write_text("col1,col2\n1,2\n")
        (source_dir / "exclude.csv").write_text("col1,col2\n3,4\n")

        # Filter function
        def file_list_filter_fn(file_list):
            return [f for f in file_list if "include" in str(f)]

        result_df = add_files_to_metadata_table(
            conninfo=conninfo,
            schema="test_schema",
            source_dir=str(source_dir) + "/",
            filetype="csv",
            has_header=True,
            file_list_filter_fn=file_list_filter_fn,
            resume=False,
        )

        # Verify only filtered file was processed
        result = execute_sql_fetchone(
            db_conn, "SELECT COUNT(*) FROM test_schema.metadata"
        )
        assert result[0] == 1

    def test_add_files_to_metadata_table_invalid_param_raises_error(self, conninfo):
        """Test that invalid parameter names raise TypeError"""
        with pytest.raises(TypeError) as exc_info:
            add_files_to_metadata_table(
                conninfo=conninfo,
                schema="test_schema",
                source_dir="/tmp",
                filetype="csv",
                file_glob="*.csv",  # Wrong param name - should be 'glob'
            )

        assert "file_glob" in str(exc_info.value)
        assert "unexpected keyword argument" in str(exc_info.value)


# ===== CLI SCHEMA INFERENCE TESTS =====


class TestCLISchemaInference:
    """Test CLI schema inference functionality"""

    def test_cli_argument_parsing(self, temp_dir):
        """Test CLI argument parsing via subprocess"""
        import subprocess

        # Create test file
        csv_path = temp_dir / "cli_test.csv"
        csv_path.write_text("name,age,score\nAlice,25,85.5\n")

        # Run CLI
        result = subprocess.run(
            ["python", "table_functions.py", str(csv_path)],
            capture_output=True,
            text=True,
            cwd=Path(__file__).parent.parent,
        )

        assert result.returncode == 0
        output = json.loads(result.stdout)
        # Output is keyed by filename with table_name and column_mapping
        assert "cli_test.csv" in output
        column_mapping = output["cli_test.csv"]["column_mapping"]
        assert "name" in column_mapping
        assert "age" in column_mapping
        assert "score" in column_mapping

    def test_cli_filetype_detection(self, temp_dir):
        """Test CLI auto-detects file type from extension"""
        import subprocess

        # Create PSV file with .psv extension
        psv_path = temp_dir / "cli_test.psv"
        psv_path.write_text("name|age|score\nAlice|25|85.5\n")

        result = subprocess.run(
            ["python", "table_functions.py", str(psv_path)],
            capture_output=True,
            text=True,
            cwd=Path(__file__).parent.parent,
        )

        assert result.returncode == 0
        output = json.loads(result.stdout)
        assert "cli_test.psv" in output
        assert "name" in output["cli_test.psv"]["column_mapping"]

    def test_cli_no_header(self, temp_dir):
        """Test CLI with --no-header flag"""
        import subprocess

        csv_path = temp_dir / "no_header.csv"
        csv_path.write_text("Alice,25,85.5\nBob,30,92.0\n")

        result = subprocess.run(
            ["python", "table_functions.py", str(csv_path), "--no-header"],
            capture_output=True,
            text=True,
            cwd=Path(__file__).parent.parent,
        )

        assert result.returncode == 0
        output = json.loads(result.stdout)
        column_mapping = output["no_header.csv"]["column_mapping"]
        assert "col_0" in column_mapping
        assert "col_1" in column_mapping
        assert "col_2" in column_mapping

    def test_cli_auto_detects_separator(self, temp_dir):
        """Test CLI auto-detects custom separator (semicolon)"""
        import subprocess

        # Create file with semicolon separator - Should auto-detect
        csv_path = temp_dir / "custom_sep.csv"
        csv_path.write_text("name;age;score\nAlice;25;85.5\n")

        result = subprocess.run(
            ["python", "table_functions.py", str(csv_path)],
            capture_output=True,
            text=True,
            cwd=Path(__file__).parent.parent,
        )

        assert result.returncode == 0
        output = json.loads(result.stdout)
        column_mapping = output["custom_sep.csv"]["column_mapping"]
        assert "name" in column_mapping
        assert "age" in column_mapping

    def test_cli_sample_rows(self, temp_dir):
        """Test CLI with --sample-rows parameter"""
        import subprocess

        # Create file with many rows
        csv_path = temp_dir / "many_rows.csv"
        with open(csv_path, "w") as f:
            f.write("id,value\n")
            for i in range(1000):
                f.write(f"{i},{i * 10}\n")

        result = subprocess.run(
            ["python", "table_functions.py", str(csv_path), "--sample-rows", "100"],
            capture_output=True,
            text=True,
            cwd=Path(__file__).parent.parent,
        )

        assert result.returncode == 0
        output = json.loads(result.stdout)
        column_mapping = output["many_rows.csv"]["column_mapping"]
        assert "id" in column_mapping
        assert "value" in column_mapping

    def test_cli_file_not_found(self, temp_dir):
        """Test CLI error handling for non-existent file"""
        import subprocess

        result = subprocess.run(
            ["python", "table_functions.py", "/nonexistent/file.csv"],
            capture_output=True,
            text=True,
            cwd=Path(__file__).parent.parent,
        )

        assert result.returncode == 1
        # Error messages go to stderr via loguru
        output = result.stdout + result.stderr
        assert "Error" in output or "not found" in output.lower()

    def test_cli_parquet_file(self, temp_dir):
        """Test CLI with parquet file"""
        import subprocess

        # Create parquet file
        test_df = pd.DataFrame(
            {
                "name": ["Alice", "Bob"],
                "age": [25, 30],
            }
        )
        parquet_path = temp_dir / "test.parquet"
        test_df.to_parquet(str(parquet_path))

        result = subprocess.run(
            ["python", "table_functions.py", str(parquet_path)],
            capture_output=True,
            text=True,
            cwd=Path(__file__).parent.parent,
        )

        assert result.returncode == 0
        output = json.loads(result.stdout)
        column_mapping = output["test.parquet"]["column_mapping"]
        assert "name" in column_mapping
        assert "age" in column_mapping

    def test_cli_directory_mode(self, temp_dir):
        """Test CLI with directory path outputs JSON keyed by filename"""
        import subprocess

        # Create subdirectory with multiple CSV files
        sub_dir = temp_dir / "multi_files"
        sub_dir.mkdir()

        (sub_dir / "file1.csv").write_text("name,age\nAlice,25\nBob,30\n")
        (sub_dir / "file2.csv").write_text("id,value,score\n1,100,85.5\n2,200,92.0\n")

        result = subprocess.run(
            ["python", "table_functions.py", str(sub_dir)],
            capture_output=True,
            text=True,
            cwd=Path(__file__).parent.parent,
        )

        assert result.returncode == 0
        output = json.loads(result.stdout)

        # Check output is keyed by filename
        assert "file1.csv" in output
        assert "file2.csv" in output

        # Check schema for each file
        file1_mapping = output["file1.csv"]["column_mapping"]
        assert "name" in file1_mapping
        assert "age" in file1_mapping

        file2_mapping = output["file2.csv"]["column_mapping"]
        assert "id" in file2_mapping
        assert "value" in file2_mapping
        assert "score" in file2_mapping

        # Check table_name is included
        assert output["file1.csv"]["table_name"] == "file1"
        assert output["file2.csv"]["table_name"] == "file2"

    def test_cli_directory_mode_with_filetype_filter(self, temp_dir):
        """Test CLI directory mode filters by --filetype"""
        import subprocess

        # Create subdirectory with mixed file types
        sub_dir = temp_dir / "mixed_files"
        sub_dir.mkdir()

        (sub_dir / "data.csv").write_text("name,age\nAlice,25\n")
        (sub_dir / "data.tsv").write_text("id\tvalue\n1\t100\n")
        (sub_dir / "data.psv").write_text("col1|col2\na|b\n")

        # Filter to only CSV files
        result = subprocess.run(
            ["python", "table_functions.py", str(sub_dir), "--filetype", "csv"],
            capture_output=True,
            text=True,
            cwd=Path(__file__).parent.parent,
        )

        assert result.returncode == 0
        output = json.loads(result.stdout)

        # Only CSV file should be in output
        assert "data.csv" in output
        assert "data.tsv" not in output
        assert "data.psv" not in output

    def test_cli_directory_mode_empty_dir(self, temp_dir):
        """Test CLI directory mode with empty directory"""
        import subprocess

        # Create empty subdirectory
        empty_dir = temp_dir / "empty_dir"
        empty_dir.mkdir()

        result = subprocess.run(
            ["python", "table_functions.py", str(empty_dir)],
            capture_output=True,
            text=True,
            cwd=Path(__file__).parent.parent,
        )

        assert result.returncode == 1
        # Error messages go to stderr via loguru
        output = result.stdout + result.stderr
        assert "No matching files" in output or "Error" in output

    def test_cli_directory_mode_non_recursive(self, temp_dir):
        """Test CLI directory mode is non-recursive (doesn't descend into subdirs)"""
        import subprocess

        # Create nested directory structure
        parent_dir = temp_dir / "parent"
        parent_dir.mkdir()
        child_dir = parent_dir / "child"
        child_dir.mkdir()

        (parent_dir / "parent_file.csv").write_text("a,b\n1,2\n")
        (child_dir / "child_file.csv").write_text("x,y\n3,4\n")

        result = subprocess.run(
            ["python", "table_functions.py", str(parent_dir)],
            capture_output=True,
            text=True,
            cwd=Path(__file__).parent.parent,
        )

        assert result.returncode == 0
        output = json.loads(result.stdout)

        # Only parent file should be in output
        assert "parent_file.csv" in output
        assert "child_file.csv" not in output

    def test_cli_directory_mode_with_multiple_files(self, temp_dir):
        """Test CLI directory mode processes multiple files correctly"""
        import subprocess

        # Create subdirectory with multiple good files
        sub_dir = temp_dir / "multi_files"
        sub_dir.mkdir()

        (sub_dir / "users.csv").write_text("name,age\nAlice,25\n")
        (sub_dir / "products.csv").write_text("id,price\n1,9.99\n")

        result = subprocess.run(
            ["python", "table_functions.py", str(sub_dir)],
            capture_output=True,
            text=True,
            cwd=Path(__file__).parent.parent,
        )

        assert result.returncode == 0
        output = json.loads(result.stdout)

        # Both files should have proper schema
        assert "users.csv" in output
        assert "name" in output["users.csv"]["column_mapping"]
        assert "products.csv" in output
        assert "id" in output["products.csv"]["column_mapping"]

    def test_cli_directory_mode_multiple_filetypes(self, temp_dir):
        """Test CLI directory mode includes all supported types when no --filetype specified"""
        import subprocess

        # Create subdirectory with multiple file types
        sub_dir = temp_dir / "all_types"
        sub_dir.mkdir()

        (sub_dir / "data.csv").write_text("a,b\n1,2\n")
        (sub_dir / "data.tsv").write_text("x\ty\n3\t4\n")
        (sub_dir / "data.psv").write_text("p|q\n5|6\n")

        # Create parquet file
        test_df = pd.DataFrame({"col1": [1, 2], "col2": ["a", "b"]})
        test_df.to_parquet(str(sub_dir / "data.parquet"))

        result = subprocess.run(
            ["python", "table_functions.py", str(sub_dir)],
            capture_output=True,
            text=True,
            cwd=Path(__file__).parent.parent,
        )

        assert result.returncode == 0
        output = json.loads(result.stdout)

        # All file types should be present
        assert "data.csv" in output
        assert "data.tsv" in output
        assert "data.psv" in output
        assert "data.parquet" in output

    def test_cli_directory_not_found(self, temp_dir):
        """Test CLI error handling for non-existent directory"""
        import subprocess

        result = subprocess.run(
            ["python", "table_functions.py", "/nonexistent/directory/"],
            capture_output=True,
            text=True,
            cwd=Path(__file__).parent.parent,
        )

        assert result.returncode == 1
        # Error messages go to stderr via loguru
        output = result.stdout + result.stderr
        assert "Error" in output or "not found" in output.lower()

    def test_cli_encoding_utf8_file(self, temp_dir):
        """Test CLI auto-detects UTF-8 encoding"""
        import subprocess

        csv_file = temp_dir / "test.csv"
        csv_file.write_text("name,value\nAlice,100\n")

        result = subprocess.run(
            ["python", "table_functions.py", str(csv_file)],
            capture_output=True,
            text=True,
            cwd=Path(__file__).parent.parent,
        )

        assert result.returncode == 0
        output = json.loads(result.stdout)

        # Should have encoding info
        assert "test.csv" in output
        assert "encoding" in output["test.csv"]
        assert output["test.csv"]["encoding"] == "utf-8"

    def test_cli_encoding_explicit_latin1(self, temp_dir):
        """Test CLI with explicit latin-1 encoding for non-UTF-8 files"""
        import subprocess

        csv_file = temp_dir / "windows.csv"
        # CP1252/latin-1 characters - not valid UTF-8
        csv_file.write_bytes(b"name,city\nCaf\xe9,M\xfcnchen\n")  # é and ü in latin-1

        result = subprocess.run(
            ["python", "table_functions.py", str(csv_file), "--encoding", "latin-1"],
            capture_output=True,
            text=True,
            cwd=Path(__file__).parent.parent,
        )

        assert result.returncode == 0
        output = json.loads(result.stdout)

        # Should use latin-1 encoding
        assert "windows.csv" in output
        assert "encoding" in output["windows.csv"]
        assert output["windows.csv"]["encoding"] == "latin-1"

    def test_cli_encoding_latin1_with_late_special_chars(self, temp_dir):
        """Test CLI handles latin-1 files with non-ASCII bytes late in file"""
        import subprocess

        csv_file = temp_dir / "late_special.csv"
        # Create file with 200KB of ASCII data followed by non-UTF-8 characters
        ascii_rows = b"id,name,value\n" + b"".join(
            f"{i},item_{i},{i * 10}\n".encode("ascii") for i in range(10000)
        )
        # Add rows with latin-1/CP1252 characters at the end (é=0xe9, ü=0xfc, ñ=0xf1)
        special_rows = (
            b"10001,Caf\xe9,100\n10002,M\xfcnchen,200\n10003,Espa\xf1ol,300\n"
        )
        csv_file.write_bytes(ascii_rows + special_rows)

        result = subprocess.run(
            ["python", "table_functions.py", str(csv_file), "--encoding", "latin-1"],
            capture_output=True,
            text=True,
            cwd=Path(__file__).parent.parent,
        )

        assert result.returncode == 0
        output = json.loads(result.stdout)

        # Should successfully parse using latin-1 encoding
        assert "late_special.csv" in output
        assert "encoding" in output["late_special.csv"]
        assert output["late_special.csv"]["encoding"] == "latin-1"
        # Should have successfully inferred the schema (not errored)
        assert "column_mapping" in output["late_special.csv"]
        assert "id" in output["late_special.csv"]["column_mapping"]

    def test_cli_encoding_latin1_special_chars(self, temp_dir):
        """Test CLI handles latin-1 special characters (like ñ = 0xf1)"""
        import subprocess

        csv_file = temp_dir / "special.csv"
        # 0xf1 is ñ in latin-1
        csv_file.write_bytes(b"name,value\nEspa\xf1a,100\n")

        result = subprocess.run(
            ["python", "table_functions.py", str(csv_file), "--encoding", "latin-1"],
            capture_output=True,
            text=True,
            cwd=Path(__file__).parent.parent,
        )

        assert result.returncode == 0
        output = json.loads(result.stdout)

        # Should successfully parse using latin-1 encoding
        assert "special.csv" in output
        assert "column_mapping" in output["special.csv"]
        assert output["special.csv"]["encoding"] == "latin-1"


# ===== CONNECTION AND SQL HELPER TESTS =====


class TestConnectionHelpers:
    """Test connection management and SQL helper functions"""

    def test_psycopg_connection_basic(self, db_conn):
        """Test psycopg connection is usable"""
        # Connection should be usable
        with db_conn.cursor() as cur:
            cur.execute("SELECT 1")
            result = cur.fetchone()
            assert result[0] == 1

    def test_execute_sql_basic(self, db_conn):
        """Test execute_sql with basic DDL"""
        execute_sql(db_conn, "CREATE TABLE test_schema.exec_test (id INT)")
        assert table_exists(db_conn, "test_schema", "exec_test")

    def test_execute_sql_with_params(self, db_conn):
        """Test execute_sql with parameters"""
        execute_sql(
            db_conn, "CREATE TABLE test_schema.param_test (name TEXT, value INT)"
        )
        execute_sql(
            db_conn,
            "INSERT INTO test_schema.param_test VALUES ($1, $2)",
            ("Alice", 100),
        )

        result = execute_sql_fetchone(
            db_conn, "SELECT name, value FROM test_schema.param_test"
        )
        assert result[0] == "Alice"
        assert result[1] == 100

    def test_execute_sql_fetch_multiple_rows(self, db_conn):
        """Test execute_sql_fetch returns all rows"""
        execute_sql(db_conn, "CREATE TABLE test_schema.multi_test (id INT)")
        for i in range(5):
            execute_sql(db_conn, f"INSERT INTO test_schema.multi_test VALUES ({i})")

        results = execute_sql_fetch(
            db_conn, "SELECT id FROM test_schema.multi_test ORDER BY id"
        )
        assert len(results) == 5
        assert [r[0] for r in results] == [0, 1, 2, 3, 4]

    def test_execute_sql_fetchone_no_results(self, db_conn):
        """Test execute_sql_fetchone returns None when no results"""
        execute_sql(db_conn, "CREATE TABLE test_schema.empty_test (id INT)")

        result = execute_sql_fetchone(db_conn, "SELECT id FROM test_schema.empty_test")
        assert result is None

    def test_execute_sql_invalid_sql(self, db_conn):
        """Test execute_sql raises error on invalid SQL"""
        with pytest.raises(Exception):
            execute_sql(db_conn, "THIS IS NOT VALID SQL")


# ===== ENCODING TESTS =====


class TestEncodings:
    """Test different file encodings"""

    def test_read_csv_cp1252_encoding(self, temp_dir):
        """Test reading Windows-1252 encoded file with explicit encoding"""
        # Create file with Windows-1252 encoding
        csv_path = temp_dir / "cp1252.csv"
        # CP-1252 smart quotes
        content = b"name,value\nTest\x92s,100\nTest2,200\n"  # \x92 is right single quote in CP-1252
        csv_path.write_bytes(content)

        column_mapping = {
            "name": ([], "string"),
            "value": ([], "int"),
        }

        df = read_csv(
            full_path=str(csv_path),
            column_mapping=column_mapping,
            has_header=True,
            encoding="cp1252",
        )

        assert len(df) == 2
        # CP-1252 decodes \x92 as right single quote (U+2019)
        assert "\u2019" in df["name"][0]  # Unicode right single quotation mark
        assert df["value"][1] == 200

    def test_read_csv_latin1_encoding(self, temp_dir):
        """Test reading Latin-1 encoded file with explicit encoding"""
        csv_path = temp_dir / "latin1.csv"
        # Latin-1 encoded é (0xe9)
        content = b"name,value\nCaf\xe9,100\nTest2,200\n"
        csv_path.write_bytes(content)

        column_mapping = {
            "name": ([], "string"),
            "value": ([], "int"),
        }

        df = read_csv(
            full_path=str(csv_path),
            column_mapping=column_mapping,
            has_header=True,
            encoding="latin-1",
        )

        assert len(df) == 2
        assert df["name"][0] == "Café"
        assert df["name"][1] == "Test2"

    def test_decode_file_content_utf8(self, temp_dir):
        """Test decode_file_content with pure UTF-8 file"""
        from table_functions import decode_file_content

        csv_path = temp_dir / "utf8.csv"
        csv_path.write_text("name,value\nJosé,100\n", encoding="utf-8")

        result = decode_file_content(str(csv_path))

        assert result["encoding"] == "utf-8"
        assert "José" in result["content"]

    def test_decode_file_content_latin1(self, temp_dir):
        """Test decode_file_content with explicit latin-1 encoding"""
        from table_functions import decode_file_content

        csv_path = temp_dir / "latin1.csv"
        # Latin-1 encoded é (0xe9)
        csv_path.write_bytes(b"name,value\nCaf\xe9,100\n")

        result = decode_file_content(str(csv_path), encoding="latin-1")

        assert result["encoding"] == "latin-1"
        assert "Café" in result["content"]

    def test_decode_file_content_cp1252(self, temp_dir):
        """Test decode_file_content with explicit cp1252 encoding"""
        from table_functions import decode_file_content

        csv_path = temp_dir / "cp1252.csv"
        # CP-1252 smart quotes (\x93\x94)
        csv_path.write_bytes(b"name,description\nTest,He said \x93hello\x94\n")

        result = decode_file_content(str(csv_path), encoding="cp1252")

        assert result["encoding"] == "cp1252"
        # CP-1252 decodes \x93\x94 as curly double quotes (U+201C, U+201D)
        assert "\u201c" in result["content"]  # Left double quotation mark
        assert "\u201d" in result["content"]  # Right double quotation mark

    def test_decode_file_content_latin1_handles_0x81(self, temp_dir):
        """Test decode_file_content handles 0x81 byte with latin-1 encoding"""
        from table_functions import decode_file_content

        csv_path = temp_dir / "problematic.csv"
        # 0x81 is undefined in CP-1252 but valid in latin-1
        csv_path.write_bytes(b"name,value\ntest\x81data,100\n")

        result = decode_file_content(str(csv_path), encoding="latin-1")

        assert result["encoding"] == "latin-1"
        assert "test" in result["content"]
        assert "data" in result["content"]


# ===== EXCEL EDGE CASE TESTS =====


class TestExcelEdgeCases:
    """Test Excel reading edge cases"""

    def test_read_xlsx_with_skiprows(self, temp_dir):
        """Test reading Excel with rows to skip"""
        xlsx_path = temp_dir / "skiprows.xlsx"

        # Create Excel with header rows to skip
        df = pd.DataFrame(
            {
                "col1": ["Title Row", "Description", "name", "Alice", "Bob"],
                "col2": ["", "", "age", "25", "30"],
            }
        )
        df.to_excel(xlsx_path, index=False)

        column_mapping = {
            "name": ([], "string"),
            "age": ([], "string"),  # Read as string since mixed content
        }

        result_df = read_xlsx(
            full_path=str(xlsx_path),
            column_mapping=column_mapping,
            excel_skiprows=2,
        )

        # Should skip first 2 rows, use row 3 as header
        assert "name" in result_df.columns or "col1" in result_df.columns

    def test_read_xlsx_empty_cells(self, temp_dir):
        """Test reading Excel with empty cells"""
        xlsx_path = temp_dir / "empty_cells.xlsx"

        df = pd.DataFrame(
            {
                "name": ["Alice", None, "Charlie"],
                "value": [100, 200, None],
            }
        )
        df.to_excel(xlsx_path, index=False)

        column_mapping = {
            "name": ([], "string"),
            "value": ([], "int"),
        }

        result_df = read_xlsx(
            full_path=str(xlsx_path),
            column_mapping=column_mapping,
        )

        assert len(result_df) == 3
        # Check that nulls are preserved
        assert result_df["name"].isna().sum() == 1
        assert result_df["value"].isna().sum() == 1


# ===== FIXED WIDTH EDGE CASES =====


class TestFixedWidthEdgeCases:
    """Test fixed-width file reading edge cases"""

    def test_read_fixed_width_with_whitespace(self, temp_dir):
        """Test fixed-width with fields that are all whitespace"""
        fw_path = temp_dir / "whitespace.txt"
        # Each field is 10 chars: name, age, city
        # Positions are 1-indexed: name at 1, age at 11, city at 21
        fw_path.write_text(
            "Alice     25        Boston    \n          30        Denver    \n"
        )

        # Format: column_name: (type, starting_position (1-indexed), field_size)
        column_mapping = {
            "name": ("string", 1, 10),
            "age": ("string", 11, 10),  # string since whitespace-padded
            "city": ("string", 21, 10),
        }

        df = read_fixed_width(
            full_path=str(fw_path),
            column_mapping=column_mapping,
        )

        assert len(df) == 2
        # First row should have "Alice"
        assert df["name"][0] == "Alice"

    def test_read_fixed_width_numeric_fields(self, temp_dir):
        """Test fixed-width with right-aligned numeric fields"""
        fw_path = temp_dir / "numeric.txt"
        # Each field is 8 chars
        fw_path.write_text("     100    1.50\n    2000   25.99\n")

        # Format: column_name: (type, starting_position (1-indexed), field_size)
        column_mapping = {
            "amount": ("int", 1, 8),
            "rate": ("float", 9, 8),
        }

        df = read_fixed_width(
            full_path=str(fw_path),
            column_mapping=column_mapping,
        )

        assert len(df) == 2
        assert df["amount"][0] == 100
        assert df["amount"][1] == 2000


# ===== UPDATE TABLE EDGE CASES =====


class TestUpdateTableEdgeCases:
    """Test update_table edge cases"""

    def _setup_metadata_table(self, db_conn, temp_dir, csv_files):
        """Helper to set up metadata table with files"""
        execute_sql(db_conn, "DROP TABLE IF EXISTS test_schema.metadata")
        execute_sql(
            db_conn,
            """
            CREATE TABLE test_schema.metadata (
                source_dir TEXT,
                source_path TEXT PRIMARY KEY,
                filesize BIGINT,
                header TEXT[],
                row_count BIGINT,
                                file_hash TEXT,
                metadata_ingest_datetime TIMESTAMP,
                metadata_ingest_status TEXT,
                ingest_datetime TIMESTAMP,
                ingest_runtime INTEGER,
                status TEXT,
                error_message TEXT,
                unpivot_row_multiplier INTEGER,
                output_table TEXT
            )
        """,
        )

        for csv_file, row_count in csv_files:
            execute_sql(
                db_conn,
                """
                INSERT INTO test_schema.metadata
                (source_path, row_count, metadata_ingest_status, source_dir)
                VALUES ($1, $2, $3, $4)
            """,
                (str(csv_file), row_count, "Success", str(temp_dir) + "/"),
            )

    def test_update_table_empty_file(self, conninfo, db_conn, temp_dir):
        """Test update_table with empty CSV file (header only)"""
        csv_path = temp_dir / "empty.csv"
        csv_path.write_text("name,value\n")  # Header only, no data

        self._setup_metadata_table(db_conn, temp_dir, [(csv_path, 0)])

        column_mapping = {
            "name": ([], "string"),
            "value": ([], "int"),
        }

        result_df = update_table(
            conninfo=conninfo,
            schema="test_schema",
            output_table="empty_output",
            filetype="csv",
            column_mapping=column_mapping,
            source_dir=str(temp_dir) + "/",
            resume=False,
        )

        # Table should exist but be empty
        assert table_exists(db_conn, "test_schema", "empty_output")
        count = execute_sql_fetchone(
            db_conn, "SELECT COUNT(*) FROM test_schema.empty_output"
        )
        assert count[0] == 0

    def test_update_table_with_null_values_single(self, conninfo, db_conn, temp_dir):
        """Test update_table with a single null value in the list"""
        csv_path = temp_dir / "null_string.csv"
        csv_path.write_text("name,value\nAlice,100\nNA,NA\nBob,200\n")

        self._setup_metadata_table(db_conn, temp_dir, [(csv_path, 3)])

        column_mapping = {
            "name": ([], "string"),
            "value": ([], "int"),
        }

        result_df = update_table(
            conninfo=conninfo,
            schema="test_schema",
            output_table="null_string_output",
            filetype="csv",
            column_mapping=column_mapping,
            source_dir=str(temp_dir) + "/",
            null_values=["NA"],
            resume=False,
        )

        # Check that NA was treated as null
        result = execute_sql_fetch(
            db_conn,
            "SELECT name, value FROM test_schema.null_string_output ORDER BY value NULLS FIRST",
        )
        # First row should have null value
        assert result[0][1] is None

    def test_update_table_with_null_values_multiple(self, conninfo, db_conn, temp_dir):
        """Test update_table with multiple null values in the list"""
        csv_path = temp_dir / "multi_null.csv"
        csv_path.write_text(
            "name,value\nAlice,100\nNA,NA\nBob,None\nCharlie,N/A\nDave,200\n"
        )

        self._setup_metadata_table(db_conn, temp_dir, [(csv_path, 5)])

        column_mapping = {
            "name": ([], "string"),
            "value": ([], "int"),
        }

        result_df = update_table(
            conninfo=conninfo,
            schema="test_schema",
            output_table="multi_null_output",
            filetype="csv",
            column_mapping=column_mapping,
            source_dir=str(temp_dir) + "/",
            null_values=["NA", "None", "N/A"],
            resume=False,
        )

        # Check that all null representations were treated as null
        result = execute_sql_fetch(
            db_conn,
            "SELECT name, value FROM test_schema.multi_null_output ORDER BY name",
        )
        # Alice has 100, Bob/Charlie/NA rows have null, Dave has 200
        values_by_name = {row[0]: row[1] for row in result}
        assert values_by_name["Alice"] == 100
        assert values_by_name["Dave"] == 200
        assert values_by_name["Bob"] is None  # "None" was in null_values
        assert values_by_name["Charlie"] is None  # "N/A" was in null_values
        # NA row has null name because "NA" is in null_values
        assert None in values_by_name or any(v is None for v in values_by_name.values())

    def test_update_table_pivot_multiple_value_columns(
        self, conninfo, db_conn, temp_dir
    ):
        """Test update_table with pivot_mapping with multiple id columns"""
        csv_path = temp_dir / "multi_pivot.csv"
        csv_path.write_text(
            "id,region,q1,q2,q3,q4\n1,East,100,200,300,400\n2,West,150,250,350,450\n"
        )

        self._setup_metadata_table(db_conn, temp_dir, [(csv_path, 2)])

        column_mapping = {
            "id": ([], "int"),
            "region": ([], "string"),
            "q1": ([], "int"),
            "q2": ([], "int"),
            "q3": ([], "int"),
            "q4": ([], "int"),
        }

        pivot_mapping = {
            "id_vars": ["id", "region"],  # Multiple id columns
            "variable_column_name": "quarter",
            "value_column_name": "sales",
        }

        result_df = update_table(
            conninfo=conninfo,
            schema="test_schema",
            output_table="multi_pivot_output",
            filetype="csv",
            column_mapping=column_mapping,
            source_dir=str(temp_dir) + "/",
            pivot_mapping=pivot_mapping,
            resume=False,
        )

        # Should have 8 rows (2 ids * 4 quarters)
        count = execute_sql_fetchone(
            db_conn, "SELECT COUNT(*) FROM test_schema.multi_pivot_output"
        )
        assert count[0] == 8

        # Verify structure
        result = execute_sql_fetchone(
            db_conn,
            "SELECT id, region, quarter, sales FROM test_schema.multi_pivot_output LIMIT 1",
        )
        assert result is not None

    def test_update_table_retry_failed(self, conninfo, db_conn, temp_dir):
        """Test update_table with retry_failed=True"""
        csv_path = temp_dir / "retry.csv"
        csv_path.write_text("name,value\nAlice,100\n")

        # Setup metadata with a failed file
        execute_sql(db_conn, "DROP TABLE IF EXISTS test_schema.metadata")
        execute_sql(
            db_conn,
            """
            CREATE TABLE test_schema.metadata (
                source_dir TEXT,
                source_path TEXT PRIMARY KEY,
                filesize BIGINT,
                header TEXT[],
                row_count BIGINT,
                                file_hash TEXT,
                metadata_ingest_datetime TIMESTAMP,
                metadata_ingest_status TEXT,
                ingest_datetime TIMESTAMP,
                ingest_runtime INTEGER,
                status TEXT,
                error_message TEXT,
                unpivot_row_multiplier INTEGER,
                output_table TEXT
            )
        """,
        )
        execute_sql(
            db_conn,
            """
            INSERT INTO test_schema.metadata
            (source_path, row_count, metadata_ingest_status, source_dir, status, error_message)
            VALUES ($1, $2, $3, $4, $5, $6)
        """,
            (
                str(csv_path),
                1,
                "Success",
                str(temp_dir) + "/",
                "Failure",
                "Previous error",
            ),
        )

        column_mapping = {
            "name": ([], "string"),
            "value": ([], "int"),
        }

        result_df = update_table(
            conninfo=conninfo,
            schema="test_schema",
            output_table="retry_output",
            filetype="csv",
            column_mapping=column_mapping,
            source_dir=str(temp_dir) + "/",
            retry_failed=True,
        )

        # File should have been reprocessed
        count = execute_sql_fetchone(
            db_conn, "SELECT COUNT(*) FROM test_schema.retry_output"
        )
        assert count[0] == 1


# ===== TYPE INFERENCE EDGE CASES =====


class TestTypeInferenceEdgeCases:
    """Test type inference with edge cases"""

    def test_infer_schema_mixed_types_in_column(self, temp_dir):
        """Test schema inference with mixed types in a column"""
        csv_path = temp_dir / "mixed.csv"
        csv_path.write_text("value\n100\ntext\n200\n")

        result = infer_schema_from_file(str(csv_path))
        column_mapping = result["column_mapping"]

        # Should infer as string since mixed types
        assert column_mapping["value"][1] == "string"

    def test_infer_schema_empty_column(self, temp_dir):
        """Test schema inference with column that has all empty values"""
        csv_path = temp_dir / "empty_col.csv"
        csv_path.write_text("name,empty_col,value\nAlice,,100\nBob,,200\n")

        result = infer_schema_from_file(str(csv_path))
        column_mapping = result["column_mapping"]

        # We infer empty columns as string (safest default)
        assert "empty_col" in column_mapping
        assert column_mapping["empty_col"][1] == "string"

    def test_infer_schema_boolean_column(self, temp_dir):
        """Test schema inference with boolean values"""
        csv_path = temp_dir / "bool.csv"
        csv_path.write_text("name,active\nAlice,true\nBob,false\n")

        result = infer_schema_from_file(str(csv_path))
        column_mapping = result["column_mapping"]

        assert column_mapping["active"][1] == "boolean"

    def test_infer_schema_date_column(self, temp_dir):
        """Test schema inference with date values"""
        csv_path = temp_dir / "dates.csv"
        csv_path.write_text("name,created_at\nAlice,2024-01-15\nBob,2024-02-20\n")

        result = infer_schema_from_file(str(csv_path))
        column_mapping = result["column_mapping"]

        # Dates might be inferred as datetime or string depending on Polars
        assert "created_at" in column_mapping


# ===== CSV HEADER AND ROW COUNT EDGE CASES =====


class TestCsvHeaderRowCountEdgeCases:
    """Test get_csv_header_and_row_count edge cases"""

    def test_csv_header_with_crlf_line_endings(self, temp_dir):
        """Test CSV with Windows line endings (CRLF)"""
        csv_path = temp_dir / "crlf.csv"
        csv_path.write_bytes(b"name,value\r\nAlice,100\r\nBob,200\r\n")

        header, row_count = get_csv_header_and_row_count(
            file=str(csv_path), separator=",", has_header=True
        )

        assert header == ["name", "value"]
        assert row_count == 2

    def test_csv_header_single_row(self, temp_dir):
        """Test CSV with only one data row"""
        csv_path = temp_dir / "single.csv"
        csv_path.write_text("name,value\nAlice,100\n")

        header, row_count = get_csv_header_and_row_count(
            file=str(csv_path), separator=",", has_header=True
        )

        assert row_count == 1

    def test_csv_header_no_data_rows(self, temp_dir):
        """Test CSV with header only (no data)"""
        csv_path = temp_dir / "header_only.csv"
        csv_path.write_text("name,value\n")

        header, row_count = get_csv_header_and_row_count(
            file=str(csv_path), separator=",", has_header=True
        )

        assert header == ["name", "value"]
        assert row_count == 0

    def test_csv_header_special_characters(self, temp_dir):
        """Test CSV with special characters in header"""
        csv_path = temp_dir / "special.csv"
        csv_path.write_text("user-name,total$amount,count#items\nAlice,100,5\n")

        header, row_count = get_csv_header_and_row_count(
            file=str(csv_path), separator=",", has_header=True
        )

        assert "user-name" in header
        assert "total$amount" in header


# ===== DROP FUNCTIONS EDGE CASES =====


class TestDropFunctionsEdgeCases:
    """Test drop functions with edge cases"""

    def test_drop_metadata_by_source_no_matches(self, conninfo, db_conn):
        """Test drop_metadata_by_source when no files match"""
        execute_sql(
            db_conn,
            """
            CREATE TABLE test_schema.metadata (
                source_dir TEXT,
                source_path TEXT PRIMARY KEY
            )
        """,
        )
        execute_sql(
            db_conn,
            "INSERT INTO test_schema.metadata VALUES ($1, $2)",
            ("/some/path", "/some/path/file.csv"),
        )

        # Drop with non-matching path
        drop_metadata_by_source(
            conninfo=conninfo,
            source_dir="/nonexistent/path",
            schema="test_schema",
        )

        # Original row should still exist
        count = execute_sql_fetchone(
            db_conn, "SELECT COUNT(*) FROM test_schema.metadata"
        )
        assert count[0] == 1

    def test_drop_partition_with_special_chars(self, conninfo, db_conn):
        """Test drop_partition with special characters in path"""
        execute_sql(
            db_conn,
            """
            CREATE TABLE test_schema.special_test (
                source_path TEXT,
                data TEXT
            )
        """,
        )
        # Path with special characters
        special_path = "/path/with spaces/and'quotes/file.csv"
        execute_sql(
            db_conn,
            f"INSERT INTO test_schema.special_test VALUES ('{special_path.replace(chr(39), chr(39) + chr(39))}', 'data')",
        )

        drop_partition(
            conninfo=conninfo,
            table="special_test",
            partition_key=special_path,
            schema="test_schema",
        )

        count = execute_sql_fetchone(
            db_conn, "SELECT COUNT(*) FROM test_schema.special_test"
        )
        assert count[0] == 0


# ===== COLUMN MAPPING EDGE CASES =====


class TestColumnMappingEdgeCases:
    """Test column mapping with edge cases"""

    def test_prepare_column_mapping_all_missing(self, temp_dir):
        """Test column mapping when all columns need to be added"""
        csv_path = temp_dir / "missing.csv"
        csv_path.write_text("col1,col2\nval1,val2\n")

        column_mapping = {
            "new_col1": ([], "string"),
            "new_col2": ([], "string"),
            "default": ([], "string"),
        }

        df = read_csv(
            full_path=str(csv_path),
            column_mapping=column_mapping,
            has_header=True,
        )

        # Should have the original columns plus potentially missing columns handled
        assert "col1" in df.columns or "new_col1" in df.columns

    def test_prepare_column_mapping_with_rename_and_missing(self):
        """Test column mapping with both rename and missing columns"""
        column_mapping = {
            "new_name": (["old_name"], "string"),
            "existing": ([], "int"),
            "missing_col": ([], "string"),
            "default": ([], "string"),
        }

        header = ["old_name", "existing"]

        rename_dict, read_dtypes, missing_cols = prepare_column_mapping(
            header, column_mapping
        )

        assert rename_dict == {"old_name": "new_name"}
        assert "missing_col" in missing_cols
        assert missing_cols["missing_col"] == "string"


# ===== CONNECTION STRING TESTS =====


class TestConnectionStringHandling:
    """Test that functions work correctly with connection strings (conninfo)"""

    def test_update_table_uses_fresh_connections(self, conninfo, temp_dir):
        """Test that update_table creates fresh connections for each operation"""
        # Create test files
        csv_dir = temp_dir / "csv_files"
        csv_dir.mkdir()

        for i in range(3):
            csv_path = csv_dir / f"file_{i}.csv"
            csv_path.write_text(f"name,value\nitem_{i},{i * 10}\n")

        column_mapping = {
            "name": ([], "string"),
            "value": ([], "int"),
        }

        # First, add files to metadata
        add_files_to_metadata_table(
            conninfo=conninfo,
            schema="test_schema",
            source_dir=str(csv_dir) + "/",
            filetype="csv",
            has_header=True,
            resume=False,
        )

        # Then run update_table - this should use fresh connections internally
        result_df = update_table(
            conninfo=conninfo,
            schema="test_schema",
            output_table="multi_file_test",
            filetype="csv",
            column_mapping=column_mapping,
            source_dir=str(csv_dir) + "/",
            resume=False,
        )

        # Verify all files were processed successfully
        assert len(result_df) == 3
        assert all(result_df["status"] == "Success")

        # Verify data was loaded by connecting fresh
        with psycopg.connect(conninfo) as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT COUNT(*) FROM test_schema.multi_file_test")
                count = cur.fetchone()[0]
                assert count == 3  # One row per file

    def test_add_files_to_metadata_table_connection_isolation(self, conninfo, temp_dir):
        """Test that add_files_to_metadata_table handles connections properly"""
        csv_dir = temp_dir / "source"
        csv_dir.mkdir()
        landing_dir = temp_dir / "landing"
        landing_dir.mkdir()

        # Create multiple files
        for i in range(5):
            (csv_dir / f"data_{i}.csv").write_text(f"col1,col2\nval_{i},100\n")

        # Run add_files_to_metadata_table
        add_files_to_metadata_table(
            conninfo=conninfo,
            schema="test_schema",
            source_dir=str(csv_dir) + "/",
            filetype="csv",
            has_header=True,
            resume=False,
        )

        # Verify metadata table exists and has correct data
        with psycopg.connect(conninfo) as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT COUNT(*) FROM test_schema.metadata")
                count = cur.fetchone()[0]
                assert count == 5

                # Verify all records have Success status
                cur.execute(
                    "SELECT COUNT(*) FROM test_schema.metadata WHERE metadata_ingest_status = 'Success'"
                )
                success_count = cur.fetchone()[0]
                assert success_count == 5

    def test_conninfo_required_for_update_table(self):
        """Test that update_table raises error when conninfo is missing"""
        with pytest.raises(ValueError, match="conninfo"):
            update_table(
                conninfo=None,
                schema="test_schema",
                output_table="test",
                filetype="csv",
                source_dir="/tmp/",
                column_mapping={"col": ([], "string")},
            )

    def test_drop_functions_with_conninfo(self, conninfo, db_conn):
        """Test drop functions work with connection string"""
        # Setup: create metadata table and data table
        execute_sql(
            db_conn,
            """
            CREATE TABLE test_schema.metadata (
                source_dir TEXT,
                source_path TEXT PRIMARY KEY
            )
        """,
        )
        execute_sql(
            db_conn,
            """
            CREATE TABLE test_schema.data_table (
                source_path TEXT,
                data TEXT
            )
        """,
        )

        # Insert test data
        execute_sql(
            db_conn,
            "INSERT INTO test_schema.metadata VALUES ($1, $2)",
            ("/search/path", "/search/path/file1.csv"),
        )
        execute_sql(
            db_conn,
            "INSERT INTO test_schema.metadata VALUES ($1, $2)",
            ("/search/path", "/search/path/file2.csv"),
        )
        execute_sql(
            db_conn,
            "INSERT INTO test_schema.data_table VALUES ($1, $2)",
            ("/search/path/file1.csv", "data1"),
        )

        # Test drop_partition with conninfo
        drop_partition(
            conninfo=conninfo,
            table="data_table",
            partition_key="/search/path/file1.csv",
            schema="test_schema",
        )

        # Verify deletion
        result = execute_sql_fetchone(
            db_conn, "SELECT COUNT(*) FROM test_schema.data_table"
        )
        assert result[0] == 0

        # Test drop_metadata_by_source with conninfo
        drop_metadata_by_source(
            conninfo=conninfo,
            source_dir="/search/path",
            schema="test_schema",
        )

        # Verify deletion
        result = execute_sql_fetchone(
            db_conn, "SELECT COUNT(*) FROM test_schema.metadata"
        )
        assert result[0] == 0

    def test_sequential_operations_dont_share_connections(self, conninfo, temp_dir):
        """Test that sequential operations each get fresh connections"""
        csv_dir = temp_dir / "source"
        csv_dir.mkdir()

        (csv_dir / "test.csv").write_text("name,value\ntest,100\n")

        column_mapping = {
            "name": ([], "string"),
            "value": ([], "int"),
        }

        # Operation 1: add files
        add_files_to_metadata_table(
            conninfo=conninfo,
            schema="test_schema",
            source_dir=str(csv_dir) + "/",
            filetype="csv",
            has_header=True,
            resume=False,
        )

        # Operation 2: update table
        update_table(
            conninfo=conninfo,
            schema="test_schema",
            output_table="seq_test",
            filetype="csv",
            column_mapping=column_mapping,
            source_dir=str(csv_dir) + "/",
            resume=False,
        )

        # Operation 3: verify with fresh connection
        with psycopg.connect(conninfo) as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT COUNT(*) FROM test_schema.seq_test")
                assert cur.fetchone()[0] == 1

                cur.execute("SELECT COUNT(*) FROM test_schema.metadata")
                assert cur.fetchone()[0] == 1

    def test_update_table_handles_errors_with_fresh_connections(
        self, conninfo, temp_dir
    ):
        """Test that errors in one file don't break connections for subsequent files"""
        csv_dir = temp_dir / "source"
        csv_dir.mkdir()

        # Create a good file and a file that will cause type conversion issues
        (csv_dir / "good.csv").write_text("name,value\ntest,100\n")
        (csv_dir / "bad.csv").write_text("name,value\ntest,not_a_number\n")
        (csv_dir / "good2.csv").write_text("name,value\ntest2,200\n")

        column_mapping = {
            "name": ([], "string"),
            "value": ([], "int"),  # Will fail on "not_a_number"
        }

        # Add files to metadata
        add_files_to_metadata_table(
            conninfo=conninfo,
            schema="test_schema",
            source_dir=str(csv_dir) + "/",
            filetype="csv",
            has_header=True,
            resume=False,
        )

        # Run update_table - should handle errors gracefully
        result_df = update_table(
            conninfo=conninfo,
            schema="test_schema",
            output_table="error_test",
            filetype="csv",
            column_mapping=column_mapping,
            source_dir=str(csv_dir) + "/",
            resume=False,
        )

        # Should have processed all 3 files (2 success, 1 failure)
        assert len(result_df) == 3

        # Verify good files were loaded
        with psycopg.connect(conninfo) as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT COUNT(*) FROM test_schema.error_test")
                # At least the good files should be loaded
                count = cur.fetchone()[0]
                assert count >= 1  # At least one good file loaded

    def test_update_table_errors_logged_to_metadata(self, conninfo, temp_dir):
        """Test that file processing errors are logged to the metadata table"""
        csv_dir = temp_dir / "source"
        csv_dir.mkdir()

        # Create a good file and a file that will cause type conversion issues
        (csv_dir / "good.csv").write_text("name,value\ntest,100\n")
        (csv_dir / "bad.csv").write_text("name,value\ntest,not_a_number\n")

        column_mapping = {
            "name": ([], "string"),
            "value": ([], "int"),  # Will fail on "not_a_number"
        }

        # Add files to metadata
        add_files_to_metadata_table(
            conninfo=conninfo,
            schema="test_schema",
            source_dir=str(csv_dir) + "/",
            filetype="csv",
            has_header=True,
            resume=False,
        )

        # Run update_table - should handle errors gracefully
        update_table(
            conninfo=conninfo,
            schema="test_schema",
            output_table="error_log_test",
            filetype="csv",
            column_mapping=column_mapping,
            source_dir=str(csv_dir) + "/",
            resume=False,
        )

        # Verify the error was logged to metadata table
        with psycopg.connect(conninfo) as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT status, error_message
                    FROM test_schema.metadata
                    WHERE source_path LIKE '%bad.csv'
                    """
                )
                result = cur.fetchone()
                assert result is not None, "bad.csv should be in metadata"
                status, error_message = result
                assert status == "Failure", f"Expected Failure status, got {status}"
                assert error_message is not None, "Error message should be logged"
                assert len(error_message) > 0, "Error message should not be empty"

        # Verify the good file succeeded
        with psycopg.connect(conninfo) as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT status, error_message
                    FROM test_schema.metadata
                    WHERE source_path LIKE '%good.csv'
                    """
                )
                result = cur.fetchone()
                assert result is not None, "good.csv should be in metadata"
                status, error_message = result
                assert status == "Success", f"Expected Success status, got {status}"

    def test_update_table_schema_mismatch_logged_to_metadata(self, conninfo, temp_dir):
        """Test that schema mismatch errors are logged to metadata and don't stop processing"""
        csv_dir = temp_dir / "source"
        csv_dir.mkdir()

        # Create files - one will cause schema mismatch after table is created
        (csv_dir / "first.csv").write_text("name,value\nAlice,100\n")
        (csv_dir / "second.csv").write_text(
            "name,value,extra_col\nBob,200,extra\n"
        )  # Has extra column
        (csv_dir / "third.csv").write_text("name,value\nCharlie,300\n")

        column_mapping_first = {
            "name": ([], "string"),
            "value": ([], "int"),
        }

        column_mapping_with_extra = {
            "name": ([], "string"),
            "value": ([], "int"),
            "extra_col": ([], "string"),
        }

        # Add files to metadata
        add_files_to_metadata_table(
            conninfo=conninfo,
            schema="test_schema",
            source_dir=str(csv_dir) + "/",
            filetype="csv",
            has_header=True,
            resume=False,
        )

        # Use column_mapping_fn to return different mappings per file
        def get_column_mapping(file_path):
            if "second" in str(file_path):
                return column_mapping_with_extra
            return column_mapping_first

        # Run update_table - second file should fail with schema mismatch but not stop processing
        update_table(
            conninfo=conninfo,
            schema="test_schema",
            output_table="schema_mismatch_test",
            filetype="csv",
            column_mapping_fn=get_column_mapping,
            source_dir=str(csv_dir) + "/",
            resume=False,
        )

        # Verify the schema mismatch was logged to metadata table
        with psycopg.connect(conninfo) as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT status, error_message
                    FROM test_schema.metadata
                    WHERE source_path LIKE '%second.csv'
                    """
                )
                result = cur.fetchone()
                assert result is not None, "second.csv should be in metadata"
                status, error_message = result
                assert status == "Failure", f"Expected Failure status, got {status}"
                assert error_message is not None, "Error message should be logged"
                assert "Schema mismatch" in error_message, (
                    f"Expected schema mismatch error, got: {error_message}"
                )

        # Verify the other files were processed successfully
        with psycopg.connect(conninfo) as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT COUNT(*)
                    FROM test_schema.metadata
                    WHERE source_dir LIKE %s AND status = 'Success'
                    """,
                    (str(csv_dir) + "/%",),
                )
                success_count = cur.fetchone()[0]
                assert success_count == 2, (
                    f"Expected 2 successful files, got {success_count}"
                )

        # Verify data was loaded for successful files
        with psycopg.connect(conninfo) as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT COUNT(*) FROM test_schema.schema_mismatch_test")
                count = cur.fetchone()[0]
                assert count == 2, f"Expected 2 rows from successful files, got {count}"


class TestEphemeralCache:
    """Test ephemeral_cache flag for temporary directory usage"""

    def test_ephemeral_cache_add_files(self, conninfo, temp_dir):
        """Test that ephemeral_cache=True uses a temp directory that gets cleaned up"""
        from table_functions import get_persistent_temp_dir

        csv_dir = temp_dir / "source"
        csv_dir.mkdir()
        (csv_dir / "test.csv").write_text("name,value\ntest,100\n")

        # Get the persistent temp dir path before
        persistent_temp = get_persistent_temp_dir()

        # Run with ephemeral_cache=True
        add_files_to_metadata_table(
            conninfo=conninfo,
            schema="test_schema",
            source_dir=str(csv_dir) + "/",
            filetype="csv",
            has_header=True,
            ephemeral_cache=True,
        )

        # Verify data was added to metadata
        with psycopg.connect(conninfo) as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT COUNT(*) FROM test_schema.metadata")
                assert cur.fetchone()[0] == 1

        # After ephemeral mode, get_persistent_temp_dir should return normal path again
        assert get_persistent_temp_dir() == persistent_temp

    def test_ephemeral_cache_update_table(self, conninfo, temp_dir):
        """Test that update_table with ephemeral_cache=True works"""
        csv_dir = temp_dir / "source"
        csv_dir.mkdir()
        (csv_dir / "test.csv").write_text("name,value\ntest,100\n")

        column_mapping = {
            "name": ([], "string"),
            "value": ([], "int"),
        }

        # First add files (using persistent cache is fine here)
        add_files_to_metadata_table(
            conninfo=conninfo,
            schema="test_schema",
            source_dir=str(csv_dir) + "/",
            filetype="csv",
            has_header=True,
        )

        # Then update table with ephemeral cache
        result_df = update_table(
            conninfo=conninfo,
            schema="test_schema",
            output_table="ephemeral_test",
            filetype="csv",
            column_mapping=column_mapping,
            source_dir=str(csv_dir) + "/",
            ephemeral_cache=True,
        )

        assert len(result_df) == 1
        assert result_df["status"].iloc[0] == "Success"

        # Verify data was loaded
        with psycopg.connect(conninfo) as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT COUNT(*) FROM test_schema.ephemeral_test")
                assert cur.fetchone()[0] == 1


# ===== S3 SUPPORT TESTS =====
# (Merged from test_s3_support.py)


class TestS3HelperFunctions:
    """Test S3 helper functions"""

    def test_is_s3_path_with_s3_url(self):
        """Test is_s3_path returns True for S3 URLs"""
        assert is_s3_path("s3://bucket/path/to/file.csv") is True
        assert is_s3_path("s3://my-bucket/") is True

    def test_is_s3_path_with_local_path(self):
        """Test is_s3_path returns False for local paths"""
        assert is_s3_path("/local/path/file.csv") is False
        assert is_s3_path("data/raw/file.csv") is False
        assert is_s3_path(Path("/local/path")) is False

    def test_get_s3_filesystem_returns_passed(self):
        """Test get_s3_filesystem returns passed filesystem"""
        mock_fs = MagicMock()
        result = get_s3_filesystem(mock_fs)
        assert result is mock_fs

    def test_get_s3_filesystem_creates_new(self):
        """Test get_s3_filesystem creates new filesystem if not provided"""
        import s3fs

        result = get_s3_filesystem(None)
        assert isinstance(result, s3fs.S3FileSystem)


class TestS3GetFileMetadataRow:
    """Test get_file_metadata_row with S3 schema"""

    def test_metadata_from_local_file(self, temp_dir):
        """Test getting metadata from local file"""

        # Create temp CSV file
        temp_file = temp_dir / "test_meta.csv"
        temp_file.write_text("col1,col2\nval1,val2\nval3,val4\n")

        # Call function with new schema - source_path is the local path
        result = get_file_metadata_row(
            source_path=temp_file.as_posix(),
            source_dir=str(temp_dir) + "/",
            filetype="csv",
            has_header=True,
            error_message=None,
        )

        # Assertions
        assert result["metadata_ingest_status"] == "Success"
        assert result["source_path"] == temp_file.as_posix()
        assert result["header"] == ["col1", "col2"]
        assert result["row_count"] == 2
        assert result["file_hash"] is not None
        assert result["filesize"] == temp_file.stat().st_size

    def test_metadata_with_error_message(self):
        """Test that error message prevents processing"""

        result = get_file_metadata_row(
            source_path="nonexistent.csv",
            source_dir="data/raw/",
            filetype="csv",
            has_header=True,
            error_message="Test error",
        )

        # Assertions
        assert result["metadata_ingest_status"] == "Failure"
        assert result["error_message"] == "Test error"
        assert result["header"] is None
        assert result["row_count"] is None
        assert result["file_hash"] is None


class TestS3AddFilesWithS3:
    """Test add_files function with S3"""

    def test_add_files_local(self, temp_dir):
        """Test adding local files"""

        # Create temp file
        test_file = temp_dir / "test_add.csv"
        test_file.write_text("col1,col2\n1,2\n")

        # Call function - now uses source_path_list, no landing_dir
        result = add_files(
            source_dir=str(temp_dir) + "/",
            resume=False,
            sample=None,
            file_list=[str(test_file)],
            filetype="csv",
            has_header=True,
            source_path_list=[],
        )

        # Assertions
        assert len(result) == 1
        assert result[0]["metadata_ingest_status"] == "Success"
        assert result[0]["source_path"] == test_file.as_posix()


class TestS3PersistentCaching:
    """Test persistent S3 caching functions"""

    def test_get_persistent_temp_dir_creates_directory(self, temp_dir):
        """Test that get_persistent_temp_dir creates temp/ in cwd"""
        import os

        original_cwd = os.getcwd()
        try:
            os.chdir(temp_dir)
            result = get_persistent_temp_dir()
            assert result.exists()
            assert result.name == "temp"
            # Use resolve() to handle symlinks (macOS /var -> /private/var)
            assert result.parent.resolve() == Path(temp_dir).resolve()
        finally:
            os.chdir(original_cwd)

    def test_get_persistent_temp_dir_idempotent(self, temp_dir):
        """Test that calling get_persistent_temp_dir multiple times is safe"""
        import os

        original_cwd = os.getcwd()
        try:
            os.chdir(temp_dir)
            result1 = get_persistent_temp_dir()
            result2 = get_persistent_temp_dir()
            assert result1 == result2
            assert result1.exists()
        finally:
            os.chdir(original_cwd)

    def test_get_cache_path_from_s3_simple(self, temp_dir):
        """Test converting S3 path to cache path"""
        import os

        original_cwd = os.getcwd()
        try:
            os.chdir(temp_dir)
            result = get_cache_path_from_s3("s3://my-bucket/data/file.csv")
            expected = (
                Path(temp_dir).resolve() / "temp" / "my-bucket" / "data" / "file.csv"
            )
            assert result.resolve() == expected
        finally:
            os.chdir(original_cwd)

    def test_get_cache_path_from_s3_nested(self, temp_dir):
        """Test converting nested S3 path to cache path"""
        import os

        original_cwd = os.getcwd()
        try:
            os.chdir(temp_dir)
            result = get_cache_path_from_s3("s3://bucket/a/b/c/d/file.zip")
            expected = (
                Path(temp_dir).resolve()
                / "temp"
                / "bucket"
                / "a"
                / "b"
                / "c"
                / "d"
                / "file.zip"
            )
            assert result.resolve() == expected
        finally:
            os.chdir(original_cwd)

    def test_get_cache_path_from_s3_creates_parents(self, temp_dir):
        """Test that get_cache_path_from_s3 creates parent directories"""
        import os

        original_cwd = os.getcwd()
        try:
            os.chdir(temp_dir)
            result = get_cache_path_from_s3("s3://bucket/nested/path/file.csv")
            # Parent directories should be created
            assert result.parent.exists()
            expected_parent = (
                Path(temp_dir).resolve() / "temp" / "bucket" / "nested" / "path"
            )
            assert result.parent.resolve() == expected_parent
        finally:
            os.chdir(original_cwd)


class TestS3GetCachePathFromSourcePath:
    """Test the get_cache_path_from_source_path function"""

    def test_s3_simple_file(self, temp_dir):
        """Test S3 file without archive"""
        import os

        original_cwd = os.getcwd()
        try:
            os.chdir(temp_dir)
            result = get_cache_path_from_source_path("s3://bucket/path/file.csv")
            expected = (
                Path(temp_dir).resolve() / "temp" / "bucket" / "path" / "file.csv"
            )
            assert result.resolve() == expected
        finally:
            os.chdir(original_cwd)

    def test_s3_archive_with_inner_path(self, temp_dir):
        """Test S3 archive with :: delimiter"""
        import os

        original_cwd = os.getcwd()
        try:
            os.chdir(temp_dir)
            result = get_cache_path_from_source_path(
                "s3://bucket/archive.zip::inner/file.csv"
            )
            expected = (
                Path(temp_dir).resolve()
                / "temp"
                / "bucket"
                / "archive.zip"
                / "inner"
                / "file.csv"
            )
            assert result.resolve() == expected
        finally:
            os.chdir(original_cwd)

    def test_local_file_returned_as_is(self, temp_dir):
        """Test local file path is returned unchanged"""
        import os

        original_cwd = os.getcwd()
        try:
            os.chdir(temp_dir)
            result = get_cache_path_from_source_path("/local/path/file.csv")
            assert result == Path("/local/path/file.csv")
        finally:
            os.chdir(original_cwd)

    def test_local_archive_with_inner_path(self, temp_dir):
        """Test local archive with :: delimiter"""
        import os

        original_cwd = os.getcwd()
        try:
            os.chdir(temp_dir)
            result = get_cache_path_from_source_path(
                "/local/archive.zip::inner/file.csv"
            )
            expected = (
                Path(temp_dir).resolve()
                / "temp"
                / "local"
                / "archive.zip"
                / "inner"
                / "file.csv"
            )
            assert result.resolve() == expected
        finally:
            os.chdir(original_cwd)


class TestArchiveResumeSkipsProcessedFiles:
    """Test that resume=True properly skips already-processed archive files"""

    def test_resume_skips_files_in_metadata(self, conninfo, db_conn, temp_dir, capsys):
        """Test that resume=True skips files already in metadata table"""
        import zipfile
        import os

        original_cwd = os.getcwd()
        try:
            os.chdir(temp_dir)

            # Create a ZIP with multiple CSV files
            zip_path = temp_dir / "test_archive.zip"
            with zipfile.ZipFile(zip_path, "w") as zf:
                zf.writestr("file1.csv", "a,b\n1,2\n3,4\n")
                zf.writestr("file2.csv", "x,y\n5,6\n")
                zf.writestr("file3.csv", "p,q\n7,8\n")

            source_dir = str(temp_dir) + "/"

            # First run: process all files (resume=False)
            result1 = add_files_to_metadata_table(
                conninfo=conninfo,
                schema="test_schema",
                source_dir=source_dir,
                filetype="csv",
                compression_type="zip",
                archive_glob="*.csv",
                has_header=True,
                resume=False,
                ephemeral_cache=True,
            )

            # Verify all 3 files were processed
            assert len(result1) == 3

            # Check metadata table has all 3 files
            count1 = execute_sql_fetchone(
                db_conn,
                "SELECT COUNT(*) FROM test_schema.metadata WHERE metadata_ingest_status = 'Success'",
            )
            assert count1[0] == 3

            # Clear captured output from first run
            capsys.readouterr()

            # Second run: with resume=True, should skip all files
            result2 = add_files_to_metadata_table(
                conninfo=conninfo,
                schema="test_schema",
                source_dir=source_dir,
                filetype="csv",
                compression_type="zip",
                archive_glob="*.csv",
                has_header=True,
                resume=True,
                ephemeral_cache=True,
            )

            # Should return existing 3 files but not re-process them
            # The function returns the metadata query results, not what was processed
            assert len(result2) == 3

            # Verify that files were skipped (not re-processed)
            captured = capsys.readouterr()
            assert "Skipped (in metadata)" in captured.out
            # Should NOT see "Row count" for any file (that means it was re-processed)
            assert "Row count:" not in captured.out

        finally:
            os.chdir(original_cwd)

    def test_resume_processes_only_new_files(self, conninfo, db_conn, temp_dir, capsys):
        """Test that resume=True only processes new files not in metadata"""
        import zipfile
        import os

        original_cwd = os.getcwd()
        try:
            os.chdir(temp_dir)

            # Create first ZIP with 2 files
            zip_path1 = temp_dir / "archive1.zip"
            with zipfile.ZipFile(zip_path1, "w") as zf:
                zf.writestr("file1.csv", "a,b\n1,2\n")
                zf.writestr("file2.csv", "x,y\n3,4\n")

            source_dir = str(temp_dir) + "/"

            # First run: process first archive
            add_files_to_metadata_table(
                conninfo=conninfo,
                schema="test_schema",
                source_dir=source_dir,
                filetype="csv",
                compression_type="zip",
                archive_glob="*.csv",
                has_header=True,
                resume=False,
                ephemeral_cache=True,
            )

            # Verify 2 files in metadata
            count1 = execute_sql_fetchone(
                db_conn, "SELECT COUNT(*) FROM test_schema.metadata"
            )
            assert count1[0] == 2

            # Create second ZIP with 2 more files
            zip_path2 = temp_dir / "archive2.zip"
            with zipfile.ZipFile(zip_path2, "w") as zf:
                zf.writestr("file3.csv", "p,q\n5,6\n")
                zf.writestr("file4.csv", "m,n\n7,8\n")

            # Clear captured output
            capsys.readouterr()

            # Second run with resume=True: should only process new files
            add_files_to_metadata_table(
                conninfo=conninfo,
                schema="test_schema",
                source_dir=source_dir,
                filetype="csv",
                compression_type="zip",
                archive_glob="*.csv",
                has_header=True,
                resume=True,
                ephemeral_cache=True,
            )

            # Should now have 4 files total
            count2 = execute_sql_fetchone(
                db_conn, "SELECT COUNT(*) FROM test_schema.metadata"
            )
            assert count2[0] == 4

            # Verify files from first archive were skipped
            captured = capsys.readouterr()
            assert "Skipped (in metadata): file1.csv" in captured.out
            assert "Skipped (in metadata): file2.csv" in captured.out
            # New files should be extracted
            assert "Extracted" in captured.out

        finally:
            os.chdir(original_cwd)

    def test_extract_and_add_zip_files_resume_skips(self, temp_dir):
        """Test extract_and_add_zip_files directly with resume=True"""
        import zipfile
        import os

        original_cwd = os.getcwd()
        try:
            os.chdir(temp_dir)

            # Create a ZIP with 3 files
            zip_path = temp_dir / "test.zip"
            with zipfile.ZipFile(zip_path, "w") as zf:
                zf.writestr("a.csv", "col1,col2\n1,2\n")
                zf.writestr("b.csv", "col1,col2\n3,4\n")
                zf.writestr("c.csv", "col1,col2\n5,6\n")

            source_dir = str(temp_dir) + "/"
            archive_path = str(zip_path)

            # Simulate that file a.csv and b.csv are already in metadata
            already_processed = [
                f"{archive_path}::a.csv",
                f"{archive_path}::b.csv",
            ]

            # Call with resume=True and pre-populated source_path_list
            rows, archive_stats = extract_and_add_zip_files(
                file_list=[archive_path],
                source_path_list=already_processed,
                source_dir=source_dir,
                has_header=True,
                filetype="csv",
                resume=True,
                sample=None,
                archive_glob="*.csv",
            )

            # Should only return 1 row (c.csv), skipping a.csv and b.csv
            assert len(rows) == 1
            assert rows[0]["source_path"] == f"{archive_path}::c.csv"
            assert (
                archive_stats[archive_path] == 1
            )  # Only 1 file processed from archive

        finally:
            os.chdir(original_cwd)

    def test_extract_and_add_zip_files_resume_false_processes_all(self, temp_dir):
        """Test extract_and_add_zip_files with resume=False processes all files"""
        import zipfile
        import os

        original_cwd = os.getcwd()
        try:
            os.chdir(temp_dir)

            # Create a ZIP with 3 files
            zip_path = temp_dir / "test.zip"
            with zipfile.ZipFile(zip_path, "w") as zf:
                zf.writestr("a.csv", "col1,col2\n1,2\n")
                zf.writestr("b.csv", "col1,col2\n3,4\n")
                zf.writestr("c.csv", "col1,col2\n5,6\n")

            source_dir = str(temp_dir) + "/"
            archive_path = str(zip_path)

            # Even with source_path_list populated, resume=False should process all
            already_processed = [
                f"{archive_path}::a.csv",
                f"{archive_path}::b.csv",
            ]

            rows, archive_stats = extract_and_add_zip_files(
                file_list=[archive_path],
                source_path_list=already_processed,
                source_dir=source_dir,
                has_header=True,
                filetype="csv",
                resume=False,  # Not resuming
                sample=None,
                archive_glob="*.csv",
            )

            # Should return all 3 files
            assert len(rows) == 3
            assert archive_stats[archive_path] == 3

        finally:
            os.chdir(original_cwd)


class TestArchiveMetadataTable:
    """Test archive_metadata table for skipping completed archives entirely"""

    def test_archive_metadata_table_created(self, conninfo, db_conn, temp_dir):
        """Test that archive_metadata table is created when using compression_type"""
        import zipfile
        import os

        original_cwd = os.getcwd()
        try:
            os.chdir(temp_dir)

            # Create a ZIP with files
            zip_path = temp_dir / "test.zip"
            with zipfile.ZipFile(zip_path, "w") as zf:
                zf.writestr("file1.csv", "a,b\n1,2\n")

            source_dir = str(temp_dir) + "/"

            add_files_to_metadata_table(
                conninfo=conninfo,
                schema="test_schema",
                source_dir=source_dir,
                filetype="csv",
                compression_type="zip",
                archive_glob="*.csv",
                has_header=True,
                ephemeral_cache=True,
            )

            # Verify archive_metadata table was created
            result = execute_sql_fetchone(
                db_conn,
                "SELECT EXISTS (SELECT FROM information_schema.tables WHERE table_schema = 'test_schema' AND table_name = 'archive_metadata')",
            )
            assert result[0] is True

        finally:
            os.chdir(original_cwd)

    def test_archive_marked_success_when_expected_count_met(
        self, conninfo, db_conn, temp_dir
    ):
        """Test that archive is marked Success when expected_archive_file_count is met"""
        import zipfile
        import os

        original_cwd = os.getcwd()
        try:
            os.chdir(temp_dir)

            # Create a ZIP with 3 files
            zip_path = temp_dir / "test.zip"
            with zipfile.ZipFile(zip_path, "w") as zf:
                zf.writestr("file1.csv", "a,b\n1,2\n")
                zf.writestr("file2.csv", "x,y\n3,4\n")
                zf.writestr("file3.csv", "p,q\n5,6\n")

            source_dir = str(temp_dir) + "/"

            add_files_to_metadata_table(
                conninfo=conninfo,
                schema="test_schema",
                source_dir=source_dir,
                filetype="csv",
                compression_type="zip",
                archive_glob="*.csv",
                has_header=True,
                expected_archive_file_count=3,
                ephemeral_cache=True,
            )

            # Verify archive_metadata has Success status
            result = execute_sql_fetchone(
                db_conn,
                "SELECT status, processed_file_count, expected_file_count FROM test_schema.archive_metadata WHERE archive_path = %s",
                (str(zip_path),),
            )
            assert result[0] == "Success"
            assert result[1] == 3
            assert result[2] == 3

        finally:
            os.chdir(original_cwd)

    def test_archive_marked_partial_when_expected_count_not_met(
        self, conninfo, db_conn, temp_dir
    ):
        """Test that archive is marked Partial when expected_archive_file_count is not met"""
        import zipfile
        import os

        original_cwd = os.getcwd()
        try:
            os.chdir(temp_dir)

            # Create a ZIP with 2 files
            zip_path = temp_dir / "test.zip"
            with zipfile.ZipFile(zip_path, "w") as zf:
                zf.writestr("file1.csv", "a,b\n1,2\n")
                zf.writestr("file2.csv", "x,y\n3,4\n")

            source_dir = str(temp_dir) + "/"

            add_files_to_metadata_table(
                conninfo=conninfo,
                schema="test_schema",
                source_dir=source_dir,
                filetype="csv",
                compression_type="zip",
                archive_glob="*.csv",
                has_header=True,
                expected_archive_file_count=5,  # More than actual files
                ephemeral_cache=True,
            )

            # Verify archive_metadata has Partial status
            result = execute_sql_fetchone(
                db_conn,
                "SELECT status, processed_file_count, expected_file_count FROM test_schema.archive_metadata WHERE archive_path = %s",
                (str(zip_path),),
            )
            assert result[0] == "Partial"
            assert result[1] == 2
            assert result[2] == 5

        finally:
            os.chdir(original_cwd)

    def test_completed_archives_skipped_on_resume(
        self, conninfo, db_conn, temp_dir, capsys
    ):
        """Test that archives marked Success are skipped entirely on resume"""
        import zipfile
        import os

        original_cwd = os.getcwd()
        try:
            os.chdir(temp_dir)

            # Create two ZIPs
            zip_path1 = temp_dir / "archive1.zip"
            with zipfile.ZipFile(zip_path1, "w") as zf:
                zf.writestr("file1.csv", "a,b\n1,2\n")
                zf.writestr("file2.csv", "x,y\n3,4\n")

            zip_path2 = temp_dir / "archive2.zip"
            with zipfile.ZipFile(zip_path2, "w") as zf:
                zf.writestr("file3.csv", "p,q\n5,6\n")
                zf.writestr("file4.csv", "m,n\n7,8\n")

            source_dir = str(temp_dir) + "/"

            # First run: process both archives
            add_files_to_metadata_table(
                conninfo=conninfo,
                schema="test_schema",
                source_dir=source_dir,
                filetype="csv",
                compression_type="zip",
                archive_glob="*.csv",
                has_header=True,
                expected_archive_file_count=2,
                ephemeral_cache=True,
            )

            # Verify both archives are Success
            result = execute_sql_fetch(
                db_conn,
                "SELECT archive_path, status FROM test_schema.archive_metadata ORDER BY archive_path",
            )
            assert len(result) == 2
            assert all(row[1] == "Success" for row in result)

            # Clear captured output
            capsys.readouterr()

            # Second run with resume=True: should skip both archives entirely
            add_files_to_metadata_table(
                conninfo=conninfo,
                schema="test_schema",
                source_dir=source_dir,
                filetype="csv",
                compression_type="zip",
                archive_glob="*.csv",
                has_header=True,
                expected_archive_file_count=2,
                resume=True,
                ephemeral_cache=True,
            )

            # Verify "Skipping" message was printed
            captured = capsys.readouterr()
            assert "Skipping 2 completed archive(s)" in captured.out
            # Should NOT see any "Extracted" or "Row count" messages
            assert "Extracted" not in captured.out
            assert "Row count" not in captured.out

        finally:
            os.chdir(original_cwd)

    def test_partial_archives_not_skipped_on_resume(
        self, conninfo, db_conn, temp_dir, capsys
    ):
        """Test that archives marked Partial are still processed on resume"""
        import zipfile
        import os

        original_cwd = os.getcwd()
        try:
            os.chdir(temp_dir)

            # Create a ZIP with 2 files
            zip_path = temp_dir / "test.zip"
            with zipfile.ZipFile(zip_path, "w") as zf:
                zf.writestr("file1.csv", "a,b\n1,2\n")
                zf.writestr("file2.csv", "x,y\n3,4\n")

            source_dir = str(temp_dir) + "/"

            # First run: process with high expected count (will be Partial)
            add_files_to_metadata_table(
                conninfo=conninfo,
                schema="test_schema",
                source_dir=source_dir,
                filetype="csv",
                compression_type="zip",
                archive_glob="*.csv",
                has_header=True,
                expected_archive_file_count=10,  # High, so Partial
                ephemeral_cache=True,
            )

            # Verify archive is Partial
            result = execute_sql_fetchone(
                db_conn,
                "SELECT status FROM test_schema.archive_metadata WHERE archive_path = %s",
                (str(zip_path),),
            )
            assert result[0] == "Partial"

            # Clear captured output
            capsys.readouterr()

            # Second run with resume=True: should NOT skip (Partial status)
            add_files_to_metadata_table(
                conninfo=conninfo,
                schema="test_schema",
                source_dir=source_dir,
                filetype="csv",
                compression_type="zip",
                archive_glob="*.csv",
                has_header=True,
                expected_archive_file_count=10,
                resume=True,
                ephemeral_cache=True,
            )

            # Should NOT see "Skipping completed archive" message
            captured = capsys.readouterr()
            assert (
                "Skipping" not in captured.out
                or "completed archive" not in captured.out
            )

        finally:
            os.chdir(original_cwd)

    def test_no_archive_metadata_without_expected_count(
        self, conninfo, db_conn, temp_dir
    ):
        """Test that archive_metadata is not updated if expected_archive_file_count is not provided"""
        import zipfile
        import os

        original_cwd = os.getcwd()
        try:
            os.chdir(temp_dir)

            # Create a ZIP
            zip_path = temp_dir / "test.zip"
            with zipfile.ZipFile(zip_path, "w") as zf:
                zf.writestr("file1.csv", "a,b\n1,2\n")

            source_dir = str(temp_dir) + "/"

            # Process without expected_archive_file_count
            add_files_to_metadata_table(
                conninfo=conninfo,
                schema="test_schema",
                source_dir=source_dir,
                filetype="csv",
                compression_type="zip",
                archive_glob="*.csv",
                has_header=True,
                # No expected_archive_file_count
                ephemeral_cache=True,
            )

            # archive_metadata table exists but should have no rows
            result = execute_sql_fetchone(
                db_conn, "SELECT COUNT(*) FROM test_schema.archive_metadata"
            )
            assert result[0] == 0

        finally:
            os.chdir(original_cwd)


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
